"""Лист «ORD Количество поручений»: заголовки (headers), агрегация (aggregation) из «СВОД_операций»; T/Y — из «REPs» и «СВОД_операций»."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

_ORD_QUANTITY_HEADER_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid",
)

ORD_QUANTITY_SHEET = "ORD Количество поручений"
ORD_QUANTITY_SHEET_LEGACY = "Количество поручений"
SVOD_SHEET = "СВОД_операций"
REPS_SHEET = "REPs"

# REPs: колонка «Операция» и «Тип сделки или иного основания» → bucket T/Y
REPS_OPERATION_FOR_TY = "перевод ценных бумаг по разделам счета депо"
REPS_DEAL_TYPE_FOR_T = "прием цб на учет"
REPS_DEAL_TYPE_FOR_Y = "внесение изменений"
REPS_REFUSAL_OP_NUM = "отказ"
# Блок «Хранение комплектов»: DF/DP — число строк СВОД, не уникальных ключей
STORAGE_SVOD_ROW_BUCKETS = frozenset({"DF", "DP"})

# «Пор.» по одному ключу на файл-источник (ГПБ / REGION / ВТБ СД)
FILENAME_POR_BUCKETS = frozenset(
    {"T", "Y", "RF", "RP", "VF", "VP", "F", "P", "XP", "XF"},
)

REGION_DK_DEPOSITORY = 'АО "ДК РЕГИОН"'


@dataclass
class _GroupLayout:
    merge: str
    start_col: int
    width: int
    subheaders: list[str]
    placeholder: bool
    instruction_types: list[str]
    metrics: list[str]
    skip_por: bool
    storage_subcolumns: dict[str, list[str]] = field(default_factory=dict)
    account_service: bool = False
    spr_su_archives: bool = False
    empty_column: bool = False
    sprav_enc: bool = False
    header_fill: str | None = None


def _project_root() -> Path:
    return Path(__file__).resolve().parent.parent


def default_sprav_workbook_path() -> Path:
    """Справочник ENC: REPORTS_DEPO_DIRECTORY/depo_validation/Справочник.xlsx."""
    from config import default_sprav_workbook_path as _from_config

    return _from_config(_project_root())


def load_headers_layout(path: Path | None = None) -> dict[str, Any]:
    """Читает memory/Ord_Quantity_headers.json."""
    p = path or (_project_root() / "memory" / "Ord_Quantity_headers.json")
    with p.open(encoding="utf-8") as f:
        return json.load(f)


def _build_group_layouts(data: dict[str, Any]) -> list[_GroupLayout]:
    col = 4
    out: list[_GroupLayout] = []
    for g in data["data_groups"]:
        subs = list(g["subheaders"])
        w = len(subs)
        if w == 1 and subs[0] == "":
            w = 1
        ph = bool(g.get("placeholder"))
        it = list(g.get("instruction_types") or [])
        met = list(g.get("metrics") or [])
        sk = bool(g.get("skip_por"))
        storage = {
            str(k): list(v)
            for k, v in (g.get("storage_subcolumns") or {}).items()
        }
        acct_svc = bool(g.get("account_service"))
        spr_su = bool(g.get("spr_su_archives"))
        empty_col = bool(g.get("empty_column"))
        sprav_enc = bool(g.get("sprav_enc"))
        header_fill = g.get("header_fill")
        out.append(
            _GroupLayout(
                merge=str(g["merge"]),
                start_col=col,
                width=w,
                subheaders=subs,
                placeholder=ph and not acct_svc and not spr_su and not empty_col and not sprav_enc,
                instruction_types=it,
                metrics=met,
                skip_por=sk,
                storage_subcolumns=storage,
                account_service=acct_svc,
                spr_su_archives=spr_su,
                empty_column=empty_col,
                sprav_enc=sprav_enc,
                header_fill=str(header_fill) if header_fill else None,
            ),
        )
        col += w
    return out


def _normalize_reps_deal_type(raw: str) -> str:
    t = (raw or "").strip().lower().replace("ё", "е")
    return re.sub(r"\s+", " ", t)


def _reps_operation_for_t_y(operation_raw: str) -> bool:
    return _normalize_reps_deal_type(operation_raw) == REPS_OPERATION_FOR_TY


def _reps_deal_type_bucket(deal_type_raw: str) -> str | None:
    n = _normalize_reps_deal_type(deal_type_raw)
    if n == REPS_DEAL_TYPE_FOR_T:
        return "T"
    if n == REPS_DEAL_TYPE_FOR_Y:
        return "Y"
    return None


def _reps_operation_number_is_refusal(op_num_raw: str) -> bool:
    """Строки REPs с «Номер операции» = ОТКАЗ не учитываются в ORD Количество поручений."""
    return _normalize_reps_deal_type(op_num_raw) == REPS_REFUSAL_OP_NUM


def _reps_row_t_y_bucket(operation_raw: str, deal_type_raw: str) -> str | None:
    if not _reps_operation_for_t_y(operation_raw):
        return None
    return _reps_deal_type_bucket(deal_type_raw)


def _normalize_source_filename(name: str) -> str:
    return (name or "").strip().casefold()


def _is_rsd_depository(depository: str) -> bool:
    d = (depository or "").strip().casefold()
    return "рсд" in d


def _normalize_depository_label(raw: str) -> str:
    t = (raw or "").replace("\u00a0", " ").replace("\u202f", " ")
    t = t.strip().casefold().replace("ё", "е")
    t = t.replace("«", '"').replace("»", '"')
    return re.sub(r"\s+", " ", t)


def _norm_sprav_account(s: str | None) -> str:
    return (s or "").strip().casefold()


def load_enc_by_account(path: Path) -> dict[str, str]:
    """Справочник.xlsx лист 1: счёт депо (col C) → ENC (col L)."""
    result: dict[str, str] = {}
    if not path.is_file():
        return result
    from sql_upload.excel_grid import cell_at, last_row_in_column, read_workbook_sheets
    from sql_upload.predsvod_layout import SPRAV1_ACCOUNT_COL, SPRAV1_ROW_START

    sheets = read_workbook_sheets(path)
    sheet_index = 0
    if sheet_index >= len(sheets):
        for i, (name, _) in enumerate(sheets):
            if "сопоставления 1" in (name or "").lower():
                sheet_index = i
                break

    if sheet_index >= len(sheets):
        return result

    _name, rows = sheets[sheet_index]
    last = last_row_in_column(rows, SPRAV1_ACCOUNT_COL)
    for r in range(SPRAV1_ROW_START, last + 1):
        account = cell_at(rows, r, SPRAV1_ACCOUNT_COL)
        if not account.strip():
            continue
        enc = cell_at(rows, r, 12)
        enc_low = enc.lower()
        if "кодификатор" in enc_low or "см." in enc_low:
            enc = ""
        result[_norm_sprav_account(account)] = enc.strip()
    return result


def _account_service_cell_value(depository: str, account: str) -> int | None:
    """
    Колонка «Обслуживание счета»: 1 для АО «ДК РЕГИОН» и счёта, начинающегося с VL; иначе пусто.
    """
    if _normalize_depository_label(depository) != _normalize_depository_label(
        REGION_DK_DEPOSITORY,
    ):
        return None
    if (account or "").strip().upper().startswith("VL"):
        return 1
    return None


def _svod_row_uses_filename_por_key(
    bucket: str,
    source_filename: str,
    depository: str,
) -> bool:
    """ГПБ / REGION / ВТБ СД: для T,Y,RF,RP,VF,VP,F,P,XP,XF — одно «Пор.» на имя файла (Источник)."""
    if bucket not in FILENAME_POR_BUCKETS:
        return False
    name = (source_filename or "").strip()
    if not name:
        return False
    low = name.lower()
    dep_cf = (depository or "").strip().casefold()
    stem = Path(name).stem.upper()

    if low.endswith(".xml"):
        return True
    if "втб" in dep_cf and "сд" in dep_cf:
        return True

    if stem.startswith("IS") and low.endswith((".xls", ".xlsx")):
        return True
    if "регион" in dep_cf and low.endswith((".xls", ".xlsx")):
        return True

    if _is_rsd_depository(depository):
        return False
    if re.match(r"^R-01", stem, re.IGNORECASE):
        return False
    if re.match(r"^I_.*\.MSG$", name, re.IGNORECASE):
        return True
    if re.match(r"^D_.*\.MSG$", name, re.IGNORECASE):
        return True
    if re.match(r"^REP[_-]", name, re.IGNORECASE) and low.endswith((".xls", ".xlsx")):
        return True
    return False


def _normalize_por_basis(raw: str) -> str:
    t = (raw or "").replace("\u00a0", " ").replace("\u202f", " ")
    t = t.strip().lower().replace("ё", "е")
    if t in ("", "-", "—", "–"):
        return ""
    return re.sub(r"\s+", " ", t)


def _por_key_for_svod_row(
    row: dict[str, Any],
    bucket: str,
    tech: str,
) -> str:
    dep = row.get("Депозитарий", "")
    if _is_rsd_depository(dep):
        return _normalize_por_basis(row.get("Основание", ""))

    source = row.get("Источник", "")
    if _svod_row_uses_filename_por_key(bucket, source, dep):
        key = _normalize_source_filename(source)
        return key or tech
    return tech


def _instruction_to_bucket(
    inst_raw: str,
    type_to_bucket: dict[str, str],
) -> str | None:
    inst = (inst_raw or "").strip().upper()
    if not inst or inst == "-":
        return None
    return type_to_bucket.get(inst)


def _parse_svod_date(s: str) -> datetime | None:
    s = (s or "").strip()
    if not s or s == "-":
        return None
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _in_month(d: datetime, year: int, month: int) -> bool:
    return d.year == year and d.month == month


def _parse_qty_mortgages(cell: str) -> int:
    s = (cell or "").strip()
    if not s or s == "-":
        return 1
    try:
        return max(1, int(float(s.replace(",", "."))))
    except ValueError:
        return 1


def _header_map(ws, row: int = 1) -> dict[str, int]:
    m: dict[str, int] = {}
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        key = str(v).strip()
        if key:
            m[key] = c
    return m


def _read_svod_rows(
    ws,
    year: int,
    month: int,
) -> list[dict[str, Any]]:
    hdr = _header_map(ws, 1)
    need = [
        "Дата операции",
        "Депозитарий",
        "Тип операции",
        "Владелец",
        "Счет депо",
        "Количество закладных",
        "Тип поручения",
        "ОдноТипОперТипРазд",
        "Основание",
        "Источник",
    ]
    for k in need:
        if k not in hdr:
            raise ValueError(f"В листе «{SVOD_SHEET}» нет колонки «{k}»")

    rows_out: list[dict[str, Any]] = []
    for r in range(2, (ws.max_row or 1) + 1):
        def cell(name: str) -> str:
            v = ws.cell(row=r, column=hdr[name]).value
            if v is None:
                return ""
            return str(v).strip()

        d_raw = cell("Дата операции")
        dt = _parse_svod_date(d_raw)
        if dt is None or not _in_month(dt, year, month):
            continue

        acc = cell("Счет депо")
        if not acc or acc == "-":
            continue

        rows_out.append(
            {
                "Дата операции": d_raw,
                "Депозитарий": cell("Депозитарий"),
                "Тип операции": cell("Тип операции"),
                "Владелец": cell("Владелец"),
                "Счет депо": acc,
                "Количество закладных": cell("Количество закладных"),
                "Тип поручения": cell("Тип поручения"),
                "ОдноТипОперТипРазд": cell("ОдноТипОперТипРазд"),
                "Основание": cell("Основание"),
                "Источник": cell("Источник"),
            },
        )
    return rows_out


def _find_reps_header_row(ws, max_scan: int = 10) -> int:
    from reps_excel import REPS_HEADERS

    marker = REPS_HEADERS[0]
    for r in range(1, max_scan + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and str(v).strip() == marker:
            return r
    return 1


def _read_reps_rows(
    ws,
    year: int,
    month: int,
    depo_rows: list | None = None,
) -> list[dict[str, Any]]:
    """Строки REPs для агрегации T/Y (ГПБ): фильтр по месяцу и типу сделки."""
    from depo_directory import lookup_depo_row_by_account_number

    header_row = _find_reps_header_row(ws)
    hdr = _header_map(ws, header_row)
    need = [
        "Номер операции",
        "Дата операции",
        "Операция",
        "Тип сделки или иного основания",
        "счет депо",
        "Tab1_Счета.владелец",
        "НГРИ",
        "Источник",
    ]
    for k in need:
        if k not in hdr:
            raise ValueError(f"В листе «{REPS_SHEET}» нет колонки «{k}»")

    rows_out: list[dict[str, Any]] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):

        def cell(name: str) -> str:
            v = ws.cell(row=r, column=hdr[name]).value
            if v is None:
                return ""
            return str(v).strip()

        if _reps_operation_number_is_refusal(cell("Номер операции")):
            continue

        deal = cell("Тип сделки или иного основания")
        operation = cell("Операция")
        bucket = _reps_row_t_y_bucket(operation, deal)
        if bucket is None:
            continue

        d_raw = cell("Дата операции")
        dt = _parse_svod_date(d_raw)
        if dt is None or not _in_month(dt, year, month):
            continue

        acc = cell("счет депо")
        if not acc or acc == "-":
            continue

        owner = cell("Tab1_Счета.владелец")
        depository = ""
        if depo_rows:
            depo_row = lookup_depo_row_by_account_number(depo_rows, acc)
            if depo_row:
                depository = (depo_row.depository or "").strip()
                if not owner:
                    owner = (depo_row.owner or "").strip()

        op_num = cell("Номер операции")
        ngri = cell("НГРИ")
        source = cell("Источник")
        tech = f"{op_num}_{d_raw}_{acc}_{ngri}".strip("_")
        por_key = _normalize_source_filename(source) or tech or f"{acc}_{bucket}_{d_raw}"

        rows_out.append(
            {
                "bucket": bucket,
                "Депозитарий": depository,
                "Владелец": owner,
                "Счет депо": acc,
                "por_key": por_key,
            },
        )
    return rows_out


@dataclass
class _BucketAcc:
    por_keys: set[str] = field(default_factory=set)
    zakl_sum: int = 0
    list_rows: int = 0
    svod_row_count: int = 0


def aggregate_reps_to_matrix(
    reps_rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
    depo_rows: list[Any] | None = None,
) -> tuple[dict[tuple[str, str, str], dict[str, _BucketAcc]], list[_GroupLayout]]:
    data = layout or load_headers_layout()
    groups = _build_group_layouts(data)

    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]] = defaultdict(
        lambda: defaultdict(_BucketAcc),
    )

    for row in reps_rows:
        dep = (row["Депозитарий"] or "").strip() or "-"
        owner = (row["Владелец"] or "").strip() or "-"
        acc = row["Счет депо"].strip()
        bucket = row["bucket"]
        cell_key = _matrix_row_key_for_account(
            depo_rows,
            acc,
            depository=dep,
            owner=owner,
        )
        accm = matrix[cell_key][bucket]
        accm.por_keys.add(row["por_key"])
        accm.zakl_sum += 1

    return matrix, groups


def _is_gpb_depository(depository: str) -> bool:
    d = (depository or "").strip().casefold()
    return "гпб" in d


def _is_vtb_sd_depository(depository: str) -> bool:
    d = (depository or "").strip().casefold()
    return "втб" in d and "сд" in d


def _svod_zakl_increment_for_bucket(
    depository: str,
    bucket: str,
    qty_mortgages: int,
) -> int:
    """ВТБ СД: для F/P/T/Y «Закл» = число строк СВОД; для остальных — сумма «Количество закладных»."""
    if _is_vtb_sd_depository(depository) and bucket in ("F", "P", "T", "Y"):
        return 1
    return qty_mortgages


def _is_region_dk_depository(depository: str) -> bool:
    return _normalize_depository_label(depository) == _normalize_depository_label(
        REGION_DK_DEPOSITORY,
    )


def apply_gpb_spr_su_archives_to_matrix(
    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]],
    spr_dir: Path,
) -> int:
    """
    Колонка S(u) для ГПБ: «Закл» и «Лист» = число архивов SPR*_<счёт>_*.zip|7z (GPB/SPR).
    """
    from gpb_spr_su_archive import scan_gpb_spr_su_archives

    stats_by_acc = scan_gpb_spr_su_archives(spr_dir)
    if not stats_by_acc:
        return 0

    filled = 0
    for cell_key in matrix:
        dep, _owner, acc = cell_key
        if not _is_gpb_depository(dep):
            continue
        acc_key = _matrix_account_lookup_key(acc)
        cnt = stats_by_acc.get(acc_key)
        if not cnt:
            continue
        accm = matrix[cell_key]["S(u)"]
        accm.zakl_sum = cnt
        accm.list_rows = cnt
        filled += 1
    return filled


def apply_region_spr_pdfs_to_matrix(
    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]],
    region_spr_dir: Path,
) -> int:
    """
    Колонка S(u) для Регион: «Закл» и «Лист» = число PDF [счёт]… внутри архивов REGION/SPR.
    """
    from region_spr_pdf import normalize_region_account_key, scan_region_spr_pdfs

    stats_by_acc = scan_region_spr_pdfs(region_spr_dir)
    if not stats_by_acc:
        return 0

    filled = 0
    for cell_key in matrix:
        dep, _owner, acc = cell_key
        if not _is_region_dk_depository(dep):
            continue
        acc_key = normalize_region_account_key(acc)
        cnt = stats_by_acc.get(acc_key)
        if not cnt:
            continue
        accm = matrix[cell_key]["S(u)"]
        accm.zakl_sum = cnt
        accm.list_rows = cnt
        filled += 1
    return filled


def apply_gpb_spr_q_pdf_to_matrix(
    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]],
    spr_dir: Path,
) -> int:
    """
    Колонка Q: «Закл» = число Q*.pdf по счёту, «Лист» = сумма (страниц − 1) по каждому PDF (GPB/SPR).
    Возвращает число строк матрицы, для которых заданы ненулевые метрики Q.
    """
    from gpb_spr_q_pdf import scan_gpb_spr_q_pdfs

    stats_by_acc = scan_gpb_spr_q_pdfs(spr_dir)
    if not stats_by_acc:
        return 0

    filled = 0
    for cell_key in matrix:
        _dep, _owner, acc = cell_key
        acc_key = _matrix_account_lookup_key(acc)
        st = stats_by_acc.get(acc_key)
        if st is None or (st.file_count == 0 and st.page_sum == 0):
            continue
        accm = matrix[cell_key]["Q"]
        accm.zakl_sum = st.file_count
        accm.list_rows = st.page_sum
        filled += 1
    return filled


def _matrix_account_lookup_key(account_display: str) -> str:
    from depo_directory import normalize_account_key

    return normalize_account_key(account_display)


def _merge_bucket_matrices(
    primary: dict[tuple[str, str, str], dict[str, _BucketAcc]],
    extra: dict[tuple[str, str, str], dict[str, _BucketAcc]],
) -> dict[tuple[str, str, str], dict[str, _BucketAcc]]:
    for cell_key, buckets in extra.items():
        for bucket, accm in buckets.items():
            dst = primary[cell_key][bucket]
            dst.por_keys |= accm.por_keys
            dst.zakl_sum += accm.zakl_sum
            dst.list_rows += accm.list_rows
            dst.svod_row_count += accm.svod_row_count
    return primary


def _type_to_bucket_map(groups: list[_GroupLayout]) -> dict[str, str]:
    m: dict[str, str] = {}
    for g in groups:
        for sub_name, types in g.storage_subcolumns.items():
            for t in types:
                m[t.strip().upper()] = sub_name.strip().upper()
        if g.placeholder:
            continue
        for t in g.instruction_types:
            m[t.strip().upper()] = g.merge
    return m


def _format_account_for_matrix(acc: str) -> str:
    """Как в СВОД: чистые цифры — с ведущими нулями (leading zeros)."""
    s = (acc or "").strip()
    if not s:
        return ""
    if re.fullmatch(r"[0-9]+", s):
        return s
    try:
        return str(int(float(s.replace(",", "."))))
    except (ValueError, TypeError):
        return s


def _matrix_row_key_from_depo_row(row: Any) -> tuple[str, str, str]:
    dep = (row.depository or "").strip() or "-"
    owner = (row.owner or "").strip() or "-"
    acc = _format_account_for_matrix(row.account_number)
    return (dep, owner, acc)


def _matrix_row_key_for_account(
    depo_rows: list | None,
    account: str,
    *,
    depository: str = "",
    owner: str = "",
) -> tuple[str, str, str]:
    from depo_directory import lookup_depo_row_by_account_number

    if depo_rows:
        depo_row = lookup_depo_row_by_account_number(depo_rows, account)
        if depo_row is not None:
            return _matrix_row_key_from_depo_row(depo_row)
    dep = (depository or "").strip() or "-"
    ow = (owner or "").strip() or "-"
    return (dep, ow, _format_account_for_matrix(account))


def seed_matrix_from_depo_directory(
    depo_rows: list[Any],
) -> dict[tuple[str, str, str], dict[str, _BucketAcc]]:
    """Базовые строки матрицы A–C: все счета из «Счета депо» (колонка «Номер счета депо»)."""
    from depo_directory import normalize_account_key

    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]] = defaultdict(
        lambda: defaultdict(_BucketAcc),
    )
    seen_accounts: set[str] = set()
    for row in depo_rows:
        acc_raw = (row.account_number or "").strip()
        if not acc_raw:
            continue
        acc_key = normalize_account_key(acc_raw)
        if not acc_key or acc_key in seen_accounts:
            continue
        seen_accounts.add(acc_key)
        cell_key = _matrix_row_key_from_depo_row(row)
        _ = matrix[cell_key]
    return matrix


def aggregate_svod_to_matrix(
    svod_rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
    depo_rows: list[Any] | None = None,
) -> tuple[dict[tuple[str, str, str], dict[str, _BucketAcc]], list[_GroupLayout]]:
    data = layout or load_headers_layout()
    groups = _build_group_layouts(data)
    tmap = _type_to_bucket_map(groups)

    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]] = defaultdict(
        lambda: defaultdict(_BucketAcc),
    )

    for row in svod_rows:
        dep = (row["Депозитарий"] or "").strip() or "-"
        owner = (row["Владелец"] or "").strip() or "-"
        acc = row["Счет депо"].strip()
        inst = row["Тип поручения"]
        tech = (row["ОдноТипОперТипРазд"] or "").strip()
        if not tech:
            tech = f"{acc}_{inst}_{row['Дата операции']}"

        bucket = _instruction_to_bucket(inst, tmap)
        if bucket is None:
            continue

        qty = _parse_qty_mortgages(row["Количество закладных"])
        cell_key = _matrix_row_key_for_account(
            depo_rows,
            acc,
            depository=dep,
            owner=owner,
        )
        accm = matrix[cell_key][bucket]
        accm.zakl_sum += _svod_zakl_increment_for_bucket(dep, bucket, qty)
        if bucket in STORAGE_SVOD_ROW_BUCKETS:
            accm.svod_row_count += 1
        elif bucket == "Q":
            continue
        else:
            accm.por_keys.add(_por_key_for_svod_row(row, bucket, tech))

    return matrix, groups


def _unmerge_all(ws) -> None:
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))


def _header_fill_for_color(color: str | None) -> PatternFill:
    c = (color or "C6EFCE").lstrip("#").upper()
    return PatternFill(start_color=c, end_color=c, fill_type="solid")


def _apply_ord_quantity_header_fill(
    ws,
    id_col_count: int,
    groups: list[_GroupLayout],
) -> None:
    """Заливка строк 1–2: зелёная по умолчанию, per-group header_fill из JSON."""
    default_fill = _ORD_QUANTITY_HEADER_FILL
    for r in range(1, 3):
        for c in range(1, id_col_count + 1):
            ws.cell(row=r, column=c).fill = default_fill
    for g in groups:
        fill = _header_fill_for_color(g.header_fill)
        for c in range(g.start_col, g.start_col + g.width):
            for r in range(1, 3):
                ws.cell(row=r, column=c).fill = fill


def ensure_ord_quantity_sheet(wb, layout: dict[str, Any] | None = None) -> Any:
    data = layout or load_headers_layout()
    name = data.get("sheet_name", ORD_QUANTITY_SHEET)
    if name not in wb.sheetnames and ORD_QUANTITY_SHEET_LEGACY in wb.sheetnames:
        wb[ORD_QUANTITY_SHEET_LEGACY].title = name
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    ws = wb[name]
    groups = _build_group_layouts(data)
    _unmerge_all(ws)

    last_col = max(3, groups[-1].start_col + groups[-1].width - 1)
    for r in range(1, 3):
        for c in range(1, last_col + 1):
            ws.cell(row=r, column=c, value=None)

    for i, idc in enumerate(data["id_columns"], start=1):
        ws.cell(row=1, column=i, value=idc["header"])
        ws.merge_cells(
            start_row=1,
            start_column=i,
            end_row=2,
            end_column=i,
        )

    for g in groups:
        c0 = g.start_col
        c1 = c0 + g.width - 1
        letter_s = get_column_letter(c0)
        letter_e = get_column_letter(c1)
        ws.merge_cells(f"{letter_s}1:{letter_e}1")
        ws.cell(row=1, column=c0, value=g.merge)
        for j, sub in enumerate(g.subheaders):
            ws.cell(row=2, column=c0 + j, value=sub if sub else None)

    _apply_ord_quantity_header_fill(ws, len(data["id_columns"]), groups)
    return ws


def _storage_v_tek_formula(row_idx: int, group: _GroupLayout) -> str:
    """«в тек» = «в пред» + DF − DP (эталон: =AD+AE−AF при текущей схеме колонок)."""
    letters: dict[str, str] = {}
    for j, sub in enumerate(group.subheaders):
        letters[sub.strip()] = get_column_letter(group.start_col + j)
    r = row_idx
    return (
        f"={letters['в пред']}{r}+{letters['DF']}{r}-{letters['DP']}{r}"
    )


def write_ord_quantity_matrix(
    path: Path,
    matrix: dict[tuple[str, str, str], dict[str, _BucketAcc]],
    groups: list[_GroupLayout],
    layout: dict[str, Any] | None = None,
    enc_by_account: dict[str, str] | None = None,
) -> None:
    data = layout or load_headers_layout()
    wb = load_workbook(path)
    ensure_ord_quantity_sheet(wb, data)
    ws = wb[data.get("sheet_name", ORD_QUANTITY_SHEET)]

    max_r = ws.max_row or 2
    if max_r > 2:
        for r in range(3, max_r + 1):
            for c in range(1, groups[-1].start_col + groups[-1].width + 1):
                ws.cell(row=r, column=c, value=None)

    sorted_keys = sorted(matrix.keys(), key=lambda x: (x[0], x[1], x[2]))
    row_idx = 3
    for dep, owner, acc in sorted_keys:
        ws.cell(row=row_idx, column=1, value=dep if dep != "-" else None)
        ws.cell(row=row_idx, column=2, value=owner if owner != "-" else None)
        ws.cell(row=row_idx, column=3, value=acc)
        buckets = matrix[(dep, owner, acc)]

        for g in groups:
            if g.account_service:
                ws.cell(
                    row=row_idx,
                    column=g.start_col,
                    value=_account_service_cell_value(dep, acc),
                )
                continue

            if g.empty_column:
                for j in range(g.width):
                    ws.cell(row=row_idx, column=g.start_col + j, value=None)
                continue

            if g.sprav_enc:
                enc_map = enc_by_account or {}
                enc_val = enc_map.get(_norm_sprav_account(acc))
                for j in range(g.width):
                    ws.cell(
                        row=row_idx,
                        column=g.start_col + j,
                        value=enc_val if enc_val else None,
                    )
                continue

            if g.storage_subcolumns:
                for j, sub in enumerate(g.subheaders):
                    sub_key = sub.strip().upper()
                    col = g.start_col + j
                    if sub.strip() == "в тек":
                        ws.cell(
                            row=row_idx,
                            column=col,
                            value=_storage_v_tek_formula(row_idx, g),
                        )
                    elif sub_key in g.storage_subcolumns:
                        accm = buckets.get(sub_key)
                        ws.cell(
                            row=row_idx,
                            column=col,
                            value=accm.svod_row_count if accm else 0,
                        )
                    else:
                        ws.cell(row=row_idx, column=col, value=0)
                continue

            if g.placeholder:
                for j in range(g.width):
                    ws.cell(row=row_idx, column=g.start_col + j, value=0)
                continue

            accm = buckets.get(g.merge)
            if accm is None:
                for j in range(g.width):
                    ws.cell(row=row_idx, column=g.start_col + j, value=0)
                continue

            if g.merge == "Q":
                ws.cell(row=row_idx, column=g.start_col, value=accm.zakl_sum)
                ws.cell(row=row_idx, column=g.start_col + 1, value=accm.list_rows)
            elif g.spr_su_archives:
                ws.cell(row=row_idx, column=g.start_col, value=accm.zakl_sum)
                ws.cell(row=row_idx, column=g.start_col + 1, value=accm.list_rows)
            else:
                ws.cell(row=row_idx, column=g.start_col, value=len(accm.por_keys))
                ws.cell(row=row_idx, column=g.start_col + 1, value=accm.zakl_sum)

        row_idx += 1

    wb.save(path)


def fill_ord_quantity_from_workbook(
    workbook_path: Path,
    year: int,
    month: int,
    layout_path: Path | None = None,
    *,
    gpb_spr_dir: Path | None = None,
    skip_gpb_spr_q: bool = False,
    skip_gpb_spr_su: bool = False,
    region_spr_dir: Path | None = None,
    skip_region_spr_su: bool = False,
    sprav_workbook: Path | None = None,
) -> int:
    from depo_directory import read_depo_directory_rows

    layout = (
        json.load(layout_path.open(encoding="utf-8"))
        if layout_path
        else load_headers_layout()
    )
    wb = load_workbook(workbook_path)
    if SVOD_SHEET not in wb.sheetnames:
        raise ValueError(f"В книге нет листа «{SVOD_SHEET}»")

    depo_rows = read_depo_directory_rows(str(workbook_path))
    matrix = seed_matrix_from_depo_directory(depo_rows)

    svod_rows = _read_svod_rows(wb[SVOD_SHEET], year, month)
    svod_matrix, groups = aggregate_svod_to_matrix(
        svod_rows,
        layout,
        depo_rows,
    )
    matrix = _merge_bucket_matrices(matrix, svod_matrix)

    if REPS_SHEET in wb.sheetnames:
        reps_rows = _read_reps_rows(wb[REPS_SHEET], year, month, depo_rows)
        reps_matrix, _ = aggregate_reps_to_matrix(
            reps_rows,
            layout,
            depo_rows,
        )
        matrix = _merge_bucket_matrices(matrix, reps_matrix)

    gpb_spr = gpb_spr_dir
    if gpb_spr is None:
        from period_paths import try_resolve_from_workbook

        pp = try_resolve_from_workbook(workbook_path, _project_root())
        gpb_spr = pp.gpb_spr if pp else _project_root() / "2026_01" / "GPB" / "SPR"
    gpb_spr = gpb_spr.resolve()
    if gpb_spr.is_dir():
        if not skip_gpb_spr_q:
            apply_gpb_spr_q_pdf_to_matrix(matrix, gpb_spr)
        if not skip_gpb_spr_su:
            apply_gpb_spr_su_archives_to_matrix(matrix, gpb_spr)
    elif gpb_spr_dir is not None:
        import sys

        print(f"Каталог GPB/SPR не найден (пропуск Q/S(u) ГПБ): {gpb_spr}", file=sys.stderr)

    region_spr = region_spr_dir
    if region_spr is None:
        from period_paths import try_resolve_from_workbook

        pp = try_resolve_from_workbook(workbook_path, _project_root())
        region_spr = pp.region_spr if pp else _project_root() / "2026_01" / "REGION" / "SPR"
    region_spr = region_spr.resolve()
    if region_spr.is_dir():
        if not skip_region_spr_su:
            apply_region_spr_pdfs_to_matrix(matrix, region_spr)
    elif region_spr_dir is not None:
        import sys

        print(
            f"Каталог REGION/SPR не найден (пропуск S(u) Регион): {region_spr}",
            file=sys.stderr,
        )

    sprav_path = (sprav_workbook or default_sprav_workbook_path()).resolve()
    if not sprav_path.is_file():
        import sys

        print(f"Справочник не найден (пропуск ENC): {sprav_path}", file=sys.stderr)
    enc_by_account = load_enc_by_account(sprav_path)

    write_ord_quantity_matrix(
        workbook_path,
        matrix,
        groups,
        layout,
        enc_by_account=enc_by_account,
    )
    return len(matrix)


def create_workbook_with_ord_quantity_template(out: Path, layout: dict[str, Any] | None = None) -> None:
    """Создаёт xlsx с листом «ORD Количество поручений» (только шапка)."""
    data = layout or load_headers_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("sheet_name", ORD_QUANTITY_SHEET)
    ensure_ord_quantity_sheet(wb, data)
    wb.save(out)


def create_workbook_svod_and_ord_templates(out: Path, layout: dict[str, Any] | None = None) -> None:
    """Книга с шапками «СВОД_операций» и «ORD Количество поручений» (для 2026_01_Ord_Quantity.xlsx)."""
    from svod_excel import SVOD_HEADERS

    data = layout or load_headers_layout()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("sheet_name", ORD_QUANTITY_SHEET)
    ensure_ord_quantity_sheet(wb, data)

    ws_svod = wb.create_sheet(SVOD_SHEET)
    for col, title in enumerate(SVOD_HEADERS, start=1):
        ws_svod.cell(row=1, column=col, value=title)

    wb.save(out)
