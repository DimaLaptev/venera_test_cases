"""Парсинг отчётов ГПБ REP_*.xls / REP_*.xlsx (якорный разбор шапки и таблицы «ЗАКЛАДНЫЕ»)."""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _sheet_fill_score(rows: list[list[str]]) -> int:
    """Оценка «полезности» листа: число непустых ячеек (cell)."""
    return sum(1 for row in rows for cell in row if (cell or "").strip())


def _calamine_cell_to_str(val: Any) -> str:
    """Значение ячейки calamine → строка для дальнейшего парсинга (как xlrd/openpyxl)."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (datetime, date, time)):
        return str(val)
    if isinstance(val, timedelta):
        return str(val)
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    return str(val).strip()


def _read_rows_calamine_best_sheet(path: Path) -> list[list[str]]:
    """
    Библиотека calamine (Rust) — читает ряд вариантов .xls/.xlsx/.xlsb, где xlrd/openpyxl дают пусто.
    """
    try:
        from python_calamine import CalamineWorkbook
    except ModuleNotFoundError:
        return []

    try:
        wb = CalamineWorkbook.from_path(path)
    except Exception:
        try:
            wb = CalamineWorkbook.from_object(path)
        except Exception:
            return []

    names = getattr(wb, "sheet_names", None) or []
    if not names:
        return []

    best_rows: list[list[str]] = []
    best_score = 0
    for i in range(len(names)):
        try:
            sh = wb.get_sheet_by_index(i)
        except Exception:
            continue
        try:
            raw = sh.to_python(skip_empty_area=False)
        except Exception:
            continue
        mat = [[_calamine_cell_to_str(c) for c in row] for row in raw]
        sc = _sheet_fill_score(mat)
        if sc > best_score:
            best_score = sc
            best_rows = mat
    return best_rows


def _read_rows_openpyxl_best_sheet(path: Path) -> list[list[str]]:
    """Выбор листа с максимальным содержимым (на первом вкладке иногда только подпись)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        best_rows: list[list[str]] = []
        best_score = 0
        for ws in wb.worksheets:
            if not callable(getattr(ws, "iter_rows", None)):
                continue
            try:
                mat: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    if row is None:
                        continue
                    mat.append([_cell_str(x) for x in row])
            except Exception:
                continue
            sc = _sheet_fill_score(mat)
            if sc > best_score:
                best_score = sc
                best_rows = mat
        return best_rows
    finally:
        wb.close()


def _read_rows_xlrd_best_sheet(path: Path) -> list[list[str]]:
    try:
        import xlrd
    except ModuleNotFoundError as e:
        raise RuntimeError(
            'Для чтения бинарного Excel .xls (OLE) нужен пакет xlrd (import). '
            'Установите: pip install "xlrd>=2.0.1,<3"'
        ) from e

    book = xlrd.open_workbook(str(path), formatting_info=False)
    best_rows: list[list[str]] = []
    best_score = 0
    for si in range(book.nsheets):
        sh = book.sheet_by_index(si)
        if sh.nrows == 0:
            continue
        mat: list[list[str]] = []
        for r in range(sh.nrows):
            row: list[str] = []
            for c in range(sh.ncols):
                cell = sh.cell(r, c)
                v = cell.value
                if v is None or v == "":
                    row.append("")
                elif isinstance(v, float) and v == int(v):
                    row.append(str(int(v)))
                else:
                    row.append(_cell_str(v))
            mat.append(row)
        sc = _sheet_fill_score(mat)
        if sc > best_score:
            best_score = sc
            best_rows = mat
    return best_rows


def _local_xml_tag(elem: ET.Element) -> str:
    if not elem.tag:
        return ""
    return elem.tag.split("}")[-1] if "}" in elem.tag else elem.tag


def _read_rows_xml_spreadsheet_2003(path: Path) -> list[list[str]]:
    """
    Microsoft XML Spreadsheet 2003 («электронная таблица XML»), иногда сохранена как *.xls.
    """
    try:
        raw = path.read_bytes()
    except OSError:
        return []
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    if raw.startswith(b"\xff\xfe"):
        text = raw[2:].decode("utf-16-le", errors="replace")
    elif raw.startswith(b"\xfe\xff"):
        text = raw[2:].decode("utf-16-be", errors="replace")
    else:
        text = raw.decode("utf-8", errors="replace")
    sniff = text.lstrip()[:2000].lower()
    if "<?xml" not in text[:2000].lower() and "worksheet" not in sniff:
        return []

    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return []

    def cell_text(cell: ET.Element) -> str:
        for ch in cell:
            if _local_xml_tag(ch) != "Data":
                continue
            parts: list[str] = []
            if ch.text:
                parts.append(ch.text)
            for sub in ch:
                parts.append(sub.text or "")
                if sub.tail:
                    parts.append(sub.tail)
            return "".join(parts).strip()
        tail = "".join(cell.itertext()).strip()
        return tail or ""

    best_rows: list[list[str]] = []
    best_score = 0
    for tbl in root.iter():
        if _local_xml_tag(tbl) != "Table":
            continue
        mat: list[list[str]] = []
        for row in tbl.iter():
            if _local_xml_tag(row) != "Row":
                continue
            line: list[str] = []
            for cell in row:
                if _local_xml_tag(cell) != "Cell":
                    continue
                line.append(cell_text(cell))
            if line:
                mat.append(line)
        sc = _sheet_fill_score(mat)
        if sc > best_score:
            best_score = sc
            best_rows = mat
    return best_rows


def read_rep_sheet_rows(path: Path) -> list[list[str]]:
    """
    Чтение активного содержимого книги Excel (Excel).

    - OOXML (по байтам «PK» → ZIP) — openpyxl, лист по наполнению; при пустом результате — **python-calamine**.
    - Бинарный .xls и прочее не-ZIP: **xlrd** и **python-calamine** (Rust); выбирается набор строк с большим числом непустых ячеек (при равенстве приоритет у calamine).

    - Далее: XML Spreadsheet 2003 (как *.xls); запасной openpyxl.

    При критической ошибке — пустой список строк (не падаем).
    """
    path = Path(path)
    try:
        head = path.open("rb").read(8)
    except OSError:
        return []

    if len(head) >= 2 and head[:2] == b"PK":
        try:
            oxml = _read_rows_openpyxl_best_sheet(path)
            if _sheet_fill_score(oxml) > 0:
                return oxml
        except Exception:
            pass
        cm = _read_rows_calamine_best_sheet(path)
        if _sheet_fill_score(cm) > 0:
            return cm
        return []

    rx: list[list[str]] = []
    try:
        rx = _read_rows_xlrd_best_sheet(path)
    except RuntimeError:
        raise
    except Exception:
        rx = []

    rc = _read_rows_calamine_best_sheet(path)
    sx = _sheet_fill_score(rx)
    sc = _sheet_fill_score(rc)
    rows = rc if sc >= sx else rx

    if _sheet_fill_score(rows) > 0:
        return rows

    xml_try = _read_rows_xml_spreadsheet_2003(path)
    if _sheet_fill_score(xml_try) > 0:
        return xml_try

    try:
        ob = _read_rows_openpyxl_best_sheet(path)
        if _sheet_fill_score(ob) > 0:
            return ob
    except Exception:
        pass
    return rows


def parse_rep_filename(stem: str) -> tuple[str, str, str] | None:
    """
    REP_<счёт>_<номер_поручения>_<суффикс> → (account, order, suffix).
    Пример: REP_10934_016699_1 → ('10934', '016699', '1').
    """
    m = re.match(
        r"^REP_([^_]+)_([^_]+)_([^_]+)$",
        stem,
        flags=re.IGNORECASE,
    )
    if not m:
        return None
    return m.group(1).strip(), m.group(2).strip(), m.group(3).strip()


def _row_join(row: list[str]) -> str:
    return " ".join(x for x in row if x).strip()


def _extract_after_label(text: str, label: str) -> str:
    """Значение после «Метка:» в той же строке (до перевода строки или конца)."""
    if not text or label not in text:
        return ""
    i = text.find(label)
    rest = text[i + len(label) :].lstrip()
    if rest.startswith(":"):
        rest = rest[1:].lstrip()
    line = rest.split("\n", 1)[0].strip()
    return line


def _find_field_in_rows(rows: list[list[str]], label: str) -> str:
    for row in rows:
        for cell in row:
            if not cell:
                continue
            if label.lower() in cell.lower():
                v = _extract_after_label(cell, label)
                if v:
                    return v
                v = _extract_after_label(cell, label.replace(":", ""))
                if v:
                    return v
    blob = "\n".join(_row_join(r) for r in rows[:120])
    for line in blob.split("\n"):
        if label.lower() in line.lower():
            return _extract_after_label(line, label)
    return ""


def _find_field_regex(rows: list[list[str]], pattern: re.Pattern[str]) -> str:
    blob = "\n".join(_row_join(r) for r in rows[:150])
    m = pattern.search(blob)
    return m.group(1).strip() if m else ""


_RE_EXEC_DATE = re.compile(
    r"Дата\s+исполнения\s+операции\s*:\s*(\d{2}\.\d{2}\.\d{4})",
    re.IGNORECASE,
)
_RE_STATE = re.compile(r"Состояние\s*:\s*([^\n]+)", re.IGNORECASE)


def _header_execution_rejected(state_line: str) -> bool:
    s = (state_line or "").upper().strip()
    if not s:
        return False
    if "ОТКАЗ В ИСПОЛНЕНИИ" in s:
        return True
    if "НЕ ИСПОЛНЕН" in s or "НЕИСПОЛНЕН" in s:
        return True
    if "ОТКАЗ" in s and "ИСПОЛНЕНО" not in s:
        return True
    return False


def _depo_from_header(rows: list[list[str]]) -> str:
    """Номер счёта депо из блока реквизитов (подпись + значение в соседней ячейке)."""
    labels = (
        "номер счета депо",
        "номер счёта депо",
    )
    for ri, row in enumerate(rows[:80]):
        for ci, cell in enumerate(row):
            cl = cell.lower()
            if not any(lbl in cl for lbl in labels):
                continue
            for cj in range(ci + 1, min(len(row), ci + 6)):
                cand = row[cj].strip()
                if not cand:
                    continue
                if re.match(r"^\d{3,8}$", cand):
                    return cand
            tail = _extract_after_label(cell, "Номер Счета депо")
            if not tail:
                tail = _extract_after_label(cell, "Номер счета депо")
            if tail:
                m = re.search(r"(\d{3,8})", tail)
                if m:
                    return m.group(1)
    return ""


def _normalize_exec_cell(s: str) -> str:
    t = (s or "").replace("\r\n", "\n").strip()
    t = re.sub(r"\s*\n\s*", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip()


@dataclass
class RepMortgageRow:
    seq: str
    op_num: str
    ngri: str
    fio: str
    line_state: str


@dataclass
class RepParsedFile:
    path: Path
    source_name: str
    depo_from_filename: str
    operation: str
    exec_date: str
    header_state: str
    deal_type: str
    depo_fallback: str
    mortgage_rows: list[RepMortgageRow] = field(default_factory=list)
    empty_sheet: bool = False
    header_rejected: bool = False


def _header_col_map_from_row(row: list[str]) -> dict[str, int] | None:
    """Одна строка — шапка таблицы закладных."""
    norm_cells: list[tuple[int, str]] = []
    for c, cell in enumerate(row):
        t = re.sub(r"\s+", " ", cell.lower().strip())
        norm_cells.append((c, t))
    col_map: dict[str, int] = {}
    for c, t in norm_cells:
        if not t:
            continue
        if ("п/п" in t or (t.startswith("№") and "п" in t)) and "номер операции" not in t:
            col_map["seq"] = c
        if "номер операции" in t:
            col_map["op"] = c
        if "номер закладн" in t or "нгри" in t or "регистрации ипотеки" in t:
            col_map["ngri"] = c
        if "фио" in t and "должник" in t:
            col_map["fio"] = c
        if (t == "состояние" or t.startswith("состояние")) and "фио" not in t:
            col_map["state"] = c
    if "op" in col_map and ("ngri" in col_map or "fio" in col_map):
        return col_map
    return None


def _find_mortgage_header_row(
    rows: list[list[str]], start_ri: int = 0
) -> tuple[int, dict[str, int]] | None:
    """Первая строка с «Номер операции» + закладная/ФИО после start_ri."""
    for ri in range(start_ri, len(rows)):
        cm = _header_col_map_from_row(rows[ri])
        if cm:
            return ri, cm
    return None


def _find_zakladnye_block_start(rows: list[list[str]]) -> int:
    for ri, row in enumerate(rows):
        blob = _row_join(row).upper()
        if "ЗАКЛАДНЫЕ" in blob and len(blob) < 40:
            return ri
    return -1


def _collect_mortgage_rows(
    rows: list[list[str]],
    hdr_ri: int,
    cmap: dict[str, int],
) -> list[RepMortgageRow]:
    out: list[RepMortgageRow] = []
    for rj in range(hdr_ri + 1, len(rows)):
        row = rows[rj]
        joined = _row_join(row)
        if "количество закладных" in joined.lower():
            break
        if not any(row):
            continue

        def get(ci: int) -> str:
            if 0 <= ci < len(row):
                return row[ci].strip()
            return ""

        cs = cmap.get("seq", -1)
        co = cmap["op"]
        cn = cmap.get("ngri", co + 1)
        cf = cmap.get("fio", cn + 1)
        cst = cmap.get("state", cf + 1)

        seq = get(cs) if cs >= 0 else ""
        op_n = get(co)
        if cs < 0 and co > 0:
            seq = get(0)
        ngri = get(cn)
        fio = get(cf)
        st = _normalize_exec_cell(get(cst)) if cst >= 0 else ""

        if not op_n and not ngri and not fio:
            continue

        out.append(
            RepMortgageRow(
                seq=seq,
                op_num=op_n,
                ngri=ngri,
                fio=fio,
                line_state=st,
            )
        )
    return out


def parse_rep_xls(path: Path | str) -> RepParsedFile:
    path = Path(path)
    stem = path.stem
    fn_parts = parse_rep_filename(stem)
    depo_fn = fn_parts[0] if fn_parts else ""

    rows = read_rep_sheet_rows(path)
    if not rows:
        return RepParsedFile(
            path=path,
            source_name=path.name,
            depo_from_filename=depo_fn,
            operation="",
            exec_date="",
            header_state="",
            deal_type="",
            depo_fallback="",
            empty_sheet=True,
        )

    operation = _find_field_in_rows(rows, "Операция:")
    if not operation:
        operation = _find_field_in_rows(rows, "Операция")

    exec_date = _find_field_regex(rows, _RE_EXEC_DATE)
    if not exec_date:
        exec_date = _find_field_in_rows(rows, "Дата исполнения операции:")

    header_state = _find_field_regex(rows, _RE_STATE)
    if not header_state:
        header_state = _find_field_in_rows(rows, "Состояние:")

    deal_type = _find_field_in_rows(rows, "Тип сделки или иного основания:")
    if not deal_type:
        deal_type = _find_field_in_rows(rows, "Тип сделки или иного основания")

    depo_fb = _depo_from_header(rows)

    rejected = _header_execution_rejected(header_state)

    mortgage_rows: list[RepMortgageRow] = []
    if not rejected:
        zr = _find_zakladnye_block_start(rows)
        start_ri = (zr + 1) if zr >= 0 else 0
        found = _find_mortgage_header_row(rows, start_ri=start_ri)
        if not found:
            found = _find_mortgage_header_row(rows, start_ri=0)
        if found:
            hdr_ri, cmap = found
            mortgage_rows = _collect_mortgage_rows(rows, hdr_ri, cmap)

    return RepParsedFile(
        path=path,
        source_name=path.name,
        depo_from_filename=depo_fn,
        operation=operation.strip(),
        exec_date=exec_date.strip(),
        header_state=_normalize_exec_cell(header_state),
        deal_type=deal_type.strip(),
        depo_fallback=depo_fb,
        mortgage_rows=mortgage_rows,
        header_rejected=rejected,
    )


def depo_account_for_output(parsed: RepParsedFile) -> str:
    if parsed.depo_from_filename:
        return parsed.depo_from_filename
    return parsed.depo_fallback
