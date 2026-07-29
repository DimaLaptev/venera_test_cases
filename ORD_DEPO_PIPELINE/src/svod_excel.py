"""Лист «СВОД_операций» в 2026_01_Ord_Quantity.xlsx."""

from __future__ import annotations

import re
from copy import copy
from pathlib import Path
from typing import Literal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from codifier_table import (
    CodifierRow,
    fallback_document_op_type,
    lookup_codifier,
    lookup_codifier_fields_by_instruction_type,
    lookup_codifier_r01_operation_type,
    lookup_codifier_region,
    lookup_codifier_rsd,
    lookup_codifier_vtb,
    lookup_codifier_vtb_sd_sheet,
    lookup_document_op_type_by_instruction,
)
from depo_directory import (
    CONTRACT_7830,
    DepoDirectoryRow,
    VTB_CONTRACT_7830_OWNER,
    VTB_CONTRACT_OTHER_OWNER,
    VTB_SD_DEPOSITORY_NAME,
    lookup_account_by_client_id,
    lookup_account_by_gpb_bank_code,
    lookup_account_by_gpb_dom_code,
    lookup_depo_row_by_account_number,
    lookup_depo_row_by_client_id,
    lookup_depo_row_vtb_account_owner_filter,
    depo_account_owners_equal,
    normalize_account_key,
)
from gpb_d_msg import DMsgDataRow, DMsgFile, instruction_type_df_dp
from gpb_i_msg import (
    IField,
    IMsgDataRow,
    IMsgFile,
    compute_depot_section,
    gpb_instruction_pair_key,
    is_gpb_i_row,
    normalize_ngri_for_match,
    parse_i_msg_file,
    passes_execution_filter,
)
from rsd_i_msg import (
    RsdIField,
    rsd_debit_credit_columns,
    rsd_exec_four_from_operation_type_label,
    rsd_execution_code,
)
from region_is_xls import RegionParseResult, RegionTableRow
from rsd_r01_xlsx import R01DataRow, R01ParseResult, parse_section_code
from vtb_f752_xml import VtbF752ParseResult, vtb_combine_section


SVOD_DASH = "-"
# Депозитарий (depository) для ВТБ СД, если «Счета депо» не нашлись по счёту в колонке F.
SVOD_DEPOSITORY_VTB_SD = 'АО "ВТБ СД"'

SVOD_HEADERS = [
    "Дата операции",
    "Депозитарий",
    "Код операции",
    "Тип операции",
    "Владелец",
    "Счет депо",
    "Раздел депо",
    "Счет списания",
    "Раздел списания",
    "Счет зачисления",
    "Раздел зачисления",
    "НГРИ",
    "Количество закладных",
    "Тип поручения",
    "Комментарий",
    "Основание",
    "ОдноТипОперТипРазд",
    "Источник",
]

ERRORS_SHEET_HEADERS = [
    "Источник",
    "Ключ выписки",
    "НГРИ",
    "Причина",
    "Наименование документа",
    "Дата операции",
]

REGION_ERROR_FILENAME_PARSE = "region_filename_parse"
REGION_ERROR_ACCOUNT_MISMATCH = "region_filename_account_mismatch"
REGION_ERROR_CODIFIER_INCOMPLETE = "region_codifier_incomplete"

VTB_ERROR_BASIS_NO_CONTRACT = "vtb_basis_no_contract"
VTB_ERROR_DEPO_NO_MATCH = "vtb_depo_no_match"
VTB_ERROR_CODIFIER_INCOMPLETE = "vtb_codifier_incomplete"

_BASIS_DEPO_CONTRACT_RE = re.compile(r"№\s*(\d+)\s*/", re.UNICODE)


def extract_vtb_basis_contract_number(basis: str) -> str | None:
    """Номер договора счёта депо из текста Basis: между «№» и «/»."""
    m = _BASIS_DEPO_CONTRACT_RE.search(basis or "")
    if not m:
        return None
    return (m.group(1) or "").strip()


def _svod_vtb_codifier_cell_missing(cell: str) -> bool:
    t = (cell or "").strip()
    return not t or t == SVOD_DASH


def vtb_error_row_f752(
    parsed: VtbF752ParseResult,
    ngri: str,
    *,
    reason: str,
    doc_detail: str,
    op_date: str,
) -> list:
    return [
        parsed.path.name,
        "",
        (ngri or "").strip(),
        reason,
        doc_detail.strip(),
        op_date,
    ]


def svod_technical_key(
    account_f: str,
    section_g: str,
    instruction_n: str,
    date_op: str,
) -> str:
    """Колонка Q: счёт_раздел_N_дата — счёт и раздел как в колонках «Счет депо» и «Раздел депо» строки."""
    return f"{account_f.strip()}_{section_g.strip()}_{instruction_n.strip()}_{date_op.strip()}"


def svod_technical_key_d(account_f: str, instruction_n: str, date_op: str) -> str:
    """Колонка Q для строк из отчёта D: F_N_дата (напр. 10934_DF_04.02.2025)."""
    return f"{account_f.strip()}_{instruction_n.strip()}_{date_op.strip()}"


def svod_technical_key_vtb(
    account_f: str,
    section_g: str,
    instruction_n: str,
    date_op: str,
    ngri: str,
) -> str:
    """Колонка Q для ВТБ СД: базовый ключ + НГРИ (NRN), чтобы строки по закладным не схлопывались."""
    base = svod_technical_key(account_f, section_g, instruction_n, date_op)
    n = (ngri or "").strip()
    if not n or n == SVOD_DASH:
        return base
    safe = re.sub(r"[^\w\-.:]+", "_", n, flags=re.UNICODE)
    return f"{base}_{safe}"


def reclassify_pf_to_xp_xf_for_same_owner(
    inst_type: str,
    op_type: str,
    op_code: str,
    debit_acc: str,
    credit_acc: str,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
) -> tuple[str, str, str]:
    """P/F → XP/XF при одном «Владелец счета депо» у счетов списания/зачисления; поля из «Коды»."""
    inst = (inst_type or "").strip().upper()
    if inst not in ("P", "F"):
        return inst_type, op_type, op_code

    h = (debit_acc or "").strip()
    j = (credit_acc or "").strip()
    if not h or h == SVOD_DASH or not j or j == SVOD_DASH:
        return inst_type, op_type, op_code
    if not depo_account_owners_equal(depo_rows, h, j):
        return inst_type, op_type, op_code

    new_inst = "XP" if inst == "P" else "XF"
    new_code, new_op_type = lookup_codifier_fields_by_instruction_type(
        codifier_rows,
        new_inst,
    )
    return (
        new_inst,
        new_op_type.strip() or op_type,
        new_code.strip() or op_code,
    )


def depository_fallback_from_source_path(path: Path) -> str:
    """
    Резерв для колонки «Депозитарий»: по сегменту пути входного файла.
    Порядок: RSD → GPB → REGION → VTBSD (без учёта регистра).
    """
    p = Path(path)
    try:
        parts_upper = {part.upper() for part in p.resolve().parts}
    except (OSError, RuntimeError):
        parts_upper = {part.upper() for part in p.parts}
    for segment, label in (
        ("RSD", 'ООО "РСД"'),
        ("GPB", "Банк ГПБ АО"),
        ("REGION", 'АО "ДК РЕГИОН"'),
        ("VTBSD", 'АО "ВТБ СД"'),
    ):
        if segment in parts_upper:
            return label
    return ""


def _format_stripped_cell_for_svod(s: str) -> str:
    """
    Строка уже strip: чистые цифры [0-9]+ — сохранить ведущие нули (leading zeros),
    иначе как раньше int(float).
    """
    if re.fullmatch(r"[0-9]+", s):
        return s
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        return s


def _format_depo_account_cell(acc: str | None) -> str:
    if not acc or not str(acc).strip():
        return ""
    s = str(acc).strip()
    return _format_stripped_cell_for_svod(s)


def _gpb_depo_flow_hint(
    f_svod: str,
    h_acc: str,
    i_sec: str,
    j_acc: str,
    k_sec: str,
) -> str | None:
    """
    Направление для «Коды»: «Счет депо» совпал с любым из пары списания (H/I) → debit,
    с любым из пары зачисления (J/K) → credit (формат Excel: в I могут быть и счёт, и раздел).
    """
    ks = normalize_account_key(f_svod)
    if not ks:
        return None
    debit_keys = {normalize_account_key(h_acc), normalize_account_key(i_sec)} - {""}
    credit_keys = {normalize_account_key(j_acc), normalize_account_key(k_sec)} - {""}
    on_debit = ks in debit_keys
    on_credit = ks in credit_keys
    if on_debit and not on_credit:
        return "debit"
    if on_credit and not on_debit:
        return "credit"
    return None


def _gpb_g_depo_section(
    f_svod: str,
    f_raw: str,
    g_raw: str,
    h_acc: str,
    i_sec: str,
    j_acc: str,
    k_sec: str,
) -> str:
    """Колонка «Раздел депо»: раздел списания или зачисления по стороне «Счет депо»; иначе §4.4 ТЗ."""
    ks = normalize_account_key(f_svod)
    if not ks:
        return compute_depot_section(f_raw, g_raw, h_acc, i_sec, j_acc, k_sec)
    debit_keys = {normalize_account_key(h_acc), normalize_account_key(i_sec)} - {""}
    credit_keys = {normalize_account_key(j_acc), normalize_account_key(k_sec)} - {""}
    on_debit = ks in debit_keys
    on_credit = ks in credit_keys
    if on_debit and not on_credit:
        return str(i_sec).strip()
    if on_credit and not on_debit:
        return str(k_sec).strip()
    return compute_depot_section(f_raw, g_raw, h_acc, i_sec, j_acc, k_sec)


def _format_section_cell(sec: str | None) -> str:
    if sec is None or not str(sec).strip():
        return ""
    s = str(sec).strip()
    return _format_stripped_cell_for_svod(s)


def _region_op_date_from_stem(stem: str) -> str:
    """IS250205… → 05.02.2025 (дата из имени файла (filename), если нет в шапке)."""
    m = re.match(r"^IS(\d{2})(\d{2})(\d{2})", stem, flags=re.IGNORECASE)
    if not m:
        return ""
    yy, mm, dd = m.groups()
    return f"{dd}.{mm}.20{yy}"


def _normalize_region_filename_account(s: str) -> str:
    """Единое сравнение токена из имени файла со счётом в выгрузке REGION."""
    if not s or not str(s).strip():
        return ""
    t = str(s).strip()
    try:
        f = float(t.replace(",", "."))
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return t.casefold()


def _region_accounts_equal_for_filename(token: str, account: str) -> bool:
    a = _normalize_region_filename_account(token)
    b = _normalize_region_filename_account(account)
    return bool(a and b and a == b)


def region_error_row_region(
    result: RegionParseResult,
    table_row: RegionTableRow,
    *,
    reason: str,
    doc_detail: str,
    op_date: str,
) -> list:
    """Одна строка листа «Ошибки» для конфликта имени REGION IS*.xls и счетов из файла."""
    return [
        result.path.name,
        "",
        (table_row.ngri or "").strip(),
        reason,
        doc_detail.strip(),
        op_date,
    ]


def _region_fgh_ijk(
    label_raw: str,
    op_type_resolved: str,
    from_acc: str,
    from_sec: str,
    to_acc: str,
    to_sec: str,
) -> tuple[str, str, str, str, str, str]:
    """F/G и H–K для отчёта REGION: перевод (transfer), зачисление (credit), списание (debit)."""
    lab = f"{label_raw} {op_type_resolved}".lower()
    fa, fs = (from_acc or "").strip(), (from_sec or "").strip()
    ta, ts = (to_acc or "").strip(), (to_sec or "").strip()
    has_from, has_to = bool(fa), bool(ta)

    if has_from and has_to:
        return ta, ts, fa, fs, ta, ts

    if "зачисление" in lab and "списание" not in lab:
        acc, sec = (ta, ts) if has_to else (fa, fs)
        if acc:
            return acc, sec, "", "", acc, sec
        return "", "", "", "", "", ""

    if "списание" in lab and "зачисление" not in lab:
        acc, sec = (fa, fs) if has_from else (ta, ts)
        if acc:
            return acc, sec, acc, sec, "", ""
        return "", "", "", "", "", ""

    ex = rsd_exec_four_from_operation_type_label(label_raw)
    f0 = fa or ta
    g0 = fs or ts
    if not f0:
        return "", "", "", "", "", ""
    h, i, j, k = rsd_debit_credit_columns(ex, f0, g0, op_type_resolved)
    return f0, g0, h, i, j, k


_DOC_OP_CREDIT = "Документы - зачисление"
_DOC_OP_DEBIT = "Документы - списание"


def select_d_msg_rows_for_svod(
    rows: list[DMsgDataRow],
    *,
    include_empty_doc_name: bool,
) -> list[DMsgDataRow]:
    """Внутри одного D*.MSG: для каждого НГРИ — первая строка с непустым элементом 3 (наименование документа); иначе см. include_empty_doc_name."""
    ngri_order: list[str] = []
    by_ngri: dict[str, list[DMsgDataRow]] = {}
    for data in rows:
        ng = normalize_ngri_for_match(data.ngri)
        if ng not in by_ngri:
            ngri_order.append(ng)
            by_ngri[ng] = []
        by_ngri[ng].append(data)

    out: list[DMsgDataRow] = []
    for ng in ngri_order:
        group = by_ngri[ng]
        picked: DMsgDataRow | None = None
        for data in group:
            if data.doc_type.strip():
                picked = data
                break
        if picked is None and include_empty_doc_name and group:
            picked = group[0]
        if picked is not None:
            out.append(picked)
    return out


def classify_gpb_i_flow_for_d_match(
    op_type: str,
    f_svod: str,
    h_acc: str,
    i_sec: str,
    j_acc: str,
    k_sec: str,
) -> Literal["debit", "credit"] | None:
    """
    Списание (debit) или зачисление (credit) для строки I*.MSG ГПБ: regex у конца строки «Тип операции»,
    затем _gpb_depo_flow_hint, затем эвристика по подстрокам.
    """
    ot = (op_type or "").strip().lower()
    # Дефис ASCII, en/em dash (тире) перед «списание»/«зачисление» в конце строки
    m = re.search(r"(?:-|–|—)\s*(списание|зачисление)\s*$", ot)
    if m:
        w = m.group(1).lower()
        if w == "списание":
            return "debit"
        if w == "зачисление":
            return "credit"
    hint = _gpb_depo_flow_hint(f_svod, h_acc, i_sec, j_acc, k_sec)
    if hint == "debit":
        return "debit"
    if hint == "credit":
        return "credit"
    has_s = "списание" in ot
    has_z = "зачисление" in ot
    if has_s and not has_z:
        return "debit"
    if has_z and not has_s:
        return "credit"
    return None


def gpb_i_row_op_type_and_accounts_for_d_match(
    data: IMsgDataRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
    client_id: str,
) -> tuple[str, str, str, str, str, str]:
    """
    Как в build_svod_row для ГПБ I: (op_type из lookup_codifier, f_svod, h_acc, i_sec, j_acc, k_sec).
    """
    f_raw = data.get(IField.F_DEPO_ACC)
    g_raw = data.get(IField.G_DEPO_SEC)
    h_acc = data.get(IField.H_DEBIT_ACC)
    i_sec = data.get(IField.I_DEBIT_SEC)
    j_acc = data.get(IField.J_CREDIT_ACC)
    k_sec = data.get(IField.K_CREDIT_SEC)

    op_code = data.get(IField.OP_CODE)
    aux = data.get(IField.AUX_CODE)

    depo_by_client = lookup_depo_row_by_client_id(depo_rows, client_id)

    acc_from_ref = lookup_account_by_gpb_dom_code(depo_rows, client_id)
    if not acc_from_ref:
        acc_from_ref = lookup_account_by_gpb_bank_code(depo_rows, client_id)
    f_svod = (
        _format_depo_account_cell(acc_from_ref)
        if acc_from_ref
        else (_format_depo_account_cell(f_raw) or (f_raw or "").strip())
    )

    flow_hint = _gpb_depo_flow_hint(f_svod, h_acc, i_sec, j_acc, k_sec)
    _gpb_g_depo_section(f_svod, f_raw, g_raw, h_acc, i_sec, j_acc, k_sec)

    _inst_type, op_type = lookup_codifier(
        codifier_rows,
        op_code,
        aux,
        f_svod,
        h_acc,
        j_acc,
        flow_hint=flow_hint,
    )
    _ = depo_by_client
    return op_type, f_svod, h_acc, i_sec, j_acc, k_sec


def build_gpb_i_d_match_index(
    gpb_i_dir: Path,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
) -> tuple[dict[tuple[str, str], Literal["debit", "credit"]], set[str]]:
    """
    Индекс (ключ выписки из имени I-файла, нормализованный НГРИ) → направление для D;
    множество ключей выписки по всем разобранным I*.MSG в каталоге.
    """
    index: dict[tuple[str, str], Literal["debit", "credit"]] = {}
    stmt_keys: set[str] = set()
    if not gpb_i_dir.is_dir():
        return index, stmt_keys

    seen: set[Path] = set()
    files_i: list[Path] = []
    for pat in ("I*.MSG", "i*.MSG"):
        for f in sorted(gpb_i_dir.glob(pat)):
            rp = f.resolve()
            if rp not in seen:
                seen.add(rp)
                files_i.append(f)

    for fp in files_i:
        stmt = gpb_instruction_pair_key(fp)
        if stmt is None:
            continue
        stmt_keys.add(stmt)
        try:
            parsed = parse_i_msg_file(fp)
        except OSError:
            continue
        client_id = (parsed.header.get("ClientID") or "").strip()
        for row in parsed.rows:
            parts = row.parts
            if not is_gpb_i_row(parts):
                continue
            if not passes_execution_filter(row):
                continue
            nrn = row.get(IField.NRN)
            ng_key = normalize_ngri_for_match(nrn)
            pair_key = (stmt, ng_key)
            if pair_key in index:
                continue
            op_type, f_svod, h_acc, i_sec, j_acc, k_sec = gpb_i_row_op_type_and_accounts_for_d_match(
                row,
                depo_rows,
                codifier_rows,
                client_id,
            )
            flow = classify_gpb_i_flow_for_d_match(
                op_type,
                f_svod,
                h_acc,
                i_sec,
                j_acc,
                k_sec,
            )
            if flow is None:
                continue
            index[pair_key] = flow
    return index, stmt_keys


def build_svod_row_from_d(
    parsed: DMsgFile,
    data: DMsgDataRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
    client_id: str,
    *,
    i_match_flow: Literal["debit", "credit"] | None = None,
) -> list:
    """Одна строка «СВОД_операций» из отчёта DOC_INSTRUCT (D*.MSG), эталон 2025_02."""
    h = parsed.header
    op_date = (h.get("ReportDate") or "").strip()

    acc_raw = lookup_account_by_client_id(depo_rows, client_id)
    f_str = _format_depo_account_cell(acc_raw)
    if not f_str:
        f_str = SVOD_DASH

    depo_by_client = lookup_depo_row_by_client_id(depo_rows, client_id)
    depository = (depo_by_client.depository if depo_by_client else "").strip()
    if not depository:
        depository = depository_fallback_from_source_path(parsed.path).strip()
    if not depository:
        depository = SVOD_DASH

    acc_for_inst = f_str if f_str != SVOD_DASH else ""
    if i_match_flow == "debit":
        inst_type = "DP"
        op_type = _DOC_OP_DEBIT
    elif i_match_flow == "credit":
        inst_type = "DF"
        op_type = _DOC_OP_CREDIT
    else:
        inst_type = instruction_type_df_dp(acc_for_inst)

        op_type = lookup_document_op_type_by_instruction(codifier_rows, inst_type)
        if not op_type:
            op_type = fallback_document_op_type(inst_type) or SVOD_DASH

    if f_str != SVOD_DASH:
        owner_row = lookup_depo_row_by_account_number(depo_rows, f_str)
        owner = (owner_row.owner if owner_row else (depo_by_client.owner if depo_by_client else "")).strip()
    else:
        owner = (depo_by_client.owner if depo_by_client else "").strip()
    if not owner:
        owner = SVOD_DASH

    nrn = data.ngri.strip()
    tech = svod_technical_key_d(f_str, inst_type, op_date)

    return [
        op_date or SVOD_DASH,
        depository,
        SVOD_DASH,
        op_type,
        owner,
        f_str,
        SVOD_DASH,
        SVOD_DASH,
        SVOD_DASH,
        SVOD_DASH,
        SVOD_DASH,
        nrn or SVOD_DASH,
        "1",
        inst_type,
        SVOD_DASH,
        SVOD_DASH,
        tech,
        parsed.path.name,
    ]


def build_svod_row(
    parsed: IMsgFile,
    data: IMsgDataRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
    client_id: str,
) -> list:
    h = parsed.header
    op_date = (h.get("ReportDate") or "").strip()

    f_raw = data.get(IField.F_DEPO_ACC)
    g_raw = data.get(IField.G_DEPO_SEC)
    h_acc = data.get(IField.H_DEBIT_ACC)
    i_sec = data.get(IField.I_DEBIT_SEC)
    j_acc = data.get(IField.J_CREDIT_ACC)
    k_sec = data.get(IField.K_CREDIT_SEC)

    op_code = data.get(IField.OP_CODE)
    aux = data.get(IField.AUX_CODE)

    depo_by_client = lookup_depo_row_by_client_id(depo_rows, client_id)
    depository = depo_by_client.depository if depo_by_client else ""

    acc_from_ref = lookup_account_by_gpb_dom_code(depo_rows, client_id)
    if not acc_from_ref:
        acc_from_ref = lookup_account_by_gpb_bank_code(depo_rows, client_id)
    f_svod = (
        _format_depo_account_cell(acc_from_ref)
        if acc_from_ref
        else (_format_depo_account_cell(f_raw) or (f_raw or "").strip())
    )

    flow_hint = _gpb_depo_flow_hint(f_svod, h_acc, i_sec, j_acc, k_sec)
    g_out = _gpb_g_depo_section(f_svod, f_raw, g_raw, h_acc, i_sec, j_acc, k_sec)

    inst_type, op_type = lookup_codifier(
        codifier_rows,
        op_code,
        aux,
        f_svod,
        h_acc,
        j_acc,
        flow_hint=flow_hint,
    )

    inst_type, op_type, op_code = reclassify_pf_to_xp_xf_for_same_owner(
        inst_type,
        op_type,
        op_code,
        h_acc,
        j_acc,
        depo_rows,
        codifier_rows,
    )

    owner_row = lookup_depo_row_by_account_number(depo_rows, f_svod or f_raw)
    owner = owner_row.owner if owner_row else (depo_by_client.owner if depo_by_client else "")

    nrn = data.get(IField.NRN)
    qty = data.get(IField.QTY) or "1"
    basis = data.get(IField.BASIS)

    tech = svod_technical_key(f_svod, g_out, inst_type, op_date)

    h_out = _format_depo_account_cell(h_acc) or (h_acc or "").strip()
    i_out = _format_section_cell(i_sec) or (i_sec or "").strip()
    j_out = _format_depo_account_cell(j_acc) or (j_acc or "").strip()
    k_out = _format_section_cell(k_sec) or (k_sec or "").strip()

    return [
        op_date,
        depository,
        op_code,
        op_type,
        owner,
        f_svod,
        g_out,
        h_out,
        i_out,
        j_out,
        k_out,
        nrn,
        qty,
        inst_type,
        "",
        basis,
        tech,
        parsed.path.name,
    ]


def build_svod_row_rsd(
    parsed: IMsgFile,
    data: IMsgDataRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
    client_id: str,
) -> list:
    """Одна строка «СВОД_операций» из отчёта РСД I*.MSG (формат RSD_MSG_example)."""
    h = parsed.header
    op_date = (h.get("ReportDate") or "").strip()
    parts = data.parts

    f_raw = _format_depo_account_cell(parts[RsdIField.ACCOUNT_F]) or parts[
        RsdIField.ACCOUNT_F
    ].strip()
    g_raw = _format_section_cell(parts[RsdIField.SECTION_G]) or parts[
        RsdIField.SECTION_G
    ].strip()

    op_code = parts[RsdIField.OP_CODE].strip()
    desc = parts[RsdIField.DESCRIPTION].strip() if len(parts) > RsdIField.DESCRIPTION else ""
    exec_c = rsd_execution_code(parts)

    inst_type, op_type = lookup_codifier_rsd(
        codifier_rows,
        op_code,
        exec_c,
        desc,
    )

    h_acc, i_sec, j_acc, k_sec = rsd_debit_credit_columns(
        exec_c,
        f_raw,
        g_raw,
        op_type,
    )

    inst_type, op_type, op_code = reclassify_pf_to_xp_xf_for_same_owner(
        inst_type,
        op_type,
        op_code,
        h_acc,
        j_acc,
        depo_rows,
        codifier_rows,
    )

    depo_by_client = lookup_depo_row_by_client_id(depo_rows, client_id)
    depository = (depo_by_client.depository if depo_by_client else "").strip()
    if not depository:
        depository = depository_fallback_from_source_path(parsed.path).strip()
    owner_row = lookup_depo_row_by_account_number(depo_rows, f_raw)
    owner = owner_row.owner if owner_row else (depo_by_client.owner if depo_by_client else "")

    nrn = parts[RsdIField.NRN].strip() if len(parts) > RsdIField.NRN else ""
    # РСД: колонка «Количество закладных» всегда 1 (не поле QTY строки I*.MSG).
    qty = "1"

    tech = svod_technical_key(f_raw, g_raw, inst_type, op_date)

    return [
        op_date,
        depository,
        op_code,
        op_type,
        owner,
        f_raw,
        g_raw,
        h_acc,
        i_sec,
        j_acc,
        k_sec,
        nrn,
        qty,
        inst_type,
        "",
        desc,
        tech,
        parsed.path.name,
    ]


def build_svod_row_r01(
    result: R01ParseResult,
    data: R01DataRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
) -> list | None:
    """Одна строка «СВОД_операций» из отчёта РСД R-01 *.xlsx."""
    inst_type, op_type, op_code = lookup_codifier_r01_operation_type(
        codifier_rows,
        data.operation_type,
    )

    f_raw = _format_depo_account_cell(result.depo_account) or result.depo_account.strip()
    g_raw = parse_section_code(data.section_raw)

    exec_c = rsd_exec_four_from_operation_type_label(data.operation_type)
    h_acc, i_sec, j_acc, k_sec = rsd_debit_credit_columns(
        exec_c,
        f_raw,
        g_raw,
        op_type,
    )

    inst_type, op_type, op_code = reclassify_pf_to_xp_xf_for_same_owner(
        inst_type,
        op_type,
        op_code,
        h_acc,
        j_acc,
        depo_rows,
        codifier_rows,
    )

    owner_row = lookup_depo_row_by_account_number(depo_rows, f_raw)
    depository = (owner_row.depository if owner_row else "").strip()
    if not depository:
        depository = depository_fallback_from_source_path(result.path).strip()
    owner = owner_row.owner if owner_row else ""

    op_date = data.exec_date_display.strip()
    nrn = data.reg_number.strip()
    qty = data.qty.strip() or "1"
    basis = data.basis.strip()

    op_code_cell = op_code if op_code else SVOD_DASH

    tech = svod_technical_key(f_raw, g_raw, inst_type, op_date)

    return [
        op_date,
        depository,
        op_code_cell,
        op_type,
        owner,
        f_raw,
        g_raw,
        h_acc,
        i_sec,
        j_acc,
        k_sec,
        nrn,
        qty,
        inst_type,
        "",
        basis,
        tech,
        result.path.name,
    ]


def build_svod_row_vtb(
    parsed: VtbF752ParseResult,
    ngri: str,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
    vtb_sd_codifier_rows: list | None = None,
) -> tuple[list | None, list | None]:
    """СВОД из ВТБ СД F752 + НГРИ приложения.

    Колонки F, G, E: номер договора из поля Basis (между «№» и «/»), затем поиск в «Счета депо»
    по DstAccountCode / SrcAccountCode при депозитарии АО «ВТБ СД» и владельце, зависящем от
    договора (=7830 → ПАО ДОМ РФ, иначе ООО «ДОМ.РФ ИА»). При провале — строка «Ошибки».
    Колонка B: АО «ВТБ СД». H–K: как в XML.
    """
    op_date = (
        parsed.stage_overlay_date.strip()
        or parsed.execution_date.strip()
        or parsed.order_date.strip()
    )
    if not op_date:
        op_date = SVOD_DASH

    contract = extract_vtb_basis_contract_number(parsed.basis)
    if not contract:
        detail = (parsed.basis or "").strip() or "(Basis пуст)"
        return None, vtb_error_row_f752(
            parsed,
            ngri,
            reason=VTB_ERROR_BASIS_NO_CONTRACT,
            doc_detail=f"нет номера договора №…/ в Basis: {detail[:500]}",
            op_date=op_date,
        )

    if normalize_account_key(contract) == normalize_account_key(CONTRACT_7830):
        expected_owner = VTB_CONTRACT_7830_OWNER
    else:
        expected_owner = VTB_CONTRACT_OTHER_OWNER

    hit_dst = lookup_depo_row_vtb_account_owner_filter(
        depo_rows,
        parsed.dst_account,
        expected_depository=VTB_SD_DEPOSITORY_NAME,
        expected_owner=expected_owner,
    )
    hit_src = lookup_depo_row_vtb_account_owner_filter(
        depo_rows,
        parsed.src_account,
        expected_depository=VTB_SD_DEPOSITORY_NAME,
        expected_owner=expected_owner,
    )

    src_sec = vtb_combine_section(parsed.src_section_name, parsed.src_sp_section_code)
    dst_sec = vtb_combine_section(parsed.dst_section_name, parsed.dst_sp_section_code)

    if hit_dst is not None:
        side_acc = parsed.dst_account
        side_sec = dst_sec
    elif hit_src is not None:
        side_acc = parsed.src_account
        side_sec = src_sec
    else:
        detail = (
            f"договор={contract}; filter_owner={expected_owner}; "
            f"DstAccountCode={parsed.dst_account!r}; SrcAccountCode={parsed.src_account!r}"
        )
        return None, vtb_error_row_f752(
            parsed,
            ngri,
            reason=VTB_ERROR_DEPO_NO_MATCH,
            doc_detail=detail,
            op_date=op_date,
        )

    f_raw = _format_depo_account_cell(side_acc) or (side_acc or "").strip() or SVOD_DASH
    g_raw = _format_section_cell(side_sec) or side_sec or SVOD_DASH
    if f_raw == SVOD_DASH:
        return None, vtb_error_row_f752(
            parsed,
            ngri,
            reason=VTB_ERROR_DEPO_NO_MATCH,
            doc_detail="пустой счёт после сопоставления со справочником",
            op_date=op_date,
        )

    depository = SVOD_DEPOSITORY_VTB_SD
    owner = expected_owner

    h_acc = _format_depo_account_cell(parsed.src_account) or (parsed.src_account or "").strip()
    i_sec_cell = _format_section_cell(src_sec) or src_sec
    j_acc = _format_depo_account_cell(parsed.dst_account) or (parsed.dst_account or "").strip()
    k_sec = _format_section_cell(dst_sec) or dst_sec

    src_blob = f"{parsed.src_section_name} {parsed.src_sp_section_code}".strip()
    dst_blob = f"{parsed.dst_section_name} {parsed.dst_sp_section_code}".strip()

    vtb_sd = vtb_sd_codifier_rows or []
    inst_type, op_type, op_code = "", "", ""
    if vtb_sd:
        inst_type, op_type, op_code = lookup_codifier_vtb_sd_sheet(
            vtb_sd,
            parsed.stage_name,
            f_raw,
            h_acc,
            j_acc,
        )
    if not (op_type or "").strip() and not (inst_type or "").strip() and not (op_code or "").strip():
        inst_type, op_type, op_code = lookup_codifier_vtb(
            codifier_rows,
            parsed.stage_name,
            src_sections_blob=src_blob,
            dst_sections_blob=dst_blob,
        )

    inst_type, op_type, op_code = reclassify_pf_to_xp_xf_for_same_owner(
        inst_type,
        op_type,
        op_code,
        h_acc,
        j_acc,
        depo_rows,
        codifier_rows,
    )

    op_code_cell = op_code if op_code else SVOD_DASH
    if not (op_type or "").strip():
        op_type = SVOD_DASH
    inst_cell = inst_type.strip() if (inst_type or "").strip() else SVOD_DASH

    basis_parts: list[str] = []
    if parsed.order_number.strip():
        basis_parts.append(f"№{parsed.order_number.strip()}")
    if parsed.basis.strip():
        basis_parts.append(parsed.basis.strip())
    if parsed.dop_info.strip():
        basis_parts.append(parsed.dop_info.strip())
    basis = " ".join(basis_parts) if basis_parts else SVOD_DASH

    nrn_cell = (ngri or "").strip() or SVOD_DASH
    qty_cell = (parsed.qty or "1").strip() or "1"

    tech = svod_technical_key_vtb(f_raw, g_raw, inst_cell, op_date, nrn_cell)

    comment = (parsed.stage_name or "").strip() or SVOD_DASH

    row = [
        op_date,
        depository,
        op_code_cell,
        op_type,
        owner,
        f_raw,
        g_raw,
        h_acc or SVOD_DASH,
        i_sec_cell or SVOD_DASH,
        j_acc or SVOD_DASH,
        k_sec or SVOD_DASH,
        nrn_cell,
        qty_cell,
        inst_cell,
        comment,
        basis,
        tech,
        parsed.path.name,
    ]
    if (
        _svod_vtb_codifier_cell_missing(op_code_cell)
        or _svod_vtb_codifier_cell_missing(op_type)
        or _svod_vtb_codifier_cell_missing(inst_cell)
    ):
        detail = (
            "пустые или «-» в СВОД: «Код операции» / «Тип операции» / «Тип поручения» "
            f"после листа ВТБ СД и «Коды»; StageName={comment!r}; "
            f"код={op_code_cell!r}; тип_оп={op_type!r}; тип_пор={inst_cell!r}"
        )
        detail = detail[:800]
        return row, vtb_error_row_f752(
            parsed,
            ngri,
            reason=VTB_ERROR_CODIFIER_INCOMPLETE,
            doc_detail=detail,
            op_date=op_date,
        )
    return row, None


def build_svod_row_region(
    result: RegionParseResult,
    table_row: RegionTableRow,
    depo_rows: list[DepoDirectoryRow],
    codifier_rows: list[CodifierRow],
) -> tuple[list | None, list | None]:
    """Одна строка «СВОД_операций» из REGION IS*.xls или строка «Ошибки».

    Возвращает (svod_row | None, error_row | None): при несовпадении токена счёта
    в имени файла со счётами в теле — только error_row.
    """
    stem = result.path.stem
    op_date = (result.op_date or "").strip() or _region_op_date_from_stem(stem)
    cod_label = (result.filename_operation_title or "").strip() or (
        result.operation_type or ""
    ).strip()

    inst_type, op_type, op_code = lookup_codifier_region(codifier_rows, cod_label)
    op_type_disp = op_type.strip() if op_type.strip() else SVOD_DASH
    op_code_cell = op_code.strip() if op_code.strip() else SVOD_DASH
    op_resolved = op_type.strip()

    codifier_incomplete = (
        not (op_code or "").strip()
        or not (op_type or "").strip()
        or not (inst_type or "").strip()
    )

    fa, fs = (result.from_acc or "").strip(), (result.from_sec or "").strip()
    ta, ts = (result.to_acc or "").strip(), (result.to_sec or "").strip()
    token = (result.filename_account or "").strip()
    has_from = bool(fa)
    has_to = bool(ta)
    filename_ok = stem.count("_") >= 3

    def err(reason: str, detail: str) -> tuple[list | None, list | None]:
        return None, region_error_row_region(
            result,
            table_row,
            reason=reason,
            doc_detail=detail,
            op_date=op_date,
        )

    body_has_account = has_from or has_to
    if body_has_account and not filename_ok:
        return err(
            REGION_ERROR_FILENAME_PARSE,
            "ожидается имя с тремя подчёркиваниями: ..._<счёт>_<операция>",
        )

    if has_from and has_to:
        m_to = _region_accounts_equal_for_filename(token, ta)
        m_from = _region_accounts_equal_for_filename(token, fa)
        if not m_to and not m_from:
            return err(
                REGION_ERROR_ACCOUNT_MISMATCH,
                f"token={token}; from_acc={fa}; to_acc={ta}",
            )
        h_acc, i_sec, j_acc, k_sec = fa, fs, ta, ts
        if m_to:
            f_raw, g_raw = ta, ts
        else:
            f_raw, g_raw = fa, fs
    elif has_to and not has_from:
        if not _region_accounts_equal_for_filename(token, ta):
            return err(
                REGION_ERROR_ACCOUNT_MISMATCH,
                f"token={token}; to_acc={ta}",
            )
        f_raw, g_raw, h_acc, i_sec, j_acc, k_sec = _region_fgh_ijk(
            cod_label,
            op_resolved,
            "",
            "",
            ta,
            ts,
        )
    elif has_from and not has_to:
        if not _region_accounts_equal_for_filename(token, fa):
            return err(
                REGION_ERROR_ACCOUNT_MISMATCH,
                f"token={token}; from_acc={fa}",
            )
        f_raw, g_raw, h_acc, i_sec, j_acc, k_sec = _region_fgh_ijk(
            cod_label,
            op_resolved,
            fa,
            fs,
            "",
            "",
        )
    else:
        if token:
            return err(
                REGION_ERROR_ACCOUNT_MISMATCH,
                "в файле нет счетов списания/зачисления для сверки с token",
            )
        f_raw, g_raw, h_acc, i_sec, j_acc, k_sec = _region_fgh_ijk(
            cod_label,
            op_resolved,
            "",
            "",
            "",
            "",
        )

    f_raw = _format_depo_account_cell(f_raw) or f_raw
    g_raw = _format_section_cell(g_raw) or g_raw
    h_acc = _format_depo_account_cell(h_acc) or h_acc
    i_sec = _format_section_cell(i_sec) or i_sec
    j_acc = _format_depo_account_cell(j_acc) or j_acc
    k_sec = _format_section_cell(k_sec) or k_sec

    inst_type, op_type, op_code = reclassify_pf_to_xp_xf_for_same_owner(
        inst_type,
        op_type,
        op_code,
        h_acc,
        j_acc,
        depo_rows,
        codifier_rows,
    )
    op_type_disp = op_type.strip() if op_type.strip() else SVOD_DASH
    op_code_cell = op_code.strip() if op_code.strip() else SVOD_DASH

    owner_row = lookup_depo_row_by_account_number(depo_rows, f_raw)
    depository = (owner_row.depository if owner_row else "").strip()
    if not depository:
        depository = depository_fallback_from_source_path(result.path).strip()
    owner = (owner_row.owner if owner_row else "").strip()

    nrn = (table_row.ngri or "").strip() or SVOD_DASH
    basis = (result.reg_footer or result.basis or "").strip()
    inst_cell = inst_type.strip() or SVOD_DASH

    tech = svod_technical_key(f_raw, g_raw, inst_cell, op_date)

    svod_row = [
        op_date,
        depository,
        op_code_cell,
        op_type_disp,
        owner,
        f_raw,
        g_raw,
        h_acc,
        i_sec,
        j_acc,
        k_sec,
        nrn,
        "1",
        inst_cell,
        "",
        basis,
        tech,
        result.path.name,
    ]
    if codifier_incomplete:
        detail = (
            f"метка={cod_label!r}; код={op_code_cell!r}; "
            f"тип_оп={op_type_disp!r}; тип_пор={inst_cell!r}"
        )
        return svod_row, region_error_row_region(
            result,
            table_row,
            reason=REGION_ERROR_CODIFIER_INCOMPLETE,
            doc_detail=detail[:800],
            op_date=op_date,
        )

    return (svod_row, None)


def ensure_svod_sheet(wb) -> None:
    if "СВОД_операций" not in wb.sheetnames:
        wb.create_sheet("СВОД_операций")
    ws = wb["СВОД_операций"]
    for col, title in enumerate(SVOD_HEADERS, start=1):
        ws.cell(row=1, column=col, value=title)


def ensure_errors_sheet(wb) -> None:
    if "Ошибки" not in wb.sheetnames:
        wb.create_sheet("Ошибки")
    ws = wb["Ошибки"]
    for col, title in enumerate(ERRORS_SHEET_HEADERS, start=1):
        ws.cell(row=1, column=col, value=title)


def clear_svod_data(ws: Worksheet) -> None:
    max_r = ws.max_row or 1
    if max_r <= 1:
        return
    empty_fill = PatternFill(fill_type=None)
    for r in range(2, max_r + 1):
        for c in range(1, len(SVOD_HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = empty_fill


def clear_errors_data(ws: Worksheet) -> None:
    max_r = ws.max_row or 1
    if max_r <= 1:
        return
    for r in range(2, max_r + 1):
        for c in range(1, len(ERRORS_SHEET_HEADERS) + 1):
            ws.cell(row=r, column=c, value=None)


def _header_cell_has_fill(cell) -> bool:
    fill = cell.fill
    if fill is None:
        return False
    ft = getattr(fill, "fill_type", None)
    return bool(ft) and ft != "none"


def apply_svod_column_fills_from_header(
    ws: Worksheet,
    *,
    first_row: int = 2,
    last_row: int,
) -> None:
    """Копирует заливку (fill) шапки (строка 1) на строки данных по тем же колонкам.

    В шаблоне Ord_Quantity колонки F/G/L/N и др. часто закрашены вручную; openpyxl при
    записи ``value`` не наследует заливку столбца Excel — без этого данные остаются белыми.
    """
    if last_row < first_row:
        return
    col_fills: list[PatternFill | None] = []
    for c in range(1, len(SVOD_HEADERS) + 1):
        header = ws.cell(row=1, column=c)
        if _header_cell_has_fill(header):
            col_fills.append(copy(header.fill))
        else:
            col_fills.append(None)
    for r in range(first_row, last_row + 1):
        for c, fill in enumerate(col_fills, start=1):
            if fill is not None:
                ws.cell(row=r, column=c).fill = copy(fill)


def write_svod_rows(
    path: Path,
    rows_out: list[list],
    error_rows: list[list] | None = None,
) -> None:
    wb = load_workbook(path)
    ensure_svod_sheet(wb)
    ws = wb["СВОД_операций"]
    clear_svod_data(ws)
    for i, row in enumerate(rows_out, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    if rows_out:
        apply_svod_column_fills_from_header(ws, first_row=2, last_row=1 + len(rows_out))
    ensure_errors_sheet(wb)
    ws_err = wb["Ошибки"]
    clear_errors_data(ws_err)
    if error_rows:
        for i, row in enumerate(error_rows, start=2):
            for j, val in enumerate(row, start=1):
                ws_err.cell(row=i, column=j, value=val)
    wb.save(path)


def create_workbook_with_svod_only(path: Path) -> None:
    """Минимальная книга с заголовками СВОД и «Ошибки» (если файла нет)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "СВОД_операций"
    for col, title in enumerate(SVOD_HEADERS, start=1):
        ws.cell(row=1, column=col, value=title)
    ws_err = wb.create_sheet("Ошибки")
    for col, title in enumerate(ERRORS_SHEET_HEADERS, start=1):
        ws_err.cell(row=1, column=col, value=title)
    wb.save(path)
