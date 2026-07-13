"""Лист REPs в книге Ord_Quantity (12 колонок — эталон по скрину)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from gpb_rep_xls import RepMortgageRow, RepParsedFile, depo_account_for_output


REPS_HEADERS = [
    "Источник",
    "№ п/п",
    "Номер операции",
    "НГРИ",
    "ФИО",
    "Код исполнения",
    "Дата операции",
    "Операция",
    "Тип сделки или иного основания",
    "Исполнение",
    "счет депо",
    "Tab1_Счета.владелец",
]


def reps_row_values(
    parsed: RepParsedFile,
    data: RepMortgageRow,
    owner: str,
) -> list:
    depo = depo_account_for_output(parsed)
    return [
        parsed.source_name,
        data.seq,
        data.op_num,
        data.ngri,
        data.fio,
        data.line_state,
        parsed.exec_date,
        parsed.operation,
        parsed.deal_type,
        parsed.header_state,
        depo,
        (owner or "").strip(),
    ]


def ensure_reps_sheet(wb) -> None:
    if "REPs" not in wb.sheetnames:
        wb.create_sheet("REPs")
    ws = wb["REPs"]
    for col, title in enumerate(REPS_HEADERS, start=1):
        ws.cell(row=1, column=col, value=title)


def clear_reps_data(ws) -> None:
    max_r = ws.max_row or 1
    if max_r <= 1:
        return
    for r in range(2, max_r + 1):
        for c in range(1, len(REPS_HEADERS) + 1):
            ws.cell(row=r, column=c, value=None)


def write_reps_rows(path: Path, rows_out: list[list]) -> None:
    wb = load_workbook(path)
    ensure_reps_sheet(wb)
    ws = wb["REPs"]
    clear_reps_data(ws)
    for i, row in enumerate(rows_out, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    wb.save(path)


def append_reps_rows(path: Path, rows_out: list[list]) -> None:
    wb = load_workbook(path)
    ensure_reps_sheet(wb)
    ws = wb["REPs"]
    next_r = (ws.max_row or 1) + 1
    if next_r <= 2 and not (ws.cell(row=1, column=1).value):
        for col, title in enumerate(REPS_HEADERS, start=1):
            ws.cell(row=1, column=col, value=title)
        next_r = 2
    for i, row in enumerate(rows_out):
        for j, val in enumerate(row, start=1):
            ws.cell(row=next_r + i, column=j, value=val)
    wb.save(path)


def create_workbook_with_reps(path: Path) -> None:
    """Минимальный шаблон: лист REPs + «Счета депо» (как для первого запуска)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    default.title = "REPs"
    for col, title in enumerate(REPS_HEADERS, start=1):
        default.cell(row=1, column=col, value=title)
    from dks_excel import SCHEETA_DEPO_HEADERS

    ws_dep = wb.create_sheet("Счета депо")
    for col, title in enumerate(SCHEETA_DEPO_HEADERS, start=1):
        ws_dep.cell(row=1, column=col, value=title)
    wb.save(path)
