"""Парсинг отчётов ВТБ СД F752 (*.xml) и одноимённых ZIP/.xls/.xlsx: НГРИ из приложения (напр. «Приложение 33»)."""

from __future__ import annotations

import sys
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

# Макет «Приложение 33.xls»: заголовки строка 6, данные с 8, НГРИ — колонка B (см. scrins/VTBSD_zip_2).
VTB_APPENDIX_DATA_START_1BASED = 8
VTB_APPENDIX_NGRI_COL_1BASED = 2


def _vtb_ngri_debug(debug: bool, context: str, message: str) -> None:
    if not debug:
        return
    prefix = f"[VTB НГРИ]{f' {context}' if context else ''}"
    print(f"{prefix}: {message}", file=sys.stderr)


def _cell_to_ngri_str(val: object) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _ngri_from_xlrd_sheet(sh, *, debug: bool, context: str, member: str) -> list[str]:
    """sh — лист xlrd; НГРИ с первого листа книги (см. константы старта/колонки)."""
    start = VTB_APPENDIX_DATA_START_1BASED - 1
    col = VTB_APPENDIX_NGRI_COL_1BASED - 1
    out: list[str] = []
    if col >= sh.ncols:
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: лист «{sh.name}» — мало колонок (ncols={sh.ncols}, нужна колонка {VTB_APPENDIX_NGRI_COL_1BASED})",
        )
        return out
    if start >= sh.nrows:
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: лист «{sh.name}» — нет строк с индекса {start} (nrows={sh.nrows})",
        )
        return out
    for r in range(start, sh.nrows):
        cell = sh.cell(r, col)
        s = _cell_to_ngri_str(cell.value)
        if s:
            out.append(s)
    _vtb_ngri_debug(
        debug,
        context,
        f"«{member}» (xlrd, лист «{sh.name}»): извлечено значений НГРИ: {len(out)}",
    )
    return out


def _raw_looks_like_ooxml(raw: bytes) -> bool:
    """OOXML (.xlsx / иногда .xls, сохранённый как ZIP) начинается с PK."""
    return bool(raw) and raw[:2] == b"PK"


def _ngri_from_xls_bytes(
    raw: bytes,
    member: str,
    *,
    debug: bool,
    context: str,
) -> list[str]:
    # Excel часто отдаёт «Приложение 33.xls», которое на деле OOXML — xlrd 2.x такое не читает.
    if _raw_looks_like_ooxml(raw):
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: содержимое OOXML (PK…) — читаем как .xlsx",
        )
        return _ngri_from_xlsx_bytes(raw, member, debug=debug, context=context)
    try:
        import xlrd
    except ModuleNotFoundError:
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: пакет xlrd не установлен (pip install \"xlrd>=2.0.1,<3\")",
        )
        return []
    try:
        book = xlrd.open_workbook(file_contents=raw, formatting_info=False)
    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: xlrd.open_workbook не открыл файл ({err})",
        )
        # xlrd 2: «Excel xlsx file; not supported» — пробуем openpyxl
        if "xlsx" in err.lower() or "openpyxl" in err.lower():
            return _ngri_from_xlsx_bytes(raw, member, debug=debug, context=context)
        return []
    try:
        sh = book.sheet_by_index(0)
    except Exception as e:
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: нет листа 0 ({type(e).__name__}: {e})",
        )
        return []
    return _ngri_from_xlrd_sheet(sh, debug=debug, context=context, member=member)


def _ngri_from_xlsx_bytes(
    raw: bytes,
    member: str,
    *,
    debug: bool,
    context: str,
) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        _vtb_ngri_debug(debug, context, "openpyxl не установлен — .xlsx внутри ZIP не читаем")
        return []
    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}»: load_workbook (.xlsx) не открыл ({type(e).__name__}: {e})",
        )
        return []
    out: list[str] = []
    try:
        ws = wb[wb.sheetnames[0]]
        start_row = VTB_APPENDIX_DATA_START_1BASED
        col_idx = VTB_APPENDIX_NGRI_COL_1BASED
        for row in ws.iter_rows(
            min_row=start_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            cell = row[0]
            s = _cell_to_ngri_str(cell.value)
            if s:
                out.append(s)
        _vtb_ngri_debug(
            debug,
            context,
            f"«{member}» (openpyxl, лист «{ws.title}»): извлечено значений НГРИ: {len(out)}",
        )
    finally:
        wb.close()
    return out


def _local_tag(tag: str) -> str:
    if "}" in tag:
        return tag.split("}", 1)[1]
    return tag


def _text_direct_child(parent: ET.Element | None, local: str) -> str:
    if parent is None:
        return ""
    for child in parent:
        if _local_tag(child.tag) == local:
            return (child.text or "").strip()
    return ""


def _text_first_local_in_subtree(el: ET.Element, local: str) -> str:
    """Первый узел с локальным именем ``local`` в поддереве (в т.ч. не прямой потомок)."""
    for sub in el.iter():
        if _local_tag(sub.tag) == local:
            return (sub.text or "").strip()
    return ""


def vtb_combine_section(name: str, sp_code: str) -> str:
    """Раздел счёта (account section) + «_» + раздел в месте хранения (storage section code)."""
    a = (name or "").strip()
    b = (sp_code or "").strip()
    if not a and not b:
        return ""
    if not a:
        return b
    if not b:
        return a
    return f"{a}_{b}"


@dataclass
class VtbTableStageRow:
    stage_name: str
    odk_date: str
    overlay_depositary_date: str


@dataclass
class VtbF752ParseResult:
    path: Path
    order_number: str
    order_date: str
    execution_date: str
    basis: str
    depository_name: str
    dop_info: str
    qty: str
    src_account: str
    src_section_name: str
    src_sp_section_code: str
    dst_account: str
    dst_section_name: str
    dst_sp_section_code: str
    stage_overlay_date: str
    stage_name: str
    table_stages: list[VtbTableStageRow]


def parse_vtb_f752_xml(path: Path) -> VtbF752ParseResult | None:
    """Разбор F752: шапка у корня и/или в ``ClientCashDataOrder``; этапы — все ``ClientCashDataTable`` в документе.

    Раньше требовался только блок ``ClientCashDataOrder`` и таблицы — только прямые дочерние узлы;
    в выгрузках вроде ``EDO_ODKF752_1_F753_1`` шапка и ``ClientCashDataTable`` могут идти от корня,
    а таблицы — быть вложены не на первом уровне — из‑за этого ``StageName`` оставался пустым.
    Этап для ``stage_name``: последний по документу с непустым ``OverlayDepositaryDate``, иначе последняя таблица.
    """
    try:
        tree = ET.parse(path)
    except (ET.ParseError, OSError):
        return None
    root = tree.getroot()

    order_el: ET.Element | None = None
    for el in root.iter():
        if _local_tag(el.tag) == "ClientCashDataOrder":
            order_el = el
            break

    def co_field(name: str) -> str:
        """Поле заявки: сначала прямой потомок корня, затем — блока ``ClientCashDataOrder``."""
        v = _text_direct_child(root, name)
        return v if v else _text_direct_child(order_el, name)

    stages: list[VtbTableStageRow] = []
    for el in root.iter():
        if _local_tag(el.tag) != "ClientCashDataTable":
            continue
        stages.append(
            VtbTableStageRow(
                stage_name=_text_first_local_in_subtree(el, "StageName"),
                odk_date=_text_first_local_in_subtree(el, "ODKDate"),
                overlay_depositary_date=_text_first_local_in_subtree(
                    el,
                    "OverlayDepositaryDate",
                ),
            ),
        )

    overlay_row: VtbTableStageRow | None = None
    for st in reversed(stages):
        if st.overlay_depositary_date.strip():
            overlay_row = st
            break
    if overlay_row is None and stages:
        overlay_row = stages[-1]

    overlay_date = overlay_row.overlay_depositary_date.strip() if overlay_row else ""
    stage_nm = overlay_row.stage_name.strip() if overlay_row else ""

    return VtbF752ParseResult(
        path=path.resolve(),
        order_number=co_field("OrderNumber"),
        order_date=co_field("OrderDate"),
        execution_date=co_field("ExecutionDate"),
        basis=co_field("Basis"),
        depository_name=co_field("DepositoryName"),
        dop_info=co_field("DopInfo"),
        qty=(co_field("Qty") or "1").strip() or "1",
        src_account=co_field("SrcAccountCode"),
        src_section_name=co_field("SrcSectionName"),
        src_sp_section_code=co_field("SrcSPSectionCode"),
        dst_account=co_field("DstAccountCode"),
        dst_section_name=co_field("DstSectionName"),
        dst_sp_section_code=co_field("DstSPSectionCode"),
        stage_overlay_date=overlay_date,
        stage_name=stage_nm,
        table_stages=stages,
    )


def _vtb_xls_member_sort_key(member_name: str) -> tuple:
    """Сначала типовое приложение с «Приложение» в имени, затем прочие .xls (лексикографически)."""
    base = Path(str(member_name).replace("\\", "/")).name.lower()
    priority = 0 if "приложение" in base else 1
    return (priority, base, member_name)


def _zip_resolve_inner_name(zf: zipfile.ZipFile, member_name: str) -> str:
    """Имя записи в архиве как в zf.namelist() (учёт слэшей и лишних /)."""
    names = [n for n in zf.namelist() if not n.endswith("/")]
    if member_name in names:
        return member_name
    norm_req = str(member_name).replace("\\", "/").strip().lstrip("/")
    for n in names:
        if n.replace("\\", "/").strip().lstrip("/") == norm_req:
            return n
    raise KeyError(member_name)


def _zip_read_entry(zf: zipfile.ZipFile, member_name: str) -> bytes:
    return zf.read(_zip_resolve_inner_name(zf, member_name))
def list_ngri_from_vtb_zip(
    zip_path: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> list[str]:
    """Все НГРИ из всех .xls и при необходимости .xlsx внутри ZIP.

    Строки данных — с 1-based строки 8, НГРИ — колонка B (см. константы).

    Ограничения и типовые причины пустого списка (при ``debug=True`` пишутся в stderr):
    архив не открывается; внутри нет .xls/.xlsx; неверный формат (не OLE и не OOXML);
    нет пакета xlrd; первая вкладка пустая или макет сдвинут; несколько таблиц — перебираются
    **все** вхождения .xls (сначала имена с «Приложение» в basename), затем .xlsx.

    Каждое непустое значение ячейки — отдельная запись в списке (несколько строк приложения
    дают несколько НГРИ).
    """
    ctx = context.strip()
    zp = Path(zip_path)
    if not zp.is_file():
        _vtb_ngri_debug(debug, ctx, f"ZIP не найден по пути: {zp}")
        return []
    try:
        zf = zipfile.ZipFile(zp, "r")
    except (OSError, zipfile.BadZipFile) as e:
        _vtb_ngri_debug(
            debug,
            ctx,
            f"не удалось открыть ZIP «{zp.name}» ({type(e).__name__}: {e})",
        )
        return []

    all_members: list[str] = []
    try:
        all_members = zf.namelist()
        lower_names = [n for n in all_members if not n.endswith("/")]
        xls_names = sorted(
            (
                n
                for n in lower_names
                if n.lower().endswith(".xls") and not n.lower().endswith(".xlsx")
            ),
            key=_vtb_xls_member_sort_key,
        )
        xlsx_names = sorted(n for n in lower_names if n.lower().endswith(".xlsx"))
        _vtb_ngri_debug(
            debug,
            ctx,
            f"«{zp.name}»: в архиве {len(lower_names)} элементов, .xls={len(xls_names)}, .xlsx={len(xlsx_names)}",
        )
        if debug and not xls_names and not xlsx_names:
            sample = ", ".join(f"«{n}»" for n in lower_names[:12])
            more = f" (+ещё {len(lower_names) - 12})" if len(lower_names) > 12 else ""
            _vtb_ngri_debug(
                debug,
                ctx,
                f"нет файлов .xls/.xlsx; пример имён в архиве: {sample}{more}",
            )

        out: list[str] = []
        for name in xls_names:
            try:
                raw = _zip_read_entry(zf, name)
            except (KeyError, OSError) as e:
                _vtb_ngri_debug(
                    debug,
                    ctx,
                    f"не прочитан «{name}» ({type(e).__name__}: {e})",
                )
                continue
            part = _ngri_from_xls_bytes(raw, name, debug=debug, context=ctx)
            out.extend(part)

        if not out:
            for name in xlsx_names:
                try:
                    raw = _zip_read_entry(zf, name)
                except (KeyError, OSError) as e:
                    _vtb_ngri_debug(
                        debug,
                        ctx,
                        f"не прочитан «{name}» ({type(e).__name__}: {e})",
                    )
                    continue
                out.extend(
                    _ngri_from_xlsx_bytes(
                        raw,
                        name,
                        debug=debug,
                        context=ctx,
                    ),
                )

        if not out:
            _vtb_ngri_debug(
                debug,
                ctx,
                f"«{zp.name}»: после разбора .xls/.xlsx список НГРИ пуст (проверьте строку {VTB_APPENDIX_DATA_START_1BASED}, колонку {VTB_APPENDIX_NGRI_COL_1BASED})",
            )
        return out
    finally:
        zf.close()


def find_vtb_appendix_for_xml(
    xml_path: Path,
    apps_dir: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> Path | None:
    """Если ZIP нет — одноимённое приложение ``.xls`` / ``.xlsx`` (тот же stem, что у XML).

    Порядок: ``.xls`` → ``.xlsx`` (с вариантами регистра в суффиксе).
    """
    stem = xml_path.stem
    candidates = (
        f"{stem}.xls",
        f"{stem}.XLS",
        f"{stem}.Xls",
        f"{stem}.xlsx",
        f"{stem}.XLSX",
        f"{stem}.Xlsx",
    )
    for name in candidates:
        p = (apps_dir / name).resolve()
        if p.is_file():
            _vtb_ngri_debug(
                debug,
                context,
                f"найдено одноимённое приложение (вместо ZIP): «{p.name}»",
            )
            return p
    stem_l = stem.lower()
    try:
        preferred: list[Path] = []
        for f in apps_dir.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() not in (".xls", ".xlsx"):
                continue
            if f.stem.lower() == stem_l:
                preferred.append(f.resolve())
        # .xls раньше .xlsx при совпадении stem
        preferred.sort(key=lambda p: (0 if p.suffix.lower() == ".xls" else 1, p.name.lower()))
        if preferred:
            rp = preferred[0]
            _vtb_ngri_debug(
                debug,
                context,
                f"найдено одноимённое приложение (stem без учёта регистра): «{rp.name}»",
            )
            return rp
    except OSError:
        pass
    _vtb_ngri_debug(
        debug,
        context,
        f"одноимённое .xls/.xlsx не найдено в «{apps_dir}» "
        f"(ожидалось: {stem}.xls | {stem}.xlsx)",
    )
    return None


def find_vtb_xls_for_xml(
    xml_path: Path,
    xls_dir: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> Path | None:
    """Алиас: одноимённое приложение ``.xls`` / ``.xlsx`` в каталоге приложений."""
    return find_vtb_appendix_for_xml(
        xml_path,
        xls_dir,
        debug=debug,
        context=context,
    )


def list_ngri_from_vtb_appendix_path(
    appendix_path: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> list[str]:
    """НГРИ из отдельного ``.xls`` / ``.xlsx`` (макет «Приложение 33»: строка 8, колонка B)."""
    ctx = context.strip()
    p = Path(appendix_path)
    if not p.is_file():
        _vtb_ngri_debug(debug, ctx, f"приложение не найдено: {p}")
        return []
    try:
        raw = p.read_bytes()
    except OSError as e:
        _vtb_ngri_debug(
            debug,
            ctx,
            f"не удалось прочитать «{p.name}» ({type(e).__name__}: {e})",
        )
        return []
    suf = p.suffix.lower()
    if suf == ".xlsx" or _raw_looks_like_ooxml(raw):
        return _ngri_from_xlsx_bytes(raw, p.name, debug=debug, context=ctx)
    return _ngri_from_xls_bytes(raw, p.name, debug=debug, context=ctx)


def list_ngri_from_vtb_xls_path(
    xls_path: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> list[str]:
    """Алиас ``list_ngri_from_vtb_appendix_path`` (``.xls`` и ``.xlsx``)."""
    return list_ngri_from_vtb_appendix_path(
        xls_path,
        debug=debug,
        context=context,
    )


def collect_vtb_ngri_for_xml(
    xml_path: Path,
    apps_dir: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> list[str]:
    """НГРИ для XML: ZIP (тот же stem) → иначе одноимённое ``.xls``/``.xlsx``.

    Пустой список — вызывающий код обычно подставляет ``[\"\"]`` (одна строка СВОД с «-»),
    либо останавливает сборку, если приложение обязательно (см. ``find_vtb_attachment_for_xml``).
    """
    ctx = context.strip() or xml_path.name
    att = find_vtb_attachment_for_xml(
        xml_path,
        apps_dir,
        debug=debug,
        context=ctx,
    )
    if att is None:
        return []
    if att.suffix.lower() == ".zip":
        return list_ngri_from_vtb_zip(att, debug=debug, context=ctx)
    return list_ngri_from_vtb_appendix_path(att, debug=debug, context=ctx)


def find_vtb_attachment_for_xml(
    xml_path: Path,
    apps_dir: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> Path | None:
    """Одноимённое приложение для XML: ``.zip`` → иначе ``.xls`` / ``.xlsx`` (тот же stem)."""
    ctx = context.strip() or xml_path.name
    if not apps_dir.is_dir():
        _vtb_ngri_debug(debug, ctx, f"каталог приложений не найден: {apps_dir}")
        return None
    zp = find_vtb_zip_for_xml(xml_path, apps_dir, debug=debug, context=ctx)
    if zp is not None:
        return zp
    return find_vtb_appendix_for_xml(xml_path, apps_dir, debug=debug, context=ctx)


def list_vtb_xml_missing_attachments(
    xml_paths: list[Path],
    apps_dir: Path,
) -> list[Path]:
    """XML без одноимённого ZIP/XLS/XLSX в ``apps_dir`` (порядок как во входе)."""
    missing: list[Path] = []
    for xp in xml_paths:
        if find_vtb_attachment_for_xml(xp, apps_dir) is None:
            missing.append(xp)
    return missing


def format_vtb_missing_attachments_error(
    missing: list[Path],
    apps_dir: Path,
) -> str:
    """Текст ошибки для stderr / SMTP-отбивки."""
    lines = [
        "ВТБ СД: для XML не найдено одноимённое приложение "
        f"(ожидается тот же stem: .zip / .xls / .xlsx) в каталоге: {apps_dir}",
        "Сборка остановлена. Добавьте приложения и повторите запуск.",
        "Файлы без приложения:",
    ]
    for xp in missing:
        stem = xp.stem
        lines.append(
            f"  - {xp.name}  (нужно: {stem}.zip | {stem}.xls | {stem}.xlsx)",
        )
    return "\n".join(lines)


def find_vtb_zip_for_xml(
    xml_path: Path,
    zip_dir: Path,
    *,
    debug: bool = False,
    context: str = "",
) -> Path | None:
    """Находит ZIP рядом с XML: тот же stem (имя без .xml = имя без .zip). НГРИ — в .xls внутри архива."""
    stem = xml_path.stem
    for name in (f"{stem}.ZIP", f"{stem}.zip", f"{stem}.Zip"):
        p = (zip_dir / name).resolve()
        if p.is_file():
            return p
    stem_l = stem.lower()
    try:
        for f in zip_dir.iterdir():
            if not f.is_file():
                continue
            if f.suffix.lower() != ".zip":
                continue
            if f.stem.lower() == stem_l:
                rp = f.resolve()
                _vtb_ngri_debug(
                    debug,
                    context,
                    f"найден ZIP (совпадение stem без учёта регистра): «{f.name}»",
                )
                return rp
    except OSError:
        pass
    _vtb_ngri_debug(
        debug,
        context,
        f"ZIP не найден в «{zip_dir}» (ожидаемые имена: {stem}.ZIP | .zip; то же stem без учёта регистра .zip)",
    )
    return None
