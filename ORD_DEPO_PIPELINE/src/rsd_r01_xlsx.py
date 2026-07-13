"""Парсинг отчёта РСД формы R-01 (*.xlsx): первый блок операций, без таблицы остатков."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# Вторая таблица (остатки) — не парсить
R01_HOLDINGS_MARKER = "ценные бумаги, учитываемые на счете депо"
R01_DEPO_ACCOUNT_LEN = 12


def normalize_r01_depo_account(acc: str) -> str:
    """
    Номер счёта депо R-01: эталон — 12 цифр.
    Если не хватает ровно 3 символа (9 цифр) и счёт начинается с «3» — дописать «000» в начало.
    """
    s = (acc or "").replace(" ", "").strip()
    if not s or not s.isdigit():
        return s
    if len(s) == R01_DEPO_ACCOUNT_LEN - 3 and s.startswith("3"):
        return "000" + s
    return s


@dataclass
class R01DataRow:
    operation_type: str
    basis: str
    section_raw: str
    reg_number: str
    exec_date_display: str
    qty: str


@dataclass
class R01ParseResult:
    path: Path
    depo_account: str
    rows: list[R01DataRow]


def _cell_value_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _norm_scan(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def row_has_holdings_marker(ws: Worksheet, row: int, max_col: int = 24) -> bool:
    for c in range(1, min(max_col, (ws.max_column or 24) + 1) + 1):
        t = _norm_scan(_cell_value_str(ws.cell(row=row, column=c).value))
        if R01_HOLDINGS_MARKER in t:
            return True
    return False


def _pick_sheet(wb) -> Any:
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    for name in wb.sheetnames:
        ws = wb[name]
        for r in range(1, min(25, (ws.max_row or 1) + 20)):
            row_txt = ""
            for c in range(1, min(15, (ws.max_column or 15) + 1)):
                row_txt += _norm_scan(_cell_value_str(ws.cell(row=r, column=c).value)) + " "
            if "совершении" in row_txt and "депозитар" in row_txt:
                return ws
            if "r-01" in row_txt or "отчет о совершении" in row_txt:
                return ws
    return wb[wb.sheetnames[0]]


def find_r01_depo_account(ws: Worksheet) -> str:
    """Номер счёта депо в шапке: 9 (с «3»…) или 11–14 цифр рядом с «депо»/«счет»."""
    acc_re = re.compile(r"^(\d{9}|\d{11,14})$")
    for r in range(1, min(55, (ws.max_row or 55) + 1)):
        for c in range(1, min(14, (ws.max_column or 14) + 1)):
            raw = _cell_value_str(ws.cell(row=r, column=c).value)
            compact = raw.replace(" ", "")
            if not acc_re.match(compact):
                continue
            ctx = ""
            for dr in (-2, -1, 0, 1, 2):
                for dc in (-4, -3, -2, -1, 0, 1, 2, 3, 4):
                    rr, cc = r + dr, c + dc
                    if rr < 1 or cc < 1:
                        continue
                    ctx += _norm_scan(_cell_value_str(ws.cell(row=rr, column=cc).value))
            if "депо" in ctx or "счет" in ctx or "счёт" in ctx:
                return normalize_r01_depo_account(compact)
    return ""


def _header_cell_norm(value: Any) -> str:
    return _norm_scan(_cell_value_str(value))


def find_operations_table_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    """
    Строка заголовков первой таблицы: «Тип операции», «Номер раздела», «Дата исполнения операции».
    Возвращает (номер_строки, ключ -> индекс_колонки).
    """
    max_r = min(100, (ws.max_row or 1) + 80)
    max_c = min(30, (ws.max_column or 20) + 10)
    for r in range(1, max_r + 1):
        if row_has_holdings_marker(ws, r, max_c):
            break
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            h = _header_cell_norm(ws.cell(row=r, column=c).value)
            if not h:
                continue
            if "тип операции" in h:
                col_map["op_type"] = c
            if "номер раздела" in h:
                col_map["section"] = c
            if "дата проведения" in h or "по месту хранения" in h:
                pass
            elif "дата исполнения" in h or (
                "дата" in h and "исполнен" in h and "проведен" not in h
            ):
                col_map["exec_date"] = c
            if "документ" in h and "основан" in h:
                col_map["basis"] = c
            if ("гос" in h and "рег" in h) or "isin" in h:
                col_map["reg_no"] = c
            if "количество" in h and ("ценных бумаг" in h or "цб" in h):
                col_map["qty"] = c
        if col_map.get("op_type") and col_map.get("section") and col_map.get("exec_date"):
            return r, col_map
    return 0, {}


def parse_section_code(section_cell: str) -> str:
    """«100042 - Основной» -> «100042»."""
    s = (section_cell or "").strip()
    if not s:
        return ""
    part = s.split(" - ")[0].strip()
    try:
        return str(int(float(part)))
    except (ValueError, TypeError):
        return part


def parse_r01_xlsx(path: Path) -> R01ParseResult | None:
    path = path.resolve()
    wb = None
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except OSError:
        return None
    try:
        ws = _pick_sheet(wb)
        depo = normalize_r01_depo_account(find_r01_depo_account(ws))
        hr, cmap = find_operations_table_header(ws)
        if not hr or not cmap:
            return None

        c_op = cmap["op_type"]
        c_sec = cmap["section"]
        c_dt = cmap["exec_date"]
        c_basis = cmap.get("basis")
        c_reg = cmap.get("reg_no")
        c_qty = cmap.get("qty")

        rows_out: list[R01DataRow] = []
        last_row = ws.max_row or hr
        max_r = max(last_row, hr + 1) + 400
        for r in range(hr + 1, max_r + 1):
            if row_has_holdings_marker(ws, r):
                break
            op_t = _cell_value_str(ws.cell(row=r, column=c_op).value)
            if not op_t.strip():
                continue
            sec_raw = _cell_value_str(ws.cell(row=r, column=c_sec).value)
            dt_raw = ws.cell(row=r, column=c_dt).value
            exec_disp = _cell_value_str(dt_raw)
            basis = _cell_value_str(ws.cell(row=r, column=c_basis).value) if c_basis else ""
            reg = _cell_value_str(ws.cell(row=r, column=c_reg).value) if c_reg else ""
            qty = _cell_value_str(ws.cell(row=r, column=c_qty).value) if c_qty else "1"
            if not qty.strip():
                qty = "1"

            rows_out.append(
                R01DataRow(
                    operation_type=op_t.strip(),
                    basis=basis.strip(),
                    section_raw=sec_raw.strip(),
                    reg_number=reg.strip(),
                    exec_date_display=exec_disp.strip(),
                    qty=qty.strip() or "1",
                ),
            )
        return R01ParseResult(path=path, depo_account=depo, rows=rows_out)
    finally:
        if wb is not None:
            wb.close()
