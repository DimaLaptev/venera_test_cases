"""Загрузка и поиск по листу «Счета депо» в книге Excel."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Справочник «Счета депо»: отбор строк для поручений ВТБ СД по полю Basis (номер договора).
VTB_SD_DEPOSITORY_NAME = 'АО "ВТБ СД"'
VTB_CONTRACT_7830_OWNER = "ПАО ДОМ.РФ"
VTB_CONTRACT_OTHER_OWNER = 'ООО "ДОМ.РФ ИА"'
CONTRACT_7830 = "7830"


def _catalog_compare_normalize(value: str) -> str:
    """Нормализация для сравнения ячеек справочника (пробелы, типографские кавычки)."""
    t = (value or "").strip()
    for a, b in (
        ("\u201c", '"'),
        ("\u201d", '"'),
        ("\u00ab", '"'),
        ("\u00bb", '"'),
    ):
        t = t.replace(a, b)
    t = re.sub(r"\s+", " ", t)
    return t.casefold()


@dataclass
class DepoDirectoryRow:
    portfolio_section: str
    depository: str
    account_number: str
    originator: str
    servicing_agent: str
    deal_dom_rf_ia: str
    comment: str
    gpb_code_dom: str
    gpb_code_bank: str
    owner: str


def _norm_header(cell: Any) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip().lower()


def _find_header_row(ws: Worksheet, max_scan: int = 40) -> tuple[int, dict[str, int]]:
    for r in range(1, min(max_scan, (ws.max_row or 1) + 25) + 1):
        col_map: dict[str, int] = {}
        for c in range(1, (ws.max_column or 15) + 1):
            h = _norm_header(ws.cell(row=r, column=c))
            if h:
                col_map[h] = c
        if "номер счета депо" in col_map:
            return r, col_map
    return 1, {}


def normalize_match_token(value: Any) -> str:
    """Сопоставление ClientID с кодами выписки (70, 70.0, '070' → '70')."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        f = float(s.replace(",", "."))
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def read_depo_directory_rows(
    workbook_path: str,
    sheet_name: str = "Счета депо",
) -> list[DepoDirectoryRow]:
    """Все строки справочника (для поиска счёта по ClientID)."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    header_row, col_map = _find_header_row(ws)
    if not col_map:
        wb.close()
        return []

    def col(name: str) -> int | None:
        return col_map.get(name)

    c_acc = col("номер счета депо")
    if not c_acc:
        wb.close()
        return []

    c_pf = col("раздел консолидированного портфеля")
    c_dep = col("депозитарий")
    c_orig = col("оригинатор")
    c_agent = col("агент по сопровождению")
    c_deal = col("сделка дом.рф иа")
    c_com = col("комментарий")
    c_cd = col("код выписки гпб (дом)")
    c_cb = col("код выписки гпб (банк)")
    c_own = col("владелец счета депо")

    rows_out: list[DepoDirectoryRow] = []

    def cell_val(r: int, ci: int | None) -> str:
        if not ci:
            return ""
        v = ws.cell(row=r, column=ci).value
        if v is None:
            return ""
        return str(v).strip()

    for r in range(header_row + 1, (ws.max_row or 0) + 1):
        acc_cell = ws.cell(row=r, column=c_acc).value
        if acc_cell is None or str(acc_cell).strip() == "":
            continue

        rows_out.append(
            DepoDirectoryRow(
                portfolio_section=cell_val(r, c_pf),
                depository=cell_val(r, c_dep),
                account_number=str(acc_cell).strip(),
                originator=cell_val(r, c_orig),
                servicing_agent=cell_val(r, c_agent),
                deal_dom_rf_ia=cell_val(r, c_deal),
                comment=cell_val(r, c_com),
                gpb_code_dom=cell_val(r, c_cd),
                gpb_code_bank=cell_val(r, c_cb),
                owner=cell_val(r, c_own),
            )
        )

    wb.close()
    return rows_out


def lookup_account_by_gpb_dom_code(
    rows: list[DepoDirectoryRow],
    client_id: str,
) -> str | None:
    """
    «Номер счета депо» (колонка C): строка справочника, где ClientID совпал только с
    «Код выписки ГПБ (ДОМ)» (колонка H), без сопоставления с кодом банка.
    """
    target = normalize_match_token(client_id)
    if not target:
        return None
    for row in rows:
        dom = normalize_match_token(row.gpb_code_dom)
        if target == dom:
            return row.account_number
    return None


def lookup_account_by_gpb_bank_code(
    rows: list[DepoDirectoryRow],
    client_id: str,
) -> str | None:
    """
    «Номер счета депо» (колонка C): строка справочника, где ClientID совпал только с
    «Код выписки ГПБ (БАНК)» (колонка I).
    """
    target = normalize_match_token(client_id)
    if not target:
        return None
    for row in rows:
        bank = normalize_match_token(row.gpb_code_bank)
        if target == bank:
            return row.account_number
    return None


def lookup_account_by_client_id(
    rows: list[DepoDirectoryRow],
    client_id: str,
) -> str | None:
    """
    «Номер счета депо»: строка справочника, где ClientID совпал с
    «Код выписки ГПБ (ДОМ)» или «Код выписки ГПБ (БАНК)».
    """
    target = normalize_match_token(client_id)
    if not target:
        return None
    for row in rows:
        dom = normalize_match_token(row.gpb_code_dom)
        bank = normalize_match_token(row.gpb_code_bank)
        if target == dom or target == bank:
            return row.account_number
    return None


def lookup_depo_row_by_client_id(
    rows: list[DepoDirectoryRow],
    client_id: str,
) -> DepoDirectoryRow | None:
    """Полная строка справочника по ClientID ↔ коды выписки."""
    target = normalize_match_token(client_id)
    if not target:
        return None
    for row in rows:
        dom = normalize_match_token(row.gpb_code_dom)
        bank = normalize_match_token(row.gpb_code_bank)
        if target == dom or target == bank:
            return row
    return None


def lookup_depo_row_by_account_number(
    rows: list[DepoDirectoryRow],
    account: str,
) -> DepoDirectoryRow | None:
    """Строка по колонке «Номер счета депо» (для колонки E «Владелец»)."""
    k = normalize_account_key(account)
    if not k:
        return None
    for row in rows:
        if normalize_account_key(row.account_number) == k:
            return row
    return None


def depo_account_owners_equal(
    rows: list[DepoDirectoryRow],
    account_a: str,
    account_b: str,
) -> bool:
    """Оба счёта в «Счета депо» и «Владелец счета депо» совпадает (нормализованное сравнение)."""
    row_a = lookup_depo_row_by_account_number(rows, account_a)
    row_b = lookup_depo_row_by_account_number(rows, account_b)
    if row_a is None or row_b is None:
        return False
    own_a = (row_a.owner or "").strip()
    own_b = (row_b.owner or "").strip()
    if not own_a or not own_b:
        return False
    return _catalog_compare_normalize(own_a) == _catalog_compare_normalize(own_b)


def load_depo_index(
    workbook_path: str,
    sheet_name: str = "Счета депо",
) -> dict[str, DepoDirectoryRow]:
    """Индекс по номеру счёта депо (для совместимости)."""
    rows = read_depo_directory_rows(workbook_path, sheet_name)
    index: dict[str, DepoDirectoryRow] = {}
    for row in rows:
        acc_key = normalize_account_key(row.account_number)
        if acc_key:
            index[acc_key] = row
            try:
                index[str(int(float(row.account_number)))] = row
            except (ValueError, TypeError):
                pass
    return index


def normalize_account_key(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, TypeError):
        pass
    return s


def lookup_depo(index: dict[str, DepoDirectoryRow], depo_num: str) -> DepoDirectoryRow | None:
    if not depo_num.strip():
        return None
    k = normalize_account_key(depo_num.strip())
    if k in index:
        return index[k]
    raw = depo_num.strip()
    if raw in index:
        return index[raw]
    return None


def lookup_depo_row_vtb_account_owner_filter(
    rows: list[DepoDirectoryRow],
    account: str,
    *,
    expected_depository: str,
    expected_owner: str,
) -> DepoDirectoryRow | None:
    """Строка «Счета депо»: номер счёта + депозитарий + владелец (после нормализации)."""
    acc_key = normalize_account_key(account)
    if not acc_key:
        return None
    dep_t = _catalog_compare_normalize(expected_depository)
    own_t = _catalog_compare_normalize(expected_owner)
    for row in rows:
        if normalize_account_key(row.account_number) != acc_key:
            continue
        if _catalog_compare_normalize(row.depository) != dep_t:
            continue
        if _catalog_compare_normalize(row.owner) != own_t:
            continue
        return row
    return None
