"""Чтение листа «Коды» и поиск «Тип поручения» / «Тип операции»."""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook

from depo_directory import normalize_account_key


@dataclass
class VtbSdCodifierRow:
    op_code: str
    op_type: str
    content_vtb: str
    instruction_type: str


@dataclass
class CodifierRow:
    op_code: str
    exec_code: str
    op_type: str
    content_gpb: str
    instruction_type: str
    comment: str
    content_gpb_rep: str = ""
    content_rsd: str = ""
    content_region: str = ""
    content_vtb: str = ""


def _norm_header(cell: Any) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip().lower()


def _find_header_row(ws, max_scan: int = 40) -> tuple[int, dict[str, int]]:
    for r in range(1, min(max_scan, (ws.max_row or 1) + 30) + 1):
        col_map: dict[str, int] = {}
        for c in range(1, (ws.max_column or 20) + 1):
            h = _norm_header(ws.cell(row=r, column=c))
            if h:
                col_map[h] = c
        if "код операции" in col_map and "тип поручения" in col_map:
            return r, col_map
    return 1, {}


def _col_content_gpb_rep(col_map: dict[str, int]) -> int | None:
    """Колонка «Содержание ГПБ в REP» (если есть в шапке листа «Коды»)."""
    for key, idx in col_map.items():
        if "содержание" in key and "rep" in key:
            return idx
    return None


def _col_content_region(col_map: dict[str, int]) -> int | None:
    """Колонка «Содержание РЕГИОН» или длинный заголовок «… в отчёте P»."""
    for key, idx in col_map.items():
        if "содержание" in key and "регион" in key:
            return idx
    for key, idx in col_map.items():
        if "отчете" in key and "тип операции" in key:
            return idx
    return None


def _col_content_vtb(col_map: dict[str, int]) -> int | None:
    """Колонка «Содержание ВТБСД»."""
    for key, idx in col_map.items():
        if "содержание" in key and "втб" in key:
            return idx
    return None


def _col_content_rsd(col_map: dict[str, int]) -> int | None:
    """Колонка «Содержание РСД (для R-01, R-05)» или короткий заголовок."""
    for key, idx in col_map.items():
        if "содержание" in key and "рсд" in key:
            return idx
    return None


def _cell(ws, r: int, col: int | None) -> str:
    if not col:
        return ""
    v = ws.cell(row=r, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def _is_dash(s: str) -> bool:
    t = s.strip()
    return t in ("", "—", "-", "–", "None")


def read_codifier_rows(workbook_path: str, sheet_name: str = "Коды") -> list[CodifierRow]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    hr, col_map = _find_header_row(ws)
    if not col_map:
        wb.close()
        return []

    def col(*names: str) -> int | None:
        for n in names:
            k = n.lower()
            if k in col_map:
                return col_map[k]
        return None

    c_op = col("код операции")
    c_ex = col("код исполнения")
    c_type = col("тип операции")
    c_gpb = col("содержание гпб")
    c_rsd = col("содержание рсд") or _col_content_rsd(col_map)
    c_inst = col("тип поручения")
    c_com = col("комментарий")
    c_gpb_rep = _col_content_gpb_rep(col_map)
    c_region = _col_content_region(col_map)
    c_vtb = _col_content_vtb(col_map)

    if not c_op or not c_inst:
        wb.close()
        return []

    out: list[CodifierRow] = []
    for r in range(hr + 1, (ws.max_row or 0) + 1):
        op = _cell(ws, r, c_op)
        if _is_dash(op):
            op = ""
        ex = _cell(ws, r, c_ex) if c_ex else ""
        if _is_dash(ex):
            ex = ""
        ot = _cell(ws, r, c_type) if c_type else ""
        gpb = _cell(ws, r, c_gpb) if c_gpb else ""
        rsd = _cell(ws, r, c_rsd) if c_rsd else ""
        if _is_dash(rsd):
            rsd = ""
        inst = _cell(ws, r, c_inst)
        com = _cell(ws, r, c_com) if c_com else ""
        gpb_rep = _cell(ws, r, c_gpb_rep) if c_gpb_rep else ""
        reg_c = _cell(ws, r, c_region) if c_region else ""
        if _is_dash(reg_c):
            reg_c = ""
        vtb_c = _cell(ws, r, c_vtb) if c_vtb else ""
        if _is_dash(vtb_c):
            vtb_c = ""
        if not op and not ot and not inst:
            continue
        out.append(
            CodifierRow(
                op_code=op,
                exec_code=ex,
                op_type=ot,
                content_gpb=gpb,
                instruction_type=inst,
                comment=com,
                content_gpb_rep=gpb_rep,
                content_rsd=rsd,
                content_region=reg_c,
                content_vtb=vtb_c,
            )
        )
    wb.close()
    return out


def _norm_sheet_title_for_vtb_match(name: str) -> str:
    t = (name or "").replace("\u00a0", " ").strip().lower()
    return t.replace("ё", "е")


def _sheet_name_looks_like_vtb_codifier(name: str) -> bool:
    """Лист вроде «Коды ВТБ СД» / «Коды ВТб СБ» (опечатки в имени)."""
    k = _norm_sheet_title_for_vtb_match(name)
    return "коды" in k and "втб" in k


def _vtb_sd_codifier_sheet_candidates(wb) -> list[str]:
    """
    Порядок: точное «Коды ВТБ СД», затем остальные имена с «коды» и «втб».
    Нужно, если в книге лист назван, например, «Коды ВТб СБ».
    """
    pref = "Коды ВТБ СД"
    out: list[str] = []
    seen: set[str] = set()
    if pref in wb.sheetnames:
        out.append(pref)
        seen.add(pref)
    for sn in wb.sheetnames:
        if sn in seen:
            continue
        if _sheet_name_looks_like_vtb_codifier(sn):
            out.append(sn)
            seen.add(sn)
    return out


def _find_vtb_sd_header_row(ws, max_scan: int = 30) -> tuple[int, dict[str, int]]:
    for r in range(1, min(max_scan, (ws.max_row or 1) + 15) + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max(ws.max_column or 0, 12) + 1):
            h = _norm_header(ws.cell(row=r, column=c))
            if h:
                col_map[h] = c
        if "код операции" not in col_map:
            continue
        if "тип операции" not in col_map:
            continue
        if "тип поручения" not in col_map:
            continue
        has_vtb = any(
            "содержание" in key and ("втб" in key or "втбсд" in key)
            for key in col_map
        )
        if has_vtb:
            return r, col_map
    return 1, {}


def _vtb_sd_content_col(col_map: dict[str, int]) -> int | None:
    for key, idx in col_map.items():
        if "содержание" in key and ("втб" in key or "втбсд" in key):
            return idx
    return None


def _read_vtb_sd_codifier_rows_from_ws(ws) -> list[VtbSdCodifierRow]:
    """Читает таблицу кодификатора ВТБ СД с одного листа (первая подходящая шапка)."""
    hr, col_map = _find_vtb_sd_header_row(ws)
    if not col_map:
        return []
    c_op = col_map.get("код операции")
    c_ot = col_map.get("тип операции")
    c_inst = col_map.get("тип поручения")
    c_vtb = _vtb_sd_content_col(col_map)
    if not c_op or not c_ot or not c_inst or not c_vtb:
        return []

    out: list[VtbSdCodifierRow] = []
    for r in range(hr + 1, (ws.max_row or 0) + 1):
        op = _cell(ws, r, c_op)
        if _is_dash(op):
            op = ""
        ot = _cell(ws, r, c_ot)
        inst = _cell(ws, r, c_inst)
        cv = _cell(ws, r, c_vtb)
        if _is_dash(cv):
            cv = ""
        if not cv and not op and not ot and not inst:
            continue
        out.append(
            VtbSdCodifierRow(
                op_code=op,
                op_type=ot,
                content_vtb=cv,
                instruction_type=inst,
            ),
        )
    return out


def read_vtb_sd_codifier_rows(
    workbook_path: str,
    sheet_name: str | None = None,
) -> list[VtbSdCodifierRow]:
    """
    Лист кодификатора ВТБ СД: колонки «Код операции», «Тип операции»,
    «Содержание … ВТБСД», «Тип поручения».

    Если ``sheet_name`` задан — только этот лист. Иначе: «Коды ВТБ СД»,
    затем любой лист, в имени которого есть «коды» и «втб» (например «Коды ВТб СБ»);
    берётся первый лист, с которого удалось прочитать хотя бы одну строку данных.
    """
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return []
            return _read_vtb_sd_codifier_rows_from_ws(wb[sheet_name])
        for sn in _vtb_sd_codifier_sheet_candidates(wb):
            rows = _read_vtb_sd_codifier_rows_from_ws(wb[sn])
            if rows:
                return rows
        return []
    finally:
        wb.close()


def _norm_match_snippet_vtb(s: str) -> str:
    """
    Нормализация для сравнения StageName (XML) с колонкой «Содержание ВТБСД» (Excel):
    trim, lower (регистр), неразрывные пробелы → обычный пробел, «ё» → «е»,
    подряд идущие пробелы в один. Далее проверка только подстрок (substring), без regex.
    """
    t = (s or "").replace("\u00a0", " ").replace("\u202f", " ")
    t = t.strip().lower()
    t = t.replace("ё", "е")
    t = re.sub(r"\s+", " ", t)
    return t


def _vtb_sd_op_type_is_spisanie(op_type: str) -> bool:
    t = (op_type or "").lower().replace("–", "-").replace("—", "-")
    return " - списание" in t


def _vtb_sd_op_type_is_zachislenie(op_type: str) -> bool:
    t = (op_type or "").lower().replace("–", "-").replace("—", "-")
    return " - зачисление" in t


def lookup_codifier_vtb_sd_sheet(
    rows: list[VtbSdCodifierRow],
    stage_name: str,
    f_account: str,
    h_account: str,
    j_account: str,
) -> tuple[str, str, str]:
    """
    Маппинг StageName на лист «Коды ВТБ СД» (колонка «Содержание ВТБСД»).
    Сверка текста: после :func:`_norm_match_snippet_vtb` — полное совпадение сильнее;
    допускается вложение **короткого** «Содержание ВТБСД» в StageName (``cv in lab``).
    Вложение метки в **более длинную** ячейку без полного равенства не считается.

    Направление: счёт депо (F) = счёт списания (H) → строка «… - списание»;
    F = зачисление (J) → «… - зачисление».
    Возвращает (тип_поручения, тип_операции, код_операции).
    """
    if not rows:
        return "", "", ""
    lab = _norm_match_snippet_vtb(stage_name)
    if not lab:
        return "", "", ""

    scored: list[tuple[int, VtbSdCodifierRow]] = []
    for r in rows:
        cv = _norm_match_snippet_vtb(r.content_vtb)
        if not cv:
            continue
        score = _codifier_label_token_score(lab, cv, 5)
        if score > 0:
            scored.append((score, r))

    if not scored:
        return "", "", ""

    top_s = max(s for s, _ in scored)
    candidates = [row for s, row in scored if s == top_s]

    fk = normalize_account_key(f_account)
    hk = normalize_account_key(h_account)
    jk = normalize_account_key(j_account)
    match_spis = bool(fk and hk and fk == hk)
    match_zach = bool(fk and jk and fk == jk)

    chosen: VtbSdCodifierRow | None = None
    if match_spis:
        for row in candidates:
            if _vtb_sd_op_type_is_spisanie(row.op_type):
                chosen = row
                break
    if chosen is None and match_zach:
        for row in candidates:
            if _vtb_sd_op_type_is_zachislenie(row.op_type):
                chosen = row
                break
    if chosen is None and len(candidates) == 1:
        chosen = candidates[0]

    if chosen is None:
        return "", "", ""

    opc_raw = (chosen.op_code or "").strip()
    if _is_dash(opc_raw):
        opc_raw = ""
    opc = _norm_op_code(opc_raw) if opc_raw else ""

    return (
        chosen.instruction_type.strip(),
        chosen.op_type.strip(),
        opc,
    )


def _norm_op_code(s: str) -> str:
    try:
        return str(int(float(str(s).strip())))
    except (ValueError, TypeError):
        return str(s).strip()


def _norm_exec(s: str) -> str:
    t = str(s).strip()
    if not t:
        return ""
    try:
        n = int(float(t))
        return f"{n:04d}"[-4:]
    except (ValueError, TypeError):
        return t


def lookup_codifier(
    rows: list[CodifierRow],
    op_code_raw: str,
    aux_code_raw: str,
    f_acc: str,
    h_acc: str,
    j_acc: str,
    *,
    flow_hint: str | None = None,
) -> tuple[str, str]:
    """
    Возвращает (тип_поручения, тип_операции).
    Операция: код из MSG + вспомогательное поле; при нескольких строках в кодификаторе —
    различение по коду исполнения (если задан в таблице), затем по **flow_hint** (направление
    по колонке «Тип операции»: подстроки **зачисление** / **списание**), затем по совпадению
    **f_acc** со счётом списания **H** или зачисления **J** (как в ГПБ I*.MSG).
    """
    op = _norm_op_code(op_code_raw)
    if not op:
        return "", ""

    candidates = [r for r in rows if _norm_op_code(r.op_code) == op]
    if not candidates:
        return "", ""

    aux = aux_code_raw.strip()

    def try_exec_match(pool: list[CodifierRow]) -> CodifierRow | None:
        if not aux:
            return None
        a_norm = _norm_exec(aux)
        for r in pool:
            if not r.exec_code.strip():
                continue
            ex = _norm_exec(r.exec_code)
            if ex == a_norm or r.exec_code.strip() == aux:
                return r
        return None

    m = try_exec_match(candidates)
    if m:
        return m.instruction_type, m.op_type

    if len(candidates) == 1:
        r = candidates[0]
        return r.instruction_type, r.op_type

    def matches_flow_label(r: CodifierRow) -> bool:
        ot = (r.op_type or "").lower()
        if flow_hint == "debit":
            return bool(re.search(r"списание", ot))
        if flow_hint == "credit":
            return bool(re.search(r"зачисление", ot))
        return True

    if flow_hint in ("debit", "credit"):
        flow_pool = [r for r in candidates if matches_flow_label(r)]
        if len(flow_pool) == 1:
            r = flow_pool[0]
            return r.instruction_type, r.op_type
        if len(flow_pool) > 1:
            m2 = try_exec_match(flow_pool)
            if m2:
                return m2.instruction_type, m2.op_type
            r0 = flow_pool[0]
            return r0.instruction_type, r0.op_type

    def nf(x: str) -> str:
        try:
            return str(int(float(str(x).strip())))
        except (ValueError, TypeError):
            return str(x).strip()

    fn, hn, jn = nf(f_acc), nf(h_acc), nf(j_acc)

    for r in candidates:
        ot = r.op_type.lower()
        if "зачисление" in ot and fn and jn and fn == jn:
            return r.instruction_type, r.op_type
        if "списание" in ot and fn and hn and fn == hn:
            return r.instruction_type, r.op_type

    r0 = candidates[0]
    return r0.instruction_type, r0.op_type


def _norm_match_snippet(s: str) -> str:
    t = (s or "").replace("\u00a0", " ").replace("\u202f", " ")
    t = t.strip().lower()
    t = t.replace("ё", "е")
    t = re.sub(r"\s+", " ", t)
    return t


# Запас от неполных совпадений: «метка (label) входит в начало ячейки кодификатора», но строки не равны
# (напр. «Снятие ЦБ» vs «Снятие ЦБ – перевод») — такие пары не считаем совпадением.
_CODIFIER_EXACT_MATCH_BONUS = 1_000_000


def _codifier_label_token_score(lab: str, token: str, weight: int) -> int:
    """
    Очки за сопоставление нормализованной метки ``lab`` с нормализованной ячейкой кодификатора ``token``.

    - Полное совпадение (exact match): ``token == lab`` — с бонусом, всегда выше неполного вложения.
    - Неполное: только если **каталог короче или равен** и входит в метку: ``token in lab`` (и ``token != lab`` даёт строго меньший балл).

    Вложение «метка в более длинный текст каталога» без равенства не считается («лишние символы в каталоге»).
    """
    if not lab or not token or weight <= 0:
        return 0
    if token == lab:
        return _CODIFIER_EXACT_MATCH_BONUS + len(token) * weight
    if token in lab:
        return len(token) * weight
    return 0


def lookup_codifier_rsd(
    rows: list[CodifierRow],
    op_code_raw: str,
    aux_code_raw: str,
    description: str,
) -> tuple[str, str]:
    """
    Маппинг для отчёта РСД I*.MSG: код операции + 4-битный код исполнения (0011/1100).
    При нескольких строках (напр. 12214 + 1100) — уточнение по «Содержание РСД» и полю описания в MSG.
    """
    op = _norm_op_code(op_code_raw)
    if not op:
        return "", ""

    candidates = [r for r in rows if _norm_op_code(r.op_code) == op]
    if not candidates:
        return "", ""

    aux = aux_code_raw.strip()
    exec_matched: list[CodifierRow] = []

    if aux:
        a_norm = _norm_exec(aux)
        for r in candidates:
            ex_s = r.exec_code.strip()
            if not ex_s:
                continue
            ex = _norm_exec(ex_s)
            if ex == a_norm or ex_s == aux:
                exec_matched.append(r)

    if not exec_matched:
        for r in candidates:
            if not r.exec_code.strip():
                exec_matched.append(r)
    if not exec_matched:
        exec_matched = list(candidates)

    if len(exec_matched) == 1:
        r0 = exec_matched[0]
        return r0.instruction_type, r0.op_type

    desc_n = _norm_match_snippet(description)
    best: CodifierRow | None = None
    best_score = -1

    for r in exec_matched:
        cr = (r.content_rsd or "").strip()
        if not cr:
            score = 0
        else:
            cr_n = _norm_match_snippet(cr)
            score = 0
            if cr_n and cr_n in desc_n:
                score = len(cr_n)
            elif cr_n and desc_n and (
                desc_n.startswith(cr_n[: min(12, len(cr_n))]) or desc_n.startswith(cr_n)
            ):
                score = max(1, len(cr_n) // 2)
        if score > best_score:
            best_score = score
            best = r

    chosen = best if best is not None and best_score > 0 else exec_matched[0]
    return chosen.instruction_type, chosen.op_type


def _r01_label_flow_hint(label: str) -> str | None:
    """Зачисление / списание по тексту «Тип операции» R-01."""
    t = _norm_match_snippet(label)
    has_z = "зачисление" in t
    has_s = "списание" in t
    if has_z and not has_s:
        return "credit"
    if has_s and not has_z:
        return "debit"
    return None


def _r01_row_matches_flow(row: CodifierRow, flow: str) -> bool:
    blob = _norm_match_snippet(f"{row.content_rsd} {row.op_type}")
    if flow == "credit":
        return "зачисление" in blob
    if flow == "debit":
        return "списание" in blob
    return False


def _disambiguate_r01_codifier_candidates(
    label: str,
    candidates: list[CodifierRow],
) -> CodifierRow:
    if len(candidates) == 1:
        return candidates[0]

    flow = _r01_label_flow_hint(label)
    if flow:
        matched = [r for r in candidates if _r01_row_matches_flow(r, flow)]
        if len(matched) == 1:
            return matched[0]
        if matched:
            candidates = matched
        if flow == "credit":
            for r in candidates:
                if _norm_exec(r.exec_code) == "0011":
                    return r
        elif flow == "debit":
            for r in candidates:
                if _norm_exec(r.exec_code) == "1100":
                    return r

    with_rsd = [r for r in candidates if (r.content_rsd or "").strip()]
    if len(with_rsd) == 1:
        return with_rsd[0]
    if with_rsd:
        candidates = with_rsd

    return candidates[0]


def _score_r01_codifier_row(lab: str, row: CodifierRow) -> int:
    score = 0
    cr = _norm_match_snippet(row.content_rsd)
    ot = _norm_match_snippet(row.op_type)
    cg = _norm_match_snippet(row.content_gpb)

    for token, weight in ((cr, 3), (ot, 2), (cg, 1)):
        if not token:
            continue
        s = _codifier_label_token_score(lab, token, weight)
        if s:
            score = max(score, s)
    return score


def lookup_codifier_r01_operation_type(
    rows: list[CodifierRow],
    label: str,
) -> tuple[str, str, str]:
    """
    Маппинг текста «Тип операции» из R-01 xlsx на строку кодификатора.
    Возвращает (тип_поручения, тип_операции, код_операции).
    Приоритет: «Содержание РСД», затем «Тип операции», «Содержание ГПБ».
    """
    if not (label or "").strip():
        return "", "", ""

    lab = _norm_match_snippet(label)
    if not lab:
        return "", "", ""

    scored: list[tuple[int, CodifierRow]] = []
    best_score = 0

    for r in rows:
        score = _score_r01_codifier_row(lab, r)
        if score > 0:
            scored.append((score, r))
            best_score = max(best_score, score)

    chosen: CodifierRow | None = None
    if scored:
        tied = [r for s, r in scored if s == best_score]
        chosen = _disambiguate_r01_codifier_candidates(label, tied)
    elif "перевод между разделами" in lab:
        rsd_candidates = [
            r
            for r in rows
            if (r.content_rsd or "").strip()
            and "перевод между разделами" in _norm_match_snippet(r.content_rsd)
        ]
        if rsd_candidates:
            chosen = _disambiguate_r01_codifier_candidates(label, rsd_candidates)

    if chosen is None:
        return "", "", ""

    opc_raw = (chosen.op_code or "").strip()
    if _is_dash(opc_raw):
        opc_raw = ""
    opc = _norm_op_code(opc_raw) if opc_raw else ""
    return (
        chosen.instruction_type.strip(),
        chosen.op_type.strip(),
        opc,
    )


def lookup_codifier_region(
    rows: list[CodifierRow],
    operation_type_label: str,
) -> tuple[str, str, str]:
    """
    Маппинг текста «Тип операции» из отчёта REGION IS*.xls на строку кодификатора.
    Возвращает (тип_поручения, тип_операции, код_операции).
    Приоритет: «Содержание РЕГИОН», затем «Содержание РСД», «Тип операции», «Содержание ГПБ».
    Полное совпадение ячейки с меткой важнее вложения метки в более длинный текст в кодификаторе.
    """
    if not (operation_type_label or "").strip():
        return "", "", ""

    lab = _norm_match_snippet(operation_type_label)
    if not lab:
        return "", "", ""

    best: CodifierRow | None = None
    best_score = 0

    for r in rows:
        score = 0
        creg = _norm_match_snippet(r.content_region)
        cr = _norm_match_snippet(r.content_rsd)
        ot = _norm_match_snippet(r.op_type)
        cg = _norm_match_snippet(r.content_gpb)

        for token, weight in ((creg, 4), (cr, 3), (ot, 2), (cg, 1)):
            if not token:
                continue
            s = _codifier_label_token_score(lab, token, weight)
            if s:
                score = max(score, s)

        if score > best_score:
            best_score = score
            best = r

    if best is None or best_score == 0:
        return "", "", ""

    opc_raw = (best.op_code or "").strip()
    if _is_dash(opc_raw):
        opc_raw = ""
    opc = _norm_op_code(opc_raw) if opc_raw else ""
    return (
        best.instruction_type.strip(),
        best.op_type.strip(),
        opc,
    )


def _vtb_temp_withdrawal_pick_12002(
    candidates: list[CodifierRow],
    src_blob: str,
    dst_blob: str,
) -> CodifierRow | None:
    """
    Для кода 12002 в кодификаторе две строки (T зачисление / Y списание).
    Правило: участие раздела «Временное изъятие» со стороны списания или зачисления.
    """
    needle = "временное изъятие"
    s_low = _norm_match_snippet(src_blob)
    d_low = _norm_match_snippet(dst_blob)
    has_s = needle in s_low
    has_d = needle in d_low
    if has_d and not has_s:
        for r in candidates:
            if r.instruction_type.strip().upper() == "T":
                return r
            if "зачисление" in r.op_type.lower() and "списание" not in r.op_type.lower():
                return r
    if has_s and not has_d:
        for r in candidates:
            if r.instruction_type.strip().upper() == "Y":
                return r
            if "списание" in r.op_type.lower() and "зачисление" not in r.op_type.lower():
                return r
    return None


def lookup_codifier_vtb(
    rows: list[CodifierRow],
    stage_name: str,
    *,
    src_sections_blob: str = "",
    dst_sections_blob: str = "",
) -> tuple[str, str, str]:
    """
    Маппинг этапа исполнения ВТБ СД (F752 XML) на строку кодификатора.
    Возвращает (тип_поручения, тип_операции, код_операции).
    Приоритет: «Содержание ВТБСД», затем РСД, тип операции, РЕГИОН, ГПБ.
    Полное совпадение ячейки с меткой важнее вложения метки в более длинный текст каталога.
    Для 12002 — уточнение по «Временное изъятие» в текстах разделов src/dst.
    """
    if not (stage_name or "").strip():
        return "", "", ""

    lab = _norm_match_snippet(stage_name)
    if not lab:
        return "", "", ""

    scored: list[tuple[int, CodifierRow]] = []

    for r in rows:
        score = 0
        cv = _norm_match_snippet(r.content_vtb)
        cr = _norm_match_snippet(r.content_rsd)
        ot = _norm_match_snippet(r.op_type)
        creg = _norm_match_snippet(r.content_region)
        cg = _norm_match_snippet(r.content_gpb)

        for token, weight in ((cv, 5), (cr, 3), (ot, 2), (creg, 2), (cg, 1)):
            if not token:
                continue
            s = _codifier_label_token_score(lab, token, weight)
            if s:
                score = max(score, s)

        if score > 0:
            scored.append((score, r))

    scored.sort(key=lambda x: -x[0])
    if not scored:
        return "", "", ""

    top_score = scored[0][0]
    best_rows = [r for s, r in scored if s == top_score]

    chosen: CodifierRow | None = None
    if len(best_rows) > 1:
        by_op: dict[str, list[CodifierRow]] = {}
        for r in best_rows:
            k = _norm_op_code(r.op_code) if r.op_code.strip() else ""
            by_op.setdefault(k, []).append(r)
        for op_k, group in by_op.items():
            if op_k == "12002" and len(group) >= 2:
                picked = _vtb_temp_withdrawal_pick_12002(
                    group,
                    src_sections_blob,
                    dst_sections_blob,
                )
                if picked:
                    chosen = picked
                    break
        if chosen is None:
            chosen = best_rows[0]
    else:
        chosen = best_rows[0]

    opc_raw = (chosen.op_code or "").strip()
    if _is_dash(opc_raw):
        opc_raw = ""
    opc = _norm_op_code(opc_raw) if opc_raw else ""

    # повторная попытка для 12002 если одна строка с кодом 12002 но не та сторона
    if opc == "12002":
        cands_12002 = [r for r in rows if _norm_op_code(r.op_code) == "12002"]
        if len(cands_12002) >= 2:
            picked = _vtb_temp_withdrawal_pick_12002(
                cands_12002,
                src_sections_blob,
                dst_sections_blob,
            )
            if picked and picked is not chosen:
                # если score выбрал не ту строку — заменить при явном temp withdrawal
                s_low = _norm_match_snippet(src_sections_blob)
                d_low = _norm_match_snippet(dst_sections_blob)
                if ("временное изъятие" in s_low) ^ ("временное изъятие" in d_low):
                    chosen = picked
                    opc = _norm_op_code(chosen.op_code)

    return (
        chosen.instruction_type.strip(),
        chosen.op_type.strip(),
        opc,
    )


def normalize_svod_document_op_type_display(op_type: str) -> str:
    """
    Приводит «Тип операции» к виду как на эталонном скрине Excel:
    «Документы - зачисление» (пробел — дефис — пробел), вместо длинного тире «—».
    """
    t = (op_type or "").strip()
    if not t:
        return t
    t = t.replace("—", "-").replace("–", "-")
    t = re.sub(r"\s*-\s*", " - ", t)
    return t.strip()


def lookup_codifier_fields_by_instruction_type(
    rows: list[CodifierRow],
    instruction_type_raw: str,
) -> tuple[str, str]:
    """«Код операции» и «Тип операции» по «Тип поручения» (лист «Коды»)."""
    inst = instruction_type_raw.strip().upper()
    if not inst:
        return "", ""
    op_code = ""
    op_type = ""
    for r in rows:
        if r.instruction_type.strip().upper() != inst:
            continue
        if r.op_code.strip() and not op_code:
            op_code = r.op_code.strip()
        if r.op_type.strip() and not op_type:
            op_type = r.op_type.strip()
        if op_code and op_type:
            break
    return op_code, op_type


def lookup_document_op_type_by_instruction(
    rows: list[CodifierRow],
    instruction_type_raw: str,
) -> str:
    """
    Для строк СВОД из отчёта D: тип поручения (instruction type) DF/DP и пустой код операции
    в строке кодификатора «Документы …». Возвращает нормализованный «Тип операции» или пусто.
    """
    inst = instruction_type_raw.strip().upper()
    if inst not in ("DF", "DP"):
        return ""

    for r in rows:
        if r.op_code.strip():
            continue
        if r.instruction_type.strip().upper() != inst:
            continue
        ot = r.op_type.strip()
        if not ot:
            continue
        if "документ" not in ot.lower():
            continue
        return normalize_svod_document_op_type_display(ot)

    return ""


def fallback_document_op_type(instruction_type_raw: str) -> str:
    """Если в «Коды» нет строки для DF/DP-документов — эталонные подписи колонки D."""
    inst = instruction_type_raw.strip().upper()
    if inst == "DF":
        return "Документы - зачисление"
    if inst == "DP":
        return "Документы - списание"
    return ""
