"""Создание книги и запись листа Dks (10 колонок — как эталон Ord_Quantity)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from gpb_d_msg import DMsgDataRow, DMsgFile, display_doc_date, instruction_type_df_dp


# Лист Dks: см. memory/Dks_sheet_spec.md
DKS_HEADERS = [
    "Дата операции",
    "Счет депо",
    "НГРИ",
    "Наименование документа",
    "Номер документа",
    "Дата документа",
    "Комментарий",
    "Для удаления дубликатов",
    "Тип поручения",
    "Источник",
]

SCHEETA_DEPO_HEADERS = [
    "Раздел консолидированного портфеля",
    "Депозитарий",
    "Номер счета депо",
    "Оригинатор",
    "Агент по сопровождению",
    "Сделка ДОМ.РФ ИА",
    "Комментарий",
    "Код выписки ГПБ (ДОМ)",
    "Код выписки ГПБ (БАНК)",
    "Владелец счета депо",
]


def duplicate_key(depo_account: str, ngri: str, operation_date: str) -> str:
    """Ключ для колонки H: счёт_НГРИ_дата (подчёркивания как на эталонном скрине)."""
    return f"{depo_account.strip()}_{ngri.strip()}_{operation_date.strip()}"


def row_values(
    parsed: DMsgFile,
    data: DMsgDataRow,
    depo_account: str | None,
) -> list:
    """
    Одна строка листа Dks.
    depo_account — «Номер счета депо» из справочника по ClientID, иначе пусто.
    """
    h = parsed.header
    op_date = (h.get("ReportDate") or "").strip()
    acc = (depo_account or "").strip()
    ngri = data.ngri.strip()
    doc_name = data.doc_type.strip()
    doc_num = data.doc_num.strip()
    doc_date = display_doc_date(data.doc_date)
    comment = data.comment.strip()

    dup = duplicate_key(acc, ngri, op_date)
    inst = instruction_type_df_dp(acc)

    return [
        op_date,
        acc,
        ngri,
        doc_name,
        doc_num,
        doc_date,
        comment,
        dup,
        inst,
        parsed.path.name,
    ]


def create_workbook_template(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    default.title = "Dks"
    for col, title in enumerate(DKS_HEADERS, start=1):
        default.cell(row=1, column=col, value=title)

    ws_dep = wb.create_sheet("Счета депо")
    for col, title in enumerate(SCHEETA_DEPO_HEADERS, start=1):
        ws_dep.cell(row=1, column=col, value=title)

    wb.save(path)


def clear_dks_data(ws) -> None:
    max_r = ws.max_row or 1
    if max_r <= 1:
        return
    for r in range(2, max_r + 1):
        for c in range(1, len(DKS_HEADERS) + 1):
            ws.cell(row=r, column=c, value=None)


def ensure_dks_sheet(wb) -> None:
    if "Dks" not in wb.sheetnames:
        wb.create_sheet("Dks")
    ws = wb["Dks"]
    for col, title in enumerate(DKS_HEADERS, start=1):
        ws.cell(row=1, column=col, value=title)


def write_dks_rows(path: Path, rows_out: list[list]) -> None:
    wb = load_workbook(path)
    ensure_dks_sheet(wb)
    ws = wb["Dks"]
    clear_dks_data(ws)
    for i, row in enumerate(rows_out, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    wb.save(path)
