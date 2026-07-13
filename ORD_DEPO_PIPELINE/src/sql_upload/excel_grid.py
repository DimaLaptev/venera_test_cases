"""Чтение Excel-книг как сеток строк (raw grid) для SQL-загрузки."""

from __future__ import annotations

from pathlib import Path
from typing import Iterator

from gpb_rep_xls import _cell_str, _read_rows_calamine_best_sheet, _read_rows_openpyxl_best_sheet


def cell_at(rows: list[list[str]], row_1based: int, col_1based: int) -> str:
    """Ячейка по 1-based индексам (как VBA Cells)."""
    r = row_1based - 1
    c = col_1based - 1
    if r < 0 or r >= len(rows):
        return ""
    row = rows[r]
    if c < 0 or c >= len(row):
        return ""
    return row[c]


def last_row_in_column(rows: list[list[str]], col_1based: int) -> int:
    """Последняя непустая строка в колонке (1-based), минимум 0."""
    c = col_1based - 1
    last = 0
    for r_idx, row in enumerate(rows):
        if c < len(row) and (row[c] or "").strip():
            last = r_idx + 1
    return last


def _read_xlrd_all_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        import xlrd
    except ModuleNotFoundError:
        return []

    book = xlrd.open_workbook(str(path), formatting_info=False)
    out: list[tuple[str, list[list[str]]]] = []
    for si in range(book.nsheets):
        sh = book.sheet_by_index(si)
        mat: list[list[str]] = []
        for r in range(sh.nrows):
            row: list[str] = []
            for c in range(sh.ncols):
                v = sh.cell(r, c).value
                if v is None or v == "":
                    row.append("")
                elif isinstance(v, float) and v == int(v):
                    row.append(str(int(v)))
                else:
                    row.append(_cell_str(v))
            mat.append(row)
        out.append((sh.name, mat))
    return out


def _read_openpyxl_all_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[list[str]]]] = []
    try:
        for ws in wb.worksheets:
            mat: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                mat.append([_cell_str(x) for x in row])
            out.append((ws.title, mat))
    finally:
        wb.close()
    return out


def _read_calamine_all_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    try:
        from python_calamine import CalamineWorkbook
    except ModuleNotFoundError:
        return []

    try:
        wb = CalamineWorkbook.from_path(path)
    except Exception:
        try:
            wb = CalamineWorkbook.from_object(path)
        except Exception:
            return []

    names = getattr(wb, "sheet_names", None) or []
    out: list[tuple[str, list[list[str]]]] = []
    for i in range(len(names)):
        try:
            sh = wb.get_sheet_by_index(i)
            raw = sh.to_python(skip_empty_area=False)
            mat = [[_cell_str(c) for c in row] for row in raw]
            out.append((names[i], mat))
        except Exception:
            continue
    return out


def read_workbook_sheets(path: Path) -> list[tuple[str, list[list[str]]]]:
    """Все листы книги; fallback calamine → openpyxl → xlrd → best single sheet."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xlsb"):
        sheets = _read_openpyxl_all_sheets(path)
        if sheets:
            return sheets
        sheets = _read_calamine_all_sheets(path)
        if sheets:
            return sheets

    if suffix == ".xls":
        sheets = _read_calamine_all_sheets(path)
        if sheets:
            return sheets
        sheets = _read_xlrd_all_sheets(path)
        if sheets:
            return sheets
        rows = _read_rows_calamine_best_sheet(path)
        if rows:
            return [("Sheet1", rows)]

    if suffix in (".xlsx", ".xlsm"):
        rows = _read_rows_openpyxl_best_sheet(path)
        if rows:
            return [("Sheet1", rows)]

    rows = _read_rows_calamine_best_sheet(path)
    if rows:
        return [("Sheet1", rows)]
    return []


def iter_files_in_dir(directory: Path) -> Iterator[Path]:
    if not directory.is_dir():
        return
    for entry in sorted(directory.rglob("*")):
        if entry.is_file():
            yield entry
