"""Кейс как на скринах ВТБ: XML Qty=3/38 + «Приложение 33» → N строк СВОД с НГРИ."""

from __future__ import annotations

import sys
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PIPELINE_ROOT / "src"))

from vtb_f752_xml import (  # noqa: E402
    collect_vtb_ngri_for_xml,
    find_vtb_appendix_for_xml,
    list_ngri_from_vtb_appendix_path,
)


def _write_minimal_f752_xml(path: Path, *, qty: int, order: str) -> None:
    path.write_text(
        f"""<?xml version="1.0" encoding="utf-8"?>
<EDO_ODKF752_1_F753_1>
  <OrderNumber>{order}</OrderNumber>
  <OrderDate>06.05.2026</OrderDate>
  <ExecutionDate>06.05.2026</ExecutionDate>
  <Basis>Договор счета депо №3595/ОДК_12830 от 06.05.2026</Basis>
  <Qty>{qty}</Qty>
  <SrcAccountCode>12830</SrcAccountCode>
  <SrcSectionName>Свободное обращение</SrcSectionName>
  <SrcSPSectionCode>100000</SrcSPSectionCode>
  <DstAccountCode/>
  <ClientCashDataTable>
    <StageName>Блокировка закладных под погашение</StageName>
    <ODKDate>06.05.2026</ODKDate>
    <OverlayDepositaryDate/>
  </ClientCashDataTable>
  <ClientCashDataTable>
    <StageName>Списание ценных бумаг со счёта депо</StageName>
    <ODKDate>06.05.2026</ODKDate>
    <OverlayDepositaryDate>06.05.2026</OverlayDepositaryDate>
  </ClientCashDataTable>
</EDO_ODKF752_1_F753_1>
""",
        encoding="utf-8",
    )


def _write_appendix_xlsx(path: Path, ngris: list[str]) -> None:
    """Макет «Приложение № 33»: заголовки строка 6, данные с 8, НГРИ — колонка B."""
    wb = Workbook()
    ws = wb.active
    ws["A6"] = "№ п/п"
    ws["B6"] = "Номер государственной регистрации ипотеки"
    for i, ng in enumerate(ngris, start=1):
        row = 7 + i  # 8, 9, …
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=ng)
        ws.cell(row=row, column=8, value="RUB")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _fake_ngris(n: int, prefix: str) -> list[str]:
    return [f"{prefix}:{i:05d}/2021-{i}" for i in range(1, n + 1)]


@pytest.mark.parametrize("count", [3, 38])
def test_collect_ngri_from_same_stem_xlsx(tmp_path: Path, count: int) -> None:
    """Кейс скринов: нет ZIP, есть одноимённый .xlsx «Приложение» → count НГРИ."""
    stem = f"VTB00001-003-test{count}F752-1"
    xml = tmp_path / "XML" / f"{stem}.xml"
    apps = tmp_path / "XML_приложения"
    xml.parent.mkdir(parents=True)
    _write_minimal_f752_xml(xml, qty=count, order=f"3595/{count}")
    ngris = _fake_ngris(count, "50:42")
    _write_appendix_xlsx(apps / f"{stem}.xlsx", ngris)

    found = find_vtb_appendix_for_xml(xml, apps)
    assert found is not None
    assert found.suffix.lower() == ".xlsx"

    got = collect_vtb_ngri_for_xml(xml, apps)
    assert len(got) == count
    assert got == ngris


@pytest.mark.parametrize("count", [3, 38])
def test_collect_ngri_from_zip_with_inner_xlsx(tmp_path: Path, count: int) -> None:
    """Типовой путь: ZIP с тем же stem, внутри «Приложение 33.xlsx»."""
    stem = f"VTB00001-003-zip{count}F752-1"
    xml = tmp_path / "XML" / f"{stem}.xml"
    apps = tmp_path / "XML_приложения"
    xml.parent.mkdir(parents=True)
    apps.mkdir(parents=True)
    _write_minimal_f752_xml(xml, qty=count, order=f"3596/{count}")
    ngris = _fake_ngris(count, "42:04")
    inner = tmp_path / "Приложение 33.xlsx"
    _write_appendix_xlsx(inner, ngris)
    zpath = apps / f"{stem}.zip"
    with zipfile.ZipFile(zpath, "w") as zf:
        zf.write(inner, arcname="Приложение 33.xlsx")

    got = collect_vtb_ngri_for_xml(xml, apps)
    assert len(got) == count
    assert got == ngris


def test_ooxml_bytes_named_xls_still_reads_ngri(tmp_path: Path) -> None:
    """Файл с расширением .xls, содержимое OOXML — читаем через openpyxl."""
    stem = "VTB00001-003-ooxmlF752-1"
    xml = tmp_path / f"{stem}.xml"
    apps = tmp_path / "apps"
    _write_minimal_f752_xml(xml, qty=3, order="1/2")
    ngris = _fake_ngris(3, "77:17")
    xlsx_tmp = tmp_path / "tmp.xlsx"
    _write_appendix_xlsx(xlsx_tmp, ngris)
    # Переименовываем содержимое OOXML в .xls (как иногда приходит от ВТБ).
    fake_xls = apps / f"{stem}.xls"
    apps.mkdir(parents=True)
    fake_xls.write_bytes(xlsx_tmp.read_bytes())

    got = list_ngri_from_vtb_appendix_path(fake_xls)
    assert got == ngris
    assert collect_vtb_ngri_for_xml(xml, apps) == ngris


def test_no_appendix_yields_empty_list(tmp_path: Path) -> None:
    stem = "VTB00001-003-lonelyF752-1"
    xml = tmp_path / f"{stem}.xml"
    apps = tmp_path / "apps"
    apps.mkdir()
    _write_minimal_f752_xml(xml, qty=38, order="9/9")
    assert collect_vtb_ngri_for_xml(xml, apps) == []


def test_list_missing_attachments_reports_xml_without_pair(tmp_path: Path) -> None:
    from vtb_f752_xml import (
        format_vtb_missing_attachments_error,
        list_vtb_xml_missing_attachments,
    )

    apps = tmp_path / "XML_приложения"
    apps.mkdir()
    xml_ok = tmp_path / "VTB_okF752-1.xml"
    xml_bad = tmp_path / "VTB_badF752-1.xml"
    _write_minimal_f752_xml(xml_ok, qty=1, order="1/1")
    _write_minimal_f752_xml(xml_bad, qty=1, order="2/2")
    _write_appendix_xlsx(apps / f"{xml_ok.stem}.xlsx", _fake_ngris(1, "11:11"))

    missing = list_vtb_xml_missing_attachments([xml_ok, xml_bad], apps)
    assert missing == [xml_bad]
    text = format_vtb_missing_attachments_error(missing, apps)
    assert "VTB_badF752-1.xml" in text
    assert "Сборка остановлена" in text
    assert "VTB_okF752-1.xml" not in text