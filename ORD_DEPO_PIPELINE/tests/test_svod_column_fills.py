"""Заливка колонок СВОД после записи данных (шаблон F/G/L/N и т.п.)."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PIPELINE_ROOT / "src"))

from svod_excel import SVOD_HEADERS, write_svod_rows  # noqa: E402

_HIGHLIGHT = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
# Как на скрине Ord_Quantity: F, G, L, N
_HIGHLIGHT_COLS = (6, 7, 12, 14)


def _rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or not fill.fill_type or fill.fill_type == "none":
        return None
    color = fill.start_color
    rgb = getattr(color, "rgb", None)
    if rgb is None:
        return None
    return str(rgb).upper().lstrip("0") if str(rgb).upper().startswith("00") else str(rgb).upper()


def test_write_svod_rows_copies_header_column_fills(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "СВОД_операций"
    for col, title in enumerate(SVOD_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        if col in _HIGHLIGHT_COLS:
            cell.fill = _HIGHLIGHT
    # Старая «хвостовая» строка с заливкой — должна очиститься, если данных меньше
    for c in range(1, len(SVOD_HEADERS) + 1):
        old = ws.cell(row=5, column=c, value="old")
        if c in _HIGHLIGHT_COLS:
            old.fill = _HIGHLIGHT
    wb.save(path)

    row = [f"v{c}" for c in range(1, len(SVOD_HEADERS) + 1)]
    write_svod_rows(path, [row, list(row)])

    ws2 = load_workbook(path)["СВОД_операций"]
    for c in _HIGHLIGHT_COLS:
        assert _rgb(ws2.cell(1, c)) in ("DDEBF7", "00DDEBF7")
        assert _rgb(ws2.cell(2, c)) in ("DDEBF7", "00DDEBF7")
        assert _rgb(ws2.cell(3, c)) in ("DDEBF7", "00DDEBF7")
    # Неподсвеченная колонка A остаётся без заливки
    assert _rgb(ws2.cell(2, 1)) is None
    # Хвост после данных очищен
    assert ws2.cell(5, 6).value is None
    assert _rgb(ws2.cell(5, 6)) is None
