"""Запись СВОД_пoДЕПО.xlsx и СВОД_пoДЕПО_ДОМ.xlsx (новая книга, без шаблона)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from models import (
    SVOD_DOM_SHEETS,
    SVOD_HEADERS,
    SVOD_IA_SHEETS,
    SvodRow,
)
from date_format import report_month_last_day
from references.spravochnik import Sprav1Lookup
from validate.accounts import (
    VALIDATION_HEADERS,
    VALIDATION_SHEET_NAME,
    ValidationRow,
    build_validation_rows,
)
from validate.duplicates import (
    DUPLICATES_NGRI_HEADER,
    DUPLICATES_PORTFOLIO_HEADER,
    DUPLICATES_SHEET_NAME,
    WorkbookDuplicates,
    find_workbook_duplicates,
)

HEADER_ROW = 5
DATA_START_ROW = 6
DEFAULT_VIEW_ZOOM = 70
SVOD_SHEET_TITLE = (
    "СВОДНЫЙ ДЕПОЗИТАРНЫЙ ОТЧЕТ (ДОМ.РФ, инвесторы, собственный портфель ИА) // "
    "СВОДНЫЙ ДЕПОЗИТАРНЫЙ ОТЧЕТ (ДОМ.РФ ИПОТЕЧНЫЙ АГЕНТ)"
)

HEADER_FILL = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
VALIDATION_HEADER_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
VALIDATION_HEADER_FONT = Font(color="FFFFFF", bold=True)
HIGHLIGHT_COL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HIGHLIGHT_COLUMNS = (1, 2, 5, 6, 11)  # A, B, E, F, K


def _looks_like_formula(value: str) -> bool:
    """Строки с «=» Excel/openpyxl интерпретируют как формулы."""
    return bool(value) and value.lstrip().startswith("=")


def _set_cell_value(ws: Worksheet, row: int, column: int, value: object) -> None:
    cell = ws.cell(row=row, column=column)
    if isinstance(value, str) and _looks_like_formula(value):
        cell.value = value
        cell.data_type = "s"
    else:
        cell.value = value


def _apply_sheet_view(ws: Worksheet) -> None:
    ws.sheet_view.zoomScale = DEFAULT_VIEW_ZOOM


def _apply_sheet_styles(ws: Worksheet, last_data_row: int) -> None:
    for c in range(1, len(SVOD_HEADERS) + 1):
        cell = ws.cell(row=HEADER_ROW, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    if last_data_row < DATA_START_ROW:
        return
    for col in HIGHLIGHT_COLUMNS:
        for r in range(DATA_START_ROW, last_data_row + 1):
            ws.cell(row=r, column=col).fill = HIGHLIGHT_COL_FILL


def _write_data_rows(ws: Worksheet, rows: list[SvodRow], start_row: int = DATA_START_ROW) -> int:
    r = start_row
    for row in rows:
        for c, val in enumerate(row.as_tuple(), start=1):
            _set_cell_value(ws, r, c, val)
        r += 1
    return r - start_row


def _init_sheet(ws: Worksheet, report_date: str, row_count: int) -> None:
    _set_cell_value(ws, 1, 1, report_month_last_day(report_date))
    ws.cell(row=1, column=3, value=row_count)
    _set_cell_value(ws, 1, 6, SVOD_SHEET_TITLE)
    for c, title in enumerate(SVOD_HEADERS, start=1):
        _set_cell_value(ws, HEADER_ROW, c, title)


def _write_validation_sheet(ws: Worksheet, rows: list[ValidationRow]) -> None:
    for c, title in enumerate(VALIDATION_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill = VALIDATION_HEADER_FILL
        cell.font = VALIDATION_HEADER_FONT

    for r, row in enumerate(rows, start=2):
        _set_cell_value(ws, r, 1, row.razdel_kp)
        _set_cell_value(ws, r, 2, row.depository)
        _set_cell_value(ws, r, 3, row.schet_depo)
        ws.cell(row=r, column=4, value=row.count)

    last_row = max(len(rows) + 1, 1)
    ws.auto_filter.ref = f"A1:D{last_row}"


def _write_duplicates_sheet(ws: Worksheet, duplicates: WorkbookDuplicates) -> None:
    _set_cell_value(ws, 1, 1, DUPLICATES_NGRI_HEADER)
    _set_cell_value(ws, 1, 2, DUPLICATES_PORTFOLIO_HEADER)
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.cell(row=1, column=2).font = HEADER_FONT

    for i, value in enumerate(duplicates.ngri, start=2):
        _set_cell_value(ws, i, 1, value)
    for i, value in enumerate(duplicates.portfolio, start=2):
        _set_cell_value(ws, i, 2, value)


def _write_workbook(
    output_path: Path,
    rows_by_sheet: dict[str, list[SvodRow]],
    sheet_order: tuple[str, ...],
    report_date: str,
    sprav1: Sprav1Lookup,
) -> WorkbookDuplicates:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name in sheet_order:
        ws = wb.create_sheet(sheet_name)
        rows = rows_by_sheet.get(sheet_name, [])
        _init_sheet(ws, report_date, len(rows))
        written = _write_data_rows(ws, rows)
        last_row = DATA_START_ROW + written - 1 if written else DATA_START_ROW - 1
        _apply_sheet_styles(ws, last_row)

    validation_rows = build_validation_rows(sprav1, rows_by_sheet, sheet_order)
    ws_val = wb.create_sheet(VALIDATION_SHEET_NAME)
    _write_validation_sheet(ws_val, validation_rows)

    duplicates = find_workbook_duplicates(rows_by_sheet, sheet_order)
    if duplicates.has_any:
        ws_dup = wb.create_sheet(DUPLICATES_SHEET_NAME)
        _write_duplicates_sheet(ws_dup, duplicates)

    for ws in wb.worksheets:
        _apply_sheet_view(ws)

    wb.save(output_path)
    wb.close()
    return duplicates


def write_svod_workbooks(
    svod_path: Path,
    svod_dom_path: Path,
    rows_by_sheet: dict[str, list[SvodRow]],
    report_date: str,
    sprav1: Sprav1Lookup,
) -> tuple[Path, Path, WorkbookDuplicates, WorkbookDuplicates]:
    ia_rows = {name: rows_by_sheet.get(name, []) for name in SVOD_IA_SHEETS}
    dom_rows = {name: rows_by_sheet.get(name, []) for name in SVOD_DOM_SHEETS}
    ia_dups = _write_workbook(svod_path, ia_rows, SVOD_IA_SHEETS, report_date, sprav1)
    dom_dups = _write_workbook(svod_dom_path, dom_rows, SVOD_DOM_SHEETS, report_date, sprav1)
    return svod_path, svod_dom_path, ia_dups, dom_dups
