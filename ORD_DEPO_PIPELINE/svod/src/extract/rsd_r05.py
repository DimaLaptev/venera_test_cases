"""Парсинг выписки РСД R-05 (ВЫПИСКА ПО СЧЕТУ ДЕПО) из RSD_EXL."""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from extract.rsd_section import extract_section_prefix
from models import DepoReportRow, SourceKind

_STATEMENT_MARKER = "выписка по счету депо"
_DEPO_ACCOUNT_RE = re.compile(r"^(\d{9}|\d{11,14})$")
_FOOTER_MARKERS = (
    "исполнитель:",
    "ответственное лицо",
    "ответственное лицo",
)


def _cell_value_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value).strip()


def _norm_scan(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _is_r05_statement(ws: Worksheet) -> bool:
    max_r = min(35, (ws.max_row or 1) + 5)
    for r in range(1, max_r + 1):
        for c in range(1, min(12, (ws.max_column or 12) + 1)):
            if _STATEMENT_MARKER in _norm_scan(_cell_value_str(ws.cell(row=r, column=c).value)):
                return True
            if "выписка по счету депo" in _norm_scan(_cell_value_str(ws.cell(row=r, column=c).value)):
                return True
    return False


def _find_depo_account(ws: Worksheet) -> str:
    for r in range(1, min(55, (ws.max_row or 55) + 1)):
        for c in range(1, min(14, (ws.max_column or 14) + 1)):
            raw = _cell_value_str(ws.cell(row=r, column=c).value)
            compact = raw.replace(" ", "")
            if not _DEPO_ACCOUNT_RE.match(compact):
                continue
            ctx = ""
            for dr in (-2, -1, 0, 1, 2):
                for dc in (-4, -3, -2, -1, 0, 1, 2, 3, 4):
                    rr, cc = r + dr, c + dc
                    if rr < 1 or cc < 1:
                        continue
                    ctx += _norm_scan(_cell_value_str(ws.cell(row=rr, column=cc).value))
            if "депо" in ctx or "счет" in ctx or "счёт" in ctx:
                return compact
    return ""


def _find_holdings_header(ws: Worksheet) -> tuple[int, dict[str, int]]:
    max_r = min(80, (ws.max_row or 1) + 40)
    max_c = min(20, (ws.max_column or 15) + 5)
    for r in range(1, max_r + 1):
        col_map: dict[str, int] = {}
        for c in range(1, max_c + 1):
            h = _norm_scan(_cell_value_str(ws.cell(row=r, column=c).value))
            if not h:
                continue
            if "номер раздела" in h and ("счет" in h or "депо" in h or "депo" in h):
                col_map["section"] = c
            if ("гос" in h and "рег" in h) or "registration" in h:
                col_map["ngri"] = c
        if col_map.get("section") and col_map.get("ngri"):
            return r, col_map
    return 0, {}


def _holdings_row_empty(ws: Worksheet, row: int, cmap: dict[str, int]) -> bool:
    """Строка таблицы остатков пуста по всем ключевым колонкам."""
    for col in cmap.values():
        if _cell_value_str(ws.cell(row=row, column=col).value).strip():
            return False
    return True


def _holdings_row_footer(ws: Worksheet, row: int, max_col: int) -> bool:
    """Подпись / исполнитель под таблицей — не данные остатков."""
    parts: list[str] = []
    for c in range(1, max_col + 1):
        parts.append(_cell_value_str(ws.cell(row=row, column=c).value))
    row_text = _norm_scan(" ".join(parts))
    return any(marker in row_text for marker in _FOOTER_MARKERS)


def _is_footer_cell(value: str) -> bool:
    text = _norm_scan(value)
    return any(marker in text for marker in _FOOTER_MARKERS)


def parse_r05_statement_xlsx(path: Path) -> list[DepoReportRow]:
    path = path.resolve()
    wb = None
    try:
        wb = load_workbook(path, read_only=False, data_only=True)
    except OSError:
        return []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not _is_r05_statement(ws):
                continue
            depo = _find_depo_account(ws)
            hr, cmap = _find_holdings_header(ws)
            if not hr or not cmap:
                continue
            c_sec = cmap["section"]
            c_ngri = cmap["ngri"]
            max_col = max(cmap.values())
            out: list[DepoReportRow] = []
            last_row = ws.max_row or hr
            for r in range(hr + 1, last_row + 1):
                if _holdings_row_empty(ws, r, cmap):
                    break
                if _holdings_row_footer(ws, r, max_col):
                    break

                section_raw = _cell_value_str(ws.cell(row=r, column=c_sec).value)
                ngri = _cell_value_str(ws.cell(row=r, column=c_ngri).value)
                if _is_footer_cell(ngri) or _is_footer_cell(section_raw):
                    break
                if not ngri.strip():
                    continue
                out.append(
                    DepoReportRow(
                        source_kind=SourceKind.RSD_R05,
                        source_file=path.name,
                        schet_depo=depo,
                        razdel_scheta=extract_section_prefix(section_raw),
                        ngri_v_depo=ngri.strip(),
                        status="На хранении",
                    ),
                )
            if out:
                return out
        return []
    finally:
        if wb is not None:
            wb.close()


def iter_rsd_r05_rows(directory: Path) -> list[DepoReportRow]:
    if not directory.is_dir():
        return []
    out: list[DepoReportRow] = []
    seen_paths: set[str] = set()
    for path in sorted(directory.rglob("*")):
        if not path.is_file() or path.suffix.lower() != ".xlsx":
            continue
        if path.name.startswith("~"):
            continue
        path_key = str(path.resolve()).casefold()
        if path_key in seen_paths:
            continue
        seen_paths.add(path_key)
        out.extend(parse_r05_statement_xlsx(path))
    return out
