"""Тесты парсинга REGION формат 1/2."""

from __future__ import annotations

import sys
from pathlib import Path

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = PIPELINE_ROOT.parent
sys.path.insert(0, str(PIPELINE_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from extract.region import iter_region_rows
from extract.region_formats import (
    account_from_bracket_filename,
    apply_status_from_format2,
    build_status_by_ngri,
    is_vrsn_lkk_path,
    merge_status_by_ngri,
    parse_depoaccount_balance_xml,
    parse_format1_xml,
    parse_format2_vl_xls,
    parse_vrsn_lkk_xlsx,
    schet_from_vrsn_lkk_filename,
)

_SS_URI = "urn:schemas-microsoft-com:office:spreadsheet"
from models import DepoReportRow, SourceKind


def test_account_from_bracket_filename() -> None:
    p = Path("[VL0032][][IS26040130930][][31.03.2026].xml")
    assert account_from_bracket_filename(p) == "VL0032"


def test_parse_format1_grid() -> None:
    header = [
        "DepoAccount",
        "DivisionName",
        "AgreementNumber",
        "RegistrationNumber",
        "FIO",
    ]
    data = [
        "VL0032",
        "4-02-82084-H",
        "1573",
        "45:25:070402:1407-45/016/2018-3",
        "Дыгерн Наталья Сергеевна",
    ]

    path = Path("[VL0032][][IS26040130930][][31.03.2026].xml")
    rows = parse_format1_xml(path, [header, data])
    assert len(rows) == 1
    assert rows[0].schet_depo == "VL0032"
    assert rows[0].razdel_scheta == "4-02-82084-H"
    assert rows[0].ngri_v_depo == "45:25:070402:1407-45/016/2018-3"
    assert rows[0].nomer_kd == "1573"
    assert rows[0].fio == "Дыгерн Наталья Сергеевна"
    assert rows[0].status is None


def test_parse_format2_grid() -> None:
    grid: list[list[str]] = [[""] * 8 for _ in range(6)]
    grid.append(
        [
            "Состояние",
            "НГРИ",
            "Дата государственной регистрации",
            "Дата заключения договора",
            "Номер кредитного договора",
            "Сумма кредита",
            "Депонент",
            "Раздел счета ДЕПО",
        ],
    )
    grid.append(
        [
            "На хранении",
            "45:25:070402:1407-45/016/2018-3",
            "27.05.2018",
            "21.05.2018",
            "623/2302-0003037",
            "1700000",
            "Найверт Николай",
            "VL0089OB01 DC0001OB02",
        ],
    )
    path = Path("VL0089.XLS")
    rows = parse_format2_vl_xls(path, grid)
    assert len(rows) == 1
    assert rows[0].schet_depo == "VL0089"
    assert rows[0].status == "На хранении"
    assert rows[0].ngri_v_depo == "45:25:070402:1407-45/016/2018-3"
    assert rows[0].nomer_kd == "623/2302-0003037"


def test_parse_vrsn_lkk_xlsx() -> None:
    header = [
        "",
        "",
        "",
        "Кадастровый номер",
        "Номер КД",
        "Дата КД",
        "Номер ГРИ",
        "Дата ГРИ",
        "Статус",
    ]
    data = [
        "",
        "",
        "",
        "50:01:001:1",
        "KD-1",
        "01.01.2024",
        "77:01:000111:1-77/001/2020-1",
        "01.02.2020",
        "В хранилище",
    ]
    path = Path("BpCH_VL0160_1.xlsx")
    assert schet_from_vrsn_lkk_filename(path) == "VL0160"
    assert is_vrsn_lkk_path(path)
    rows = parse_vrsn_lkk_xlsx(path, [header, data])
    assert len(rows) == 1
    assert rows[0].schet_depo == "VL0160"
    assert rows[0].ngri_v_depo == "77:01:000111:1-77/001/2020-1"
    assert rows[0].status == "В хранилище"


def test_merge_status_from_vrsn_lkk_fills_gaps() -> None:
    f2 = [
        DepoReportRow(
            source_kind=SourceKind.REGION,
            source_file="VL0089.XLS",
            schet_depo="VL0089",
            razdel_scheta="x",
            ngri_v_depo="45:25:070402:1407-45/016/2018-3",
            status="На хранении",
        ),
    ]
    vrsn = [
        DepoReportRow(
            source_kind=SourceKind.REGION,
            source_file="BpCH_VL0160_1.xlsx",
            schet_depo="VL0160",
            razdel_scheta="",
            ngri_v_depo="77:01:000111:1-77/001/2020-1",
            status="В хранилище",
        ),
    ]
    xml_row = [
        DepoReportRow(
            source_kind=SourceKind.REGION,
            source_file="[VL0160][].xml",
            schet_depo="VL0160",
            razdel_scheta="4-02",
            ngri_v_depo="77:01:000111:1-77/001/2020-1",
            status=None,
        ),
    ]
    status_map = merge_status_by_ngri(build_status_by_ngri(f2), vrsn)
    assert status_map["45:25:070402:1407-45/016/2018-3"] == "На хранении"
    assert status_map["77:01:000111:1-77/001/2020-1"] == "В хранилище"
    apply_status_from_format2(xml_row, status_map)
    assert xml_row[0].status == "В хранилище"


def test_status_merge_from_format2() -> None:
    f2 = [
        DepoReportRow(
            source_kind=SourceKind.REGION,
            source_file="VL0089.XLS",
            schet_depo="VL0089",
            razdel_scheta="x",
            ngri_v_depo="45:25:070402:1407-45/016/2018-3",
            status="На хранении",
        ),
    ]
    f1 = [
        DepoReportRow(
            source_kind=SourceKind.REGION,
            source_file="[VL0032][].xml",
            schet_depo="VL0032",
            razdel_scheta="4-02",
            ngri_v_depo="45:25:070402:1407-45/016/2018-3",
            status=None,
        ),
    ]
    status_map = build_status_by_ngri(f2)
    apply_status_from_format2(f1, status_map)
    assert f1[0].status == "На хранении"


def test_parse_depoaccount_balance_xml() -> None:
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<DEPOACCOUNT_BALANCE_REPORT>
  <ReportHeader>
    <DepoAccount>
      <DepoAccount>10224</DepoAccount>
    </DepoAccount>
  </ReportHeader>
  <ReportBody>
    <PositionList>
      <Position>
        <PositionInfo>
          <Mortgage>
            <RegistrationNumber>50:28:0050105:8424-50/018/2024-2</RegistrationNumber>
            <AgreementNumber>0040-00767/ИКР-23РБ</AgreementNumber>
          </Mortgage>
        </PositionInfo>
        <DivisionList>
          <Division>
            <DivisionInfo>
              <DivisionName>4B02-60-00307-R-001P от 19.12.2025</DivisionName>
              <DivisionCode>1022414039000041002</DivisionCode>
            </DivisionInfo>
          </Division>
        </DivisionList>
      </Position>
      <Position>
        <PositionInfo>
          <Mortgage>
            <RegistrationNumber>66:41:0604021:509-66/199/2023-2</RegistrationNumber>
            <AgreementNumber>0040-00755/ИКР-23РБ</AgreementNumber>
          </Mortgage>
        </PositionInfo>
        <DivisionList>
          <Division>
            <DivisionInfo>
              <DivisionName>4B02-60-00307-R-001P от 19.12.2025</DivisionName>
            </DivisionInfo>
          </Division>
        </DivisionList>
      </Position>
    </PositionList>
  </ReportBody>
</DEPOACCOUNT_BALANCE_REPORT>"""
    path = Path("[10224][][IS26050516750][][31.03.2026].xml")
    root = _read_xml_root_from_text(xml)
    rows = parse_depoaccount_balance_xml(path, root)
    assert len(rows) == 2
    assert rows[0].schet_depo == "10224"
    assert rows[0].ngri_v_depo == "50:28:0050105:8424-50/018/2024-2"
    assert rows[0].nomer_kd == "0040-00767/ИКР-23РБ"
    assert "4B02-60-00307-R-001P" in rows[0].razdel_scheta
    assert rows[1].ngri_v_depo == "66:41:0604021:509-66/199/2023-2"


def _read_xml_root_from_text(xml: str):
    import xml.etree.ElementTree as ET

    return ET.fromstring(xml)


def test_iter_region_parses_balance_xml(tmp_path: Path) -> None:
    reg = tmp_path / "REGION"
    reg.mkdir()
    xml = """<?xml version="1.0" encoding="UTF-8"?>
<DEPOACCOUNT_BALANCE_REPORT>
  <ReportHeader><DepoAccount><DepoAccount>10224</DepoAccount></DepoAccount></ReportHeader>
  <ReportBody>
    <PositionList>
      <Position>
        <PositionInfo><Mortgage>
          <RegistrationNumber>50:28:0050105:8424-50/018/2024-2</RegistrationNumber>
          <AgreementNumber>0040-00767/ИКР-23РБ</AgreementNumber>
        </Mortgage></PositionInfo>
        <DivisionList><Division><DivisionInfo>
          <DivisionName>4B02-60-00307-R-001P от 19.12.2025</DivisionName>
        </DivisionInfo></Division></DivisionList>
      </Position>
    </PositionList>
  </ReportBody>
</DEPOACCOUNT_BALANCE_REPORT>"""
    (reg / "[10224][][IS26050516750][][31.03.2026].xml").write_text(xml, encoding="utf-8")

    rows = iter_region_rows(reg)
    assert len(rows) == 1
    assert rows[0].schet_depo == "10224"
    assert rows[0].ngri_v_depo == "50:28:0050105:8424-50/018/2024-2"


def test_read_spreadsheet_ml_utf16_default_namespace(tmp_path: Path) -> None:
    """Region xml: UTF-16 + xmlns без префикса ss: (типичный экспорт Excel)."""
    from extract.region_formats import read_region_sheet_rows

    xml = f"""<?xml version="1.0" encoding="UTF-16"?>
<Workbook xmlns="{_SS_URI}">
<Worksheet><Table>
<Row>
<Cell><Data>DepoAccount</Data></Cell>
<Cell><Data>RegistrationNumber</Data></Cell>
<Cell><Data>DivisionName</Data></Cell>
</Row>
<Row>
<Cell><Data>10224</Data></Cell>
<Cell><Data>50:28:0050105:8424-50/018/2024-2</Data></Cell>
<Cell><Data>4B02-60-00307-R-001P</Data></Cell>
</Row>
</Table></Worksheet></Workbook>"""
    path = tmp_path / "[10224][][IS26050516750][][31.03.2026].xml"
    path.write_text(xml, encoding="utf-16")

    sheet = read_region_sheet_rows(path)
    assert len(sheet) == 2
    rows = parse_format1_xml(path, sheet)
    assert len(rows) == 1
    assert rows[0].schet_depo == "10224"
    assert rows[0].ngri_v_depo == "50:28:0050105:8424-50/018/2024-2"


def test_iter_region_logs_empty_xml(tmp_path: Path) -> None:
    from io import StringIO

    reg = tmp_path / "REGION"
    reg.mkdir()
    (reg / "[10224][][IS26050516750][][31.03.2026].xml").write_text(
        "not spreadsheet ml",
        encoding="utf-8",
    )

    buf = StringIO()
    rows = iter_region_rows(reg, warn=buf)
    assert rows == []
    text = buf.getvalue()
    assert "0 строк из xml" in text
    assert "[10224]" in text
    assert "Region xml без строк: 1 файл(ов)" in text


def test_iter_region_merges_status(tmp_path: Path) -> None:
    reg = tmp_path / "REGION"
    reg.mkdir()
    # format 2
    grid_f2: list[list[str]] = [[""] * 8 for _ in range(6)]
    grid_f2.append(["Состояние", "НГРИ", "", "", "Номер кредитного договора", "", "", "Раздел"])
    grid_f2.append(["На хранении", "77:01:000111:1-77/001/2020-1", "", "", "KD-1", "", "", "R1"])

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for r, row in enumerate(grid_f2, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(reg / "VL0099.xlsx")
    wb.close()

    # format 1 as SpreadsheetML
    xml = """<?xml version="1.0"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">
<Worksheet ss:Name="Sheet1"><Table>
<Row>
<Cell><Data ss:Type="String">DepoAccount</Data></Cell>
<Cell><Data ss:Type="String">DivisionName</Data></Cell>
<Cell><Data ss:Type="String">AgreementNumber</Data></Cell>
<Cell><Data ss:Type="String">RegistrationNumber</Data></Cell>
</Row>
<Row>
<Cell><Data ss:Type="String">VL0032</Data></Cell>
<Cell><Data ss:Type="String">4-02</Data></Cell>
<Cell><Data ss:Type="String">1573</Data></Cell>
<Cell><Data ss:Type="String">77:01:000111:1-77/001/2020-1</Data></Cell>
</Row>
</Table></Worksheet></Workbook>"""
    (reg / "[VL0032][][IS26040130930][][31.03.2026].xml").write_text(xml, encoding="utf-8")

    rows = iter_region_rows(reg)
    by_ngri = {r.ngri_v_depo: r for r in rows}
    assert "77:01:000111:1-77/001/2020-1" in by_ngri
    assert by_ngri["77:01:000111:1-77/001/2020-1"].status == "На хранении"


def test_iter_region_merges_vrsn_lkk_status(tmp_path: Path) -> None:
    reg = tmp_path / "REGION"
    lkk = reg / "ОтчетыВрСн из ЛКК"
    lkk.mkdir(parents=True)

    header = ["", "", "", "Кадастровый номер", "Номер КД", "Дата КД", "Номер ГРИ", "Дата ГРИ", "Статус"]
    data = ["", "", "", "50:01:001:1", "KD-1", "01.01.2024", "77:01:000111:1-77/001/2020-1", "01.02.2020", "В хранилище"]

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for r, row in enumerate([header, data], start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
    wb.save(lkk / "BpCH_VL0160_1.xlsx")
    wb.close()

    xml = """<?xml version="1.0" encoding="UTF-8"?>
<DEPOACCOUNT_BALANCE_REPORT>
  <ReportHeader><DepoAccount><DepoAccount>VL0160</DepoAccount></DepoAccount></ReportHeader>
  <ReportBody>
    <PositionList>
      <Position>
        <PositionInfo><Mortgage>
          <RegistrationNumber>77:01:000111:1-77/001/2020-1</RegistrationNumber>
          <AgreementNumber>KD-1</AgreementNumber>
        </Mortgage></PositionInfo>
        <DivisionList><Division><DivisionInfo>
          <DivisionName>4B02-60-00307-R-001P</DivisionName>
        </DivisionInfo></Division></DivisionList>
      </Position>
    </PositionList>
  </ReportBody>
</DEPOACCOUNT_BALANCE_REPORT>"""
    (reg / "[VL0160][][IS26050516750][][31.03.2026].xml").write_text(xml, encoding="utf-8")

    rows = iter_region_rows(reg)
    assert len(rows) == 1
    assert rows[0].status == "В хранилище"
