"""Тесты поиска дублей НГРИ / номера закладной в портфеле."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = PIPELINE_ROOT.parent
sys.path.insert(0, str(PIPELINE_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from models import PORTFOLIO_SHEET_DOM_IA, PORTFOLIO_SHEET_VTB, SVOD_IA_SHEETS, SvodRow
from references.spravochnik import Sprav1Lookup
from validate.accounts import VALIDATION_SHEET_NAME
from validate.duplicates import DUPLICATES_SHEET_NAME, find_workbook_duplicates
from write.excel_writer import _write_workbook

_EMPTY_SPRAV = Sprav1Lookup()


def test_find_workbook_duplicates_across_sheets() -> None:
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [
            SvodRow(n_zakladnoy_portfel="P-1", ngri_v_depo="NGRI-A"),
            SvodRow(n_zakladnoy_portfel="P-2", ngri_v_depo="NGRI-DUP"),
        ],
        PORTFOLIO_SHEET_VTB: [
            SvodRow(n_zakladnoy_portfel="P-1", ngri_v_depo="NGRI-B"),
            SvodRow(n_zakladnoy_portfel="P-3", ngri_v_depo="NGRI-DUP"),
        ],
    }
    dups = find_workbook_duplicates(rows_by_sheet, SVOD_IA_SHEETS)
    assert dups.ngri == ("NGRI-DUP",)
    assert dups.portfolio == ("P-1",)


def test_find_workbook_duplicates_ignores_empty() -> None:
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [
            SvodRow(n_zakladnoy_portfel=None, ngri_v_depo=""),
            SvodRow(n_zakladnoy_portfel="  ", ngri_v_depo="NGRI-1"),
        ],
    }
    dups = find_workbook_duplicates(rows_by_sheet, SVOD_IA_SHEETS)
    assert dups.ngri == ()
    assert dups.portfolio == ()


def test_write_workbook_adds_dubli_sheet(tmp_path: Path) -> None:
    out = tmp_path / "svod.xlsx"
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [
            SvodRow(n_zakladnoy_portfel="P-1", ngri_v_depo="NGRI-X"),
            SvodRow(n_zakladnoy_portfel="P-1", ngri_v_depo="NGRI-Y"),
        ],
    }
    dups = _write_workbook(out, rows_by_sheet, (PORTFOLIO_SHEET_DOM_IA,), "31.03.2026", _EMPTY_SPRAV)
    assert dups.portfolio == ("P-1",)

    wb = load_workbook(out)
    assert DUPLICATES_SHEET_NAME in wb.sheetnames
    assert wb.sheetnames[-1] == DUPLICATES_SHEET_NAME
    assert wb.sheetnames[-2] == VALIDATION_SHEET_NAME
    ws = wb[DUPLICATES_SHEET_NAME]
    assert ws.cell(1, 1).value == "НГРИ в депо"
    assert ws.cell(1, 2).value == "Номер закладной в портфеле"
    assert ws.cell(2, 2).value == "P-1"
    wb.close()


def test_write_workbook_skips_dubli_sheet_when_clean(tmp_path: Path) -> None:
    out = tmp_path / "svod.xlsx"
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [
            SvodRow(n_zakladnoy_portfel="P-1", ngri_v_depo="NGRI-1"),
            SvodRow(n_zakladnoy_portfel="P-2", ngri_v_depo="NGRI-2"),
        ],
    }
    dups = _write_workbook(out, rows_by_sheet, (PORTFOLIO_SHEET_DOM_IA,), "31.03.2026", _EMPTY_SPRAV)
    assert not dups.has_any

    wb = load_workbook(out)
    assert DUPLICATES_SHEET_NAME not in wb.sheetnames
    assert VALIDATION_SHEET_NAME in wb.sheetnames
    wb.close()


def test_write_workbook_default_zoom_70(tmp_path: Path) -> None:
    out = tmp_path / "svod.xlsx"
    rows_by_sheet = {PORTFOLIO_SHEET_DOM_IA: [SvodRow(schet_depo="123", ngri_v_depo="NGRI-1")]}
    _write_workbook(out, rows_by_sheet, (PORTFOLIO_SHEET_DOM_IA,), "31.03.2026", _EMPTY_SPRAV)

    wb = load_workbook(out)
    assert wb[PORTFOLIO_SHEET_DOM_IA].sheet_view.zoomScale == 70
    wb.close()
