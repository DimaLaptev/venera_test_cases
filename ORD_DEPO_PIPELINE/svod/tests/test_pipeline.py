"""Smoke-тесты SVOD DEPO pipeline."""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = PIPELINE_ROOT.parent
sys.path.insert(0, str(PIPELINE_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from assemble.builder import build_all_rows
from date_format import format_date_ddmmyyyy, report_month_last_day
from extract.gpb_r_msg import parse_gpb_r_msg
from extract.rsd_msg import parse_rsd_r_msg
from extract.rsd_r05 import parse_r05_statement_xlsx
from extract.rsd_section import extract_section_prefix
from extract.vtb import _parse_vtb_sheet
from models import ALL_PORTFOLIO_SHEETS, SVOD_HEADERS
from references.spravochnik import Sprav1Lookup, Sprav2Lookup
from references.vtb_codifier import lookup_vtb_section


GPB_SAMPLE = """ClientID:70
DepoNum:102
InstrNum:1650183
ClientName:ПАО ДОМ.РФ
ReportType:REMAINDERS
ReportName:R_102_20260331_00070.MSG
ReportDate:31.03.2026

102;10934;1;3;1;98040.000;102080004;ZKL102080004;26-26-28/005/2010-124;На хранении
"""

RSD_SAMPLE = """ClientID:000300008201
InstrNum:040515
ClientName:ПАО ДОМ.РФ
ReportType:R
ReportDate:31.03.2026
ReportName:R_260331_000300008201.MSG

ПАО ДОМ.РФ;000300008201;100014;3477587;1;53-53-11/083/2011-061
"""


def _make_r05_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A14"] = "ВЫПИСКА ПО СЧЕТУ ДЕПО"
    ws["A20"] = "Счет депо:"
    ws["B20"] = "000300043501"
    ws["A24"] = "Наименование места хранения"
    ws["B24"] = "Номер раздела счета депo"
    ws["E24"] = "Гос. рег. № выпуска"
    ws["A25"] = 'ООО "РСД"'
    ws["B25"] = "260000 - Вне обращения (блокированы)"
    ws["E25"] = "50-50/001-50/062/001/2015-1947/1"
    wb.save(path)
    wb.close()


def _make_spravochnik_xlsx(path: Path, account: str, portfolio: str) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Таблица сопоставления 1"
    ws1.append(["", "Депозитарий", "Номер счета депo"])
    ws1.append(["", "Банк ГПБ АО", account, "", "", "", "", "", "", "Владелец", portfolio])
    ws2 = wb.create_sheet("Таблица сопоставления 2")
    ws2.append(["Депозитарий", "Номер раздела (текст)", "Город - место хранения"])
    ws2.append(["", "260000", "Москва"])
    ws2.append(["", "1", "Санкт-Петербург"])
    ws2.append(['АО "ДК РЕГИОН"', "", "МОСКВА"])
    ws2.append(['АО "ВТБ СД"', "", "МОСКВА"])
    wb.save(path)
    wb.close()


def _make_sprav2_xlsx(path: Path) -> None:
    _make_spravochnik_xlsx(path, "000000000000", "ДОМ ИА")


def test_parse_gpb_r_msg(tmp_path: Path) -> None:
    p = tmp_path / "R_102_20260331_00070.MSG"
    p.write_text(GPB_SAMPLE, encoding="utf-8")
    rows = parse_gpb_r_msg(p)
    assert len(rows) == 1
    r = rows[0]
    assert r.schet_depo == "10934"
    assert r.razdel_scheta == "1"
    assert r.ngri_v_depo == "26-26-28/005/2010-124"
    assert r.status == "На хранении"


def test_parse_rsd_r_msg(tmp_path: Path) -> None:
    p = tmp_path / "R_260331_000300008201.MSG"
    p.write_text(RSD_SAMPLE, encoding="utf-8")
    rows = parse_rsd_r_msg(p)
    assert len(rows) == 1
    r = rows[0]
    assert r.schet_depo == "000300008201"
    assert r.razdel_scheta == "100014"
    assert r.ngri_v_depo == "53-53-11/083/2011-061"
    assert r.status == "На хранении"


def test_extract_section_prefix() -> None:
    assert extract_section_prefix("260000 - Вне обращения (блокированы)") == "260000"
    assert extract_section_prefix("100014 - Основной") == "100014"


def test_parse_r05_statement(tmp_path: Path) -> None:
    p = tmp_path / "R-05_test.xlsx"
    _make_r05_xlsx(p)
    rows = parse_r05_statement_xlsx(p)
    assert len(rows) == 1
    r = rows[0]
    assert r.schet_depo == "000300043501"
    assert r.razdel_scheta == "260000"
    assert r.ngri_v_depo == "50-50/001-50/062/001/2015-1947/1"
    assert r.status == "На хранении"


def test_parse_r05_stops_at_empty_rows_and_footer(tmp_path: Path) -> None:
    p = tmp_path / "R-05_footer.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A14"] = "ВЫПИСКА ПО СЧЕТУ ДЕПО"
    ws["A20"] = "Счет депо:"
    ws["B20"] = "000300043501"
    ws["A24"] = "Наименование места хранения"
    ws["B24"] = "Номер раздела счета депo"
    ws["E24"] = "Гос. рег. № выпуска"
    ws["A25"] = 'ООО "РСД"'
    ws["B25"] = "260000 - Вне обращения (блокированы)"
    ws["E25"] = "50-50/001-50/062/001/2015-1947/1"
    ws["A26"] = 'ПАО Сбербанк'
    ws["B26"] = "100008 - Основной"
    ws["E26"] = "24-24/012-24/012/003/2015-4367/1"
    # пустые строки — конец таблицы
    # row 27-28 empty
    ws["A29"] = "Исполнитель: Анохин Антон Андреевич"
    ws["E29"] = "Ответственное лицо: _____________________"
    wb.save(p)
    wb.close()

    rows = parse_r05_statement_xlsx(p)
    assert len(rows) == 2
    assert rows[0].ngri_v_depo == "50-50/001-50/062/001/2015-1947/1"
    assert rows[1].ngri_v_depo == "24-24/012-24/012/003/2015-4367/1"
    assert all("Ответственное" not in (r.ngri_v_depo or "") for r in rows)


def test_sprav2_lookup_by_section(tmp_path: Path) -> None:
    p = tmp_path / "sprav.xlsx"
    _make_spravochnik_xlsx(p, "000000000000", "ДОМ ИА")
    lookup = Sprav2Lookup.load(p)
    assert lookup.get("260000") == "Москва"
    assert lookup.get("1") == "Санкт-Петербург"
    assert lookup.get("999") is None
    assert lookup.get_by_depository_substring("регион") == "МОСКВА"
    assert lookup.get_by_depository_substring("втб") == "МОСКВА"


def test_format_date_ddmmyyyy() -> None:
    assert format_date_ddmmyyyy(date(2026, 3, 31)) == "31.03.2026"
    assert format_date_ddmmyyyy("31.3.2026") == "31.03.2026"
    assert format_date_ddmmyyyy("2026-03-31") == "31.03.2026"
    assert format_date_ddmmyyyy(None) is None


def test_report_month_last_day() -> None:
    assert report_month_last_day("15.03.2026") == "31.03.2026"
    assert report_month_last_day("31.03.2026") == "31.03.2026"
    assert report_month_last_day("2026-02-10") == "28.02.2026"


def test_vtb_razdel_and_kd_from_cols() -> None:
    sheet = [[""] * 12 for _ in range(14)]
    sheet[10][5] = "\u0417\u0430\u043b\u043e\u0433\u043e\u0434\u0430\u0442\u0435\u043b\u044c"
    sheet[5][6] = "12666"
    sheet[12][1] = "section account"
    sheet[12][2] = "1"
    sheet[12][3] = "[00000] \u0412\u043d\u0435 \u0445\u0440\u0430\u043d\u0438\u043b\u0438\u0449\u0430."
    sheet[12][8] = "50-50/001/2015-1"
    sheet[12][10] = "KD-1573"
    rows = _parse_vtb_sheet(sheet, "test.xlsx")
    assert len(rows) == 1
    assert rows[0].razdel_scheta == "[00000] \u0412\u043d\u0435 \u0445\u0440\u0430\u043d\u0438\u043b\u0438\u0449\u0430."
    assert rows[0].nomer_kd == "KD-1573"


def test_vtb_codifier_electronic() -> None:
    r = lookup_vtb_section(
        "Свободное обращение/АО ВТБ Специализированный депозитарий "
        "(Электронное хранилище) (Тип раздела:Свободное обращение)",
    )
    assert r.enc == "ENC3"
    assert r.vid_ucheta == "Депозитарный учет ЭЗ"


def test_svod_headers_no_empty_col() -> None:
    assert len(SVOD_HEADERS) == 26
    assert "" not in SVOD_HEADERS


def test_rsd_dedupe_by_key() -> None:
    from extract.rsd_dedupe import dedupe_rsd_do_rows
    from models import DepoReportRow, SourceKind

    base = DepoReportRow(
        source_kind=SourceKind.RSD_R05,
        source_file="a.xlsx",
        schet_depo="000300043501",
        razdel_scheta="260000",
        ngri_v_depo="50-50/001/2015-1",
    )
    dup = DepoReportRow(
        source_kind=SourceKind.RSD_R05,
        source_file="a.xlsx",
        schet_depo="000300043501",
        razdel_scheta="260000",
        ngri_v_depo="50-50/001/2015-1",
    )
    other_file = DepoReportRow(
        source_kind=SourceKind.RSD_R05,
        source_file="b.xlsx",
        schet_depo="000300043501",
        razdel_scheta="260000",
        ngri_v_depo="50-50/001/2015-1",
    )
    gpb = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="r.MSG",
        schet_depo="10934",
        razdel_scheta="1",
        ngri_v_depo="50-50/001/2015-1",
    )
    rows = dedupe_rsd_do_rows([base, dup, other_file, gpb, dup])
    assert len(rows) == 3
    assert rows[0].source_kind == SourceKind.RSD_R05
    assert rows[1].source_kind == SourceKind.RSD_R05
    assert rows[1].source_file == "b.xlsx"
    assert rows[2].source_kind == SourceKind.GPB


def test_depo_prior_routing_by_portfolio(tmp_path: Path) -> None:
    from openpyxl import Workbook

    from models import DepoReportRow, SourceKind
    from references.depo_dom import DepoDomIndex

    def _make_prior(path: Path, sheet: str, ngri: str, portfolio_no: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        for r in range(1, 4):
            ws.append([""] * 24)
        ws.append([portfolio_no, "dom", "", "", "", "", "", "", "fio", "kd", ngri])
        wb.save(path)

    ia_path = tmp_path / "ia.xlsx"
    dom_path = tmp_path / "dom.xlsx"
    _make_prior(ia_path, "GPB", "NGRI-IA", "100")
    _make_prior(dom_path, "GPB", "NGRI-DOM", "200")

    depo_ia = DepoDomIndex.load(ia_path, ("GPB",))
    depo_dom = DepoDomIndex.load(dom_path, ("GPB",))

    sprav_path = tmp_path / "sprav.xlsx"
    _make_spravochnik_xlsx(sprav_path, "10934", "ДОМ ИА")
    sprav1 = Sprav1Lookup.load(sprav_path)
    sprav2 = Sprav2Lookup.load(sprav_path)

    do = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="t.MSG",
        schet_depo="10934",
        razdel_scheta="1",
        ngri_v_depo="NGRI-IA",
    )
    rows_ia = build_all_rows([do], depo_ia, depo_dom, sprav1, sprav2)
    assert rows_ia["ДОМ ИА"][0].n_zakladnoy_portfel == "100"

    _make_spravochnik_xlsx(sprav_path, "10935", "ДОМ_Оригинатор ДОМ_РезСервис")
    sprav1 = Sprav1Lookup.load(sprav_path)
    do_dom = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="t.MSG",
        schet_depo="10935",
        razdel_scheta="1",
        ngri_v_depo="NGRI-DOM",
    )
    rows_dom = build_all_rows([do_dom], depo_ia, depo_dom, sprav1, sprav2)
    assert rows_dom["ДОМ_Оригинатор ДОМ_РезСервис"][0].n_zakladnoy_portfel == "200"


def _make_svod_prior_xlsx(
    path: Path,
    *,
    sheet: str,
    ngri: str,
    portfolio_no: str,
    dom_no: str = "dom",
    depository: str = "Банк ГПБ АО",
    istochnik: str = "R_102_20260331_00070.MSG",
) -> None:
    """Мини-книга в формате СВОД_поДЕПО: шапка строка 5, данные с 6."""
    from models import SVOD_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws["A1"] = "31.03.2026"
    for c, title in enumerate(SVOD_HEADERS, start=1):
        ws.cell(row=5, column=c, value=title)
    row = [
        portfolio_no,
        dom_no,
        depository,
        "",
        "",
        "",
        "",
        "1",
        "fio",
        "kd",
        ngri,
    ]
    for c, val in enumerate(row, start=1):
        ws.cell(row=6, column=c, value=val)
    ws.cell(row=6, column=25, value=istochnik)
    wb.save(path)
    wb.close()


def test_depo_prior_reads_rows_with_empty_portfolio_col(tmp_path: Path) -> None:
    """Строки с пустым A, но заполненным K (НГРИ), должны попадать в индекс."""
    from openpyxl import Workbook

    from models import DEPO_IA_SHEETS
    from references.depo_dom import DepoDomIndex

    path = tmp_path / "ia.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "GPB"
    for _ in range(3):
        ws.append([""] * 24)
    # A пусто, K = НГРИ, B = номер ДОМа
    row = [""] * 24
    row[1] = "DOM-99"  # B
    row[10] = "NGRI-EMPTY-A"  # K
    ws.append(row)
    wb.save(path)
    wb.close()

    idx = DepoDomIndex.load(path, DEPO_IA_SHEETS, layout="predsvod")
    rec = idx.by_ngri("GPB", "NGRI-EMPTY-A")
    assert rec is not None
    assert rec.dom_no == "DOM-99"


def test_depo_prior_auto_detects_svod_layout(tmp_path: Path) -> None:
    from models import DEPO_IA_SHEETS, DepoReportRow, SourceKind
    from references.depo_dom import DepoDomIndex

    ia_path = tmp_path / "СВОД_поДЕПО.xlsx"
    _make_svod_prior_xlsx(
        ia_path,
        sheet="ДОМ ИА",
        ngri="NGRI-AUTO",
        portfolio_no="500",
        dom_no="501",
    )
    # layout=auto / без --prior-from-svod
    depo_ia = DepoDomIndex.load(ia_path, DEPO_IA_SHEETS, layout="auto")
    rec = depo_ia.by_ngri("GPB", "NGRI-AUTO")
    assert rec is not None
    assert rec.portfolio_no == "500"
    assert rec.dom_no == "501"

    sprav_path = tmp_path / "sprav.xlsx"
    _make_spravochnik_xlsx(sprav_path, "10934", "ДОМ ИА")
    sprav1 = Sprav1Lookup.load(sprav_path)
    sprav2 = Sprav2Lookup.load(sprav_path)
    do = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="t.MSG",
        schet_depo="10934",
        razdel_scheta="1",
        ngri_v_depo="NGRI-AUTO",
    )
    rows = build_all_rows([do], depo_ia, DepoDomIndex(), sprav1, sprav2)
    assert rows["ДОМ ИА"][0].n_zakladnoy_portfel == "500"
    assert rows["ДОМ ИА"][0].n_zakladnoy_dom == "501"


def test_depo_prior_from_svod_layout(tmp_path: Path) -> None:
    from models import DEPO_DOM_SHEETS, DEPO_IA_SHEETS, DepoReportRow, SourceKind
    from references.depo_dom import DepoDomIndex, route_svod_row_to_sheet

    assert (
        route_svod_row_to_sheet(
            "Банк ГПБ АО",
            "R_102_20260331_00070.MSG",
            DEPO_IA_SHEETS,
        )
        == "GPB"
    )
    assert (
        route_svod_row_to_sheet(
            'АО "ВТБ СД"',
            "statement.xlsx",
            DEPO_IA_SHEETS,
        )
        == "VTBSD"
    )
    assert (
        route_svod_row_to_sheet(
            'ООО "РСД"',
            "R_260331_000300008201.MSG",
            DEPO_DOM_SHEETS,
        )
        == "RSD_MSG"
    )

    ia_path = tmp_path / "svod_ia.xlsx"
    dom_path = tmp_path / "svod_dom.xlsx"
    _make_svod_prior_xlsx(
        ia_path,
        sheet="ДОМ ИА",
        ngri="NGRI-SVOD-IA",
        portfolio_no="300",
        dom_no="301",
    )
    _make_svod_prior_xlsx(
        dom_path,
        sheet="ДОМ_Оригинатор ДОМ_РезСервис",
        ngri="NGRI-SVOD-DOM",
        portfolio_no="400",
        dom_no="401",
    )

    depo_ia = DepoDomIndex.load(ia_path, DEPO_IA_SHEETS, layout="svod")
    depo_dom = DepoDomIndex.load(dom_path, DEPO_DOM_SHEETS, layout="svod")

    rec = depo_ia.by_ngri("GPB", "NGRI-SVOD-IA")
    assert rec is not None
    assert rec.portfolio_no == "300"
    assert rec.dom_no == "301"

    sprav_path = tmp_path / "sprav.xlsx"
    _make_spravochnik_xlsx(sprav_path, "10934", "ДОМ ИА")
    sprav1 = Sprav1Lookup.load(sprav_path)
    sprav2 = Sprav2Lookup.load(sprav_path)

    do = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="t.MSG",
        schet_depo="10934",
        razdel_scheta="1",
        ngri_v_depo="NGRI-SVOD-IA",
    )
    rows_ia = build_all_rows([do], depo_ia, depo_dom, sprav1, sprav2)
    assert rows_ia["ДОМ ИА"][0].n_zakladnoy_portfel == "300"
    assert rows_ia["ДОМ ИА"][0].n_zakladnoy_dom == "301"

    _make_spravochnik_xlsx(sprav_path, "10935", "ДОМ_Оригинатор ДОМ_РезСервис")
    sprav1 = Sprav1Lookup.load(sprav_path)
    do_dom = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="t.MSG",
        schet_depo="10935",
        razdel_scheta="1",
        ngri_v_depo="NGRI-SVOD-DOM",
    )
    rows_dom = build_all_rows([do_dom], depo_ia, depo_dom, sprav1, sprav2)
    assert rows_dom["ДОМ_Оригинатор ДОМ_РезСервис"][0].n_zakladnoy_portfel == "400"


def test_build_all_rows_portfolio_routing(tmp_path: Path) -> None:
    from io import StringIO
    from models import DepoReportRow, SourceKind
    from references.depo_dom import DepoDomIndex

    sprav_path = tmp_path / "sprav.xlsx"
    _make_spravochnik_xlsx(sprav_path, "10934", "ДОМ ИА")

    sprav1 = Sprav1Lookup.load(sprav_path)
    sprav2 = Sprav2Lookup.load(sprav_path)

    do = DepoReportRow(
        source_kind=SourceKind.GPB,
        source_file="test.MSG",
        schet_depo="10934",
        razdel_scheta="1",
        ngri_v_depo="26-26-28/005/2010-124",
        status="На хранении",
    )

    warn = StringIO()
    rows = build_all_rows([do], DepoDomIndex(), DepoDomIndex(), sprav1, sprav2, warn=warn)
    assert len(rows["ДОМ ИА"]) == 1
    assert rows["ДОМ ИА"][0].gorod_hraneniya == "Санкт-Петербург"


def test_write_workbook_has_no_formulas(tmp_path: Path) -> None:
    import zipfile

    from models import SvodRow
    from references.spravochnik import Sprav1Lookup
    from write.excel_writer import _write_workbook

    out = tmp_path / "out.xlsx"
    row = SvodRow(
        schet_depo="123",
        ngri_v_depo="NGRI-1",
        nomer_kd="=0040-00767/ИКР-23РБ",
        comment="=1+1",
        mestonahozhdenie="=SUM(A1:A10)",
    )
    _write_workbook(
        out,
        {"ДОМ_Оригинатор ДОМ_РезСервис": [row]},
        ("ДОМ_Оригинатор ДОМ_РезСервис",),
        "31.03.2026",
        Sprav1Lookup(),
    )
    wb2 = load_workbook(out)
    ws2 = wb2["ДОМ_Оригинатор ДОМ_РезСервис"]
    assert ws2.cell(6, 5).value == "123"
    assert ws2.cell(6, 10).value == "=0040-00767/ИКР-23РБ"
    assert ws2.max_row == 6
    wb2.close()

    with zipfile.ZipFile(out) as zf:
        sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "<f" not in sheet_xml


def test_pipeline_e2e(tmp_path: Path) -> None:
    root = tmp_path / "SVOD_DEPO_PIPELINE"
    (root / "source" / "GPB").mkdir(parents=True)
    (root / "source" / "RSD_MSG").mkdir(parents=True)
    (root / "source" / "SPRAV").mkdir(parents=True)
    (root / "output").mkdir(parents=True)

    (root / "source" / "GPB" / "R_102_20260331_00070.MSG").write_text(
        GPB_SAMPLE, encoding="utf-8",
    )
    (root / "source" / "RSD_MSG" / "R_260331_000300008201.MSG").write_text(
        RSD_SAMPLE, encoding="utf-8",
    )
    _make_spravochnik_xlsx(root / "source" / "SPRAV" / "Справочник.xlsx", "10934", "ДОМ ИА")

    out = root / "output" / "СВОД_поДЕПО.xlsx"
    out_dom = root / "output" / "СВОД_поДЕПО_ДОМ.xlsx"
    import subprocess

    proc = subprocess.run(
        [
            sys.executable,
            str(PIPELINE_ROOT / "scripts" / "run_pipeline.py"),
            "--pipeline-root",
            str(root),
            "--output",
            str(out),
            "--output-dom",
            str(out_dom),
        ],
        cwd=str(PROJECT_ROOT),
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    assert out.is_file()
    assert out_dom.is_file()

    wb = load_workbook(out, read_only=True, data_only=True)
    assert "ДОМ ИА" in wb.sheetnames
    ws = wb["ДОМ ИА"]
    assert ws.cell(1, 1).value == "31.03.2026"
    assert ws.cell(1, 3).value == 1
    assert "СВОДНЫЙ ДЕПОЗИТАРНЫЙ ОТЧЕТ" in str(ws.cell(1, 6).value)
    assert len(SVOD_HEADERS) == ws.cell(5, 26).column
    wb.close()

    wb_dom = load_workbook(out_dom, read_only=True, data_only=True)
    assert "ДОМ_Оригинатор ДОМ_РезСервис" in wb_dom.sheetnames
    wb_dom.close()

    from openpyxl.styles import PatternFill

    wb_style = load_workbook(out)
    ws = wb_style["ДОМ ИА"]
    assert ws.cell(5, 1).fill.start_color.rgb in ("0070C0", "000070C0")
    assert ws.cell(6, 1).fill.start_color.rgb in ("DDEBF7", "00DDEBF7")
    wb_style.close()
