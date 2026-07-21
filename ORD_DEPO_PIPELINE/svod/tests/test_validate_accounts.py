"""Тесты листа «Валидация»."""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

PIPELINE_ROOT = Path(__file__).resolve().parents[1]
PROJECT_ROOT = PIPELINE_ROOT.parent
sys.path.insert(0, str(PIPELINE_ROOT / "src"))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from models import (
    PORTFOLIO_SHEET_DOM_IA,
    PORTFOLIO_SHEET_REZERVIS,
    SVOD_DOM_SHEETS,
    SVOD_IA_SHEETS,
    SvodRow,
)
from references.spravochnik import Sprav1Lookup
from validate.accounts import (
    VALIDATION_SHEET_NAME,
    build_validation_rows,
    count_zero_accounts,
)
from write.excel_writer import _write_workbook


def _make_spravochnik_multi_xlsx(path: Path, accounts: list[tuple[str, str]]) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Таблица сопоставления 1"
    ws1.append(["", "Депозитарий", "Номер счета депo"])
    for account, portfolio in accounts:
        ws1.append(["", "Банк ГПБ АО", account, "", "", "", "", "", "", "Владелец", portfolio])
    ws2 = wb.create_sheet("Таблица сопоставления 2")
    ws2.append(["Депозитарий", "Номер раздела (текст)", "Город - место хранения"])
    wb.save(path)
    wb.close()


def test_build_validation_rows_counts_per_workbook(tmp_path: Path) -> None:
    path = tmp_path / "sprav.xlsx"
    _make_spravochnik_multi_xlsx(
        path,
        [("10934", "ДОМ ИА"), ("88888", "ДОМ ИА"), ("VL0001", "ДОМ_Оригинатор ДОМ_РезСервис")],
    )
    sprav1 = Sprav1Lookup.load(path)
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [
            SvodRow(schet_depo="10934", ngri_v_depo="N1"),
            SvodRow(schet_depo="10934", ngri_v_depo="N2"),
        ],
    }
    ia_rows = build_validation_rows(sprav1, rows_by_sheet, SVOD_IA_SHEETS)
    by_account = {r.schet_depo: r for r in ia_rows}
    assert len(ia_rows) == 2
    assert by_account["10934"].count == 2
    assert by_account["88888"].count == 0
    assert count_zero_accounts(ia_rows) == 1

    dom_rows = build_validation_rows(sprav1, rows_by_sheet, SVOD_DOM_SHEETS)
    assert len(dom_rows) == 1
    assert dom_rows[0].schet_depo == "VL0001"
    assert dom_rows[0].razdel_kp == PORTFOLIO_SHEET_REZERVIS
    assert dom_rows[0].count == 0


def test_write_workbook_includes_validation_sheet(tmp_path: Path) -> None:
    path = tmp_path / "sprav.xlsx"
    _make_spravochnik_multi_xlsx(path, [("10934", "ДОМ ИА"), ("88888", "ДОМ ИА")])
    sprav1 = Sprav1Lookup.load(path)

    out = tmp_path / "svod.xlsx"
    rows_by_sheet = {
        PORTFOLIO_SHEET_DOM_IA: [SvodRow(schet_depo="10934", ngri_v_depo="N1")],
    }
    _write_workbook(out, rows_by_sheet, (PORTFOLIO_SHEET_DOM_IA,), "31.03.2026", sprav1)

    wb = load_workbook(out)
    assert VALIDATION_SHEET_NAME in wb.sheetnames
    ws = wb[VALIDATION_SHEET_NAME]
    assert ws.cell(1, 1).value == "Раздел КП"
    assert ws.cell(1, 4).value == "Количество в отчет"
    assert ws.cell(2, 3).value == "10934"
    assert ws.cell(2, 4).value == 1
    assert ws.cell(3, 3).value == "88888"
    assert ws.cell(3, 4).value == 0
    assert ws.cell(1, 1).fill.start_color.rgb in ("00B050", "0000B050")
    wb.close()


def test_pipeline_completes_with_zero_count_accounts(tmp_path: Path) -> None:
    import subprocess

    GPB_SAMPLE = """ClientID:70
DepoNum:102
InstrNum:1650183
ClientName:ПАО ДОМ.РФ
ReportType:REMAINDERS
ReportName:R_102_20260331_00070.MSG
ReportDate:31.03.2026

102;10934;1;3;1;98040.000;102080004;ZKL102080004;26-26-28/005/2010-124;На хранении
"""

    root = tmp_path / "SVOD_DEPO_PIPELINE"
    (root / "source" / "GPB").mkdir(parents=True)
    (root / "source" / "SPRAV").mkdir(parents=True)
    (root / "output").mkdir(parents=True)

    (root / "source" / "GPB" / "R_102_20260331_00070.MSG").write_text(
        GPB_SAMPLE,
        encoding="utf-8",
    )
    _make_spravochnik_multi_xlsx(
        root / "source" / "SPRAV" / "Справочник.xlsx",
        [("10934", "ДОМ ИА"), ("88888", "ДОМ ИА")],
    )

    out = root / "output" / "СВОД_пoДЕПО.xlsx"
    out_dom = root / "output" / "СВОД_пoДEPO_ДOM.xlsx"

    proc = subprocess.run(
        [
            sys.executable,
            str(PIPELINE_ROOT / "scripts" / "run_pipeline.py"),
            "--pipeline-root",
            str(root),
            "--output",
            str(out),
            "--output-dom",
            str(out_dom),
        ],
        cwd=str(PROJECT_ROOT),
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr + proc.stdout
    assert out.is_file()
    assert "88888" not in proc.stdout or "количеством 0" in proc.stdout

    wb = load_workbook(out, read_only=True, data_only=True)
    assert VALIDATION_SHEET_NAME in wb.sheetnames
    wb.close()
