#!/usr/bin/env python3
"""Заполнение листа «СВОД_операций» из I*.MSG, R-01 xlsx, REGION IS*.xls, ВТБ СД F752 XML+ZIP, D*.MSG."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parent.parent
_SRC = _ROOT / "src"
if str(_SRC) not in sys.path:
    sys.path.insert(0, str(_SRC))

from codifier_table import read_codifier_rows, read_vtb_sd_codifier_rows
from depo_directory import read_depo_directory_rows
from gpb_d_msg import parse_d_msg_file
from period_cli import add_period_dir_argument, apply_period_workbook
from gpb_i_msg import (
    gpb_instruction_pair_key,
    is_gpb_i_row,
    normalize_ngri_for_match,
    parse_i_msg_file,
    passes_execution_filter,
)
from openpyxl import Workbook, load_workbook
from rsd_i_msg import (
    is_rsd_i_row,
    is_rsd_technical_row,
    rsd_execution_code,
)
from region_is_xls import (
    REGION_PARSE_SKIP_BAD_NAME,
    REGION_PARSE_SKIP_UNREADABLE,
    RegionTableRow,
    parse_region_is_xls,
)
from rsd_r01_xlsx import parse_r01_xlsx
from svod_excel import (
    ERRORS_SHEET_HEADERS,
    SVOD_HEADERS,
    build_gpb_i_d_match_index,
    build_svod_row,
    build_svod_row_from_d,
    build_svod_row_r01,
    build_svod_row_region,
    build_svod_row_rsd,
    build_svod_row_vtb,
    create_workbook_with_svod_only,
    ensure_errors_sheet,
    ensure_svod_sheet,
    select_d_msg_rows_for_svod,
    write_svod_rows,
)
from vtb_f752_xml import (
    find_vtb_xls_for_xml,
    find_vtb_zip_for_xml,
    list_ngri_from_vtb_xls_path,
    list_ngri_from_vtb_zip,
    parse_vtb_f752_xml,
)

from dks_excel import SCHEETA_DEPO_HEADERS


# Строки листа «Коды» для СВОД из D*.MSG (тип поручения DF/DP, пустой код операции)
_CODE_DOC_DF = [
    "",
    "",
    "Документы - зачисление",
    "Документы - зачисление",
    "",
    "",
    "",
    "DF",
    "только для 10934",
]
_CODE_DOC_DP = [
    "",
    "",
    "Документы - списание",
    "Документы - списание",
    "",
    "",
    "",
    "DP",
    "",
]


def _codifier_has_document_df_dp_rows(workbook_path: str) -> bool:
    for r in read_codifier_rows(workbook_path):
        if r.op_code.strip():
            continue
        if r.instruction_type.strip().upper() not in ("DF", "DP"):
            continue
        if "документ" in r.op_type.lower():
            return True
    return False


def append_document_codifier_rows_if_missing(path: Path) -> None:
    """Добавляет строки «Документы - …» / DF|DP, если в «Коды» их ещё нет."""
    if _codifier_has_document_df_dp_rows(str(path)):
        return
    wb = load_workbook(path)
    if "Коды" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["Коды"]
    ws.append(_CODE_DOC_DF)
    ws.append(_CODE_DOC_DP)
    wb.save(path)
    wb.close()


def _depo_demo_row() -> list:
    return [
        "ДОМ",
        "Банк ГПБ АО",
        33563,
        "—",
        "БДРФ",
        "—",
        "залоговый",
        248,
        "",
        "ПАО ДОМ.РФ",
    ]


def append_reference_sheets_if_missing(path: Path) -> None:
    """Добавляет «Счета депо» и «Коды» с тестовыми строками, если листов нет или только заголовки (например после fill_dks)."""
    wb = load_workbook(path)
    changed = False
    if "Счета депо" not in wb.sheetnames:
        ws_dep = wb.create_sheet("Счета депо")
        for col, title in enumerate(SCHEETA_DEPO_HEADERS, start=1):
            ws_dep.cell(row=1, column=col, value=title)
        ws_dep.append(_depo_demo_row())
        changed = True
    else:
        ws_dep = wb["Счета депо"]
        has_data = ws_dep.max_row > 1 and ws_dep.cell(row=2, column=3).value not in (None, "")
        if not has_data:
            ws_dep.append(_depo_demo_row())
            changed = True
    if "Коды" not in wb.sheetnames:
        ws_cod = wb.create_sheet("Коды")
        code_headers = [
            "Код операции",
            "Код исполнения",
            "Тип операции",
            "Содержание ГПБ",
            "Содержание РСД",
            "Содержание РЕГИОН",
            "Содержание ВТБСД",
            "Тип поручения",
            "Комментарий",
        ]
        for col, title in enumerate(code_headers, start=1):
            ws_cod.cell(row=1, column=col, value=title)
        ws_cod.append(
            [
                12005,
                "",
                "Прекращение залога - зачисление",
                "Зачисление ЦБ - прекращение залога",
                "",
                "",
                "",
                "VF",
                "",
            ]
        )
        ws_cod.append(
            [
                12005,
                "",
                "Прекращение залога - списание",
                "Списание ЦБ - прекращение залога",
                "",
                "",
                "",
                "VP",
                "",
            ]
        )
        ws_cod.append(_CODE_DOC_DF)
        ws_cod.append(_CODE_DOC_DP)
        changed = True
    else:
        ws_cod = wb["Коды"]
        has_codes = ws_cod.max_row > 1 and ws_cod.cell(row=2, column=1).value not in (None, "")
        if not has_codes:
            ws_cod.append(
                [
                    12005,
                    "",
                    "Прекращение залога - зачисление",
                    "Зачисление ЦБ - прекращение залога",
                    "",
                    "",
                    "",
                    "VF",
                    "",
                ]
            )
            ws_cod.append(
                [
                    12005,
                    "",
                    "Прекращение залога - списание",
                    "Списание ЦБ - прекращение залога",
                    "",
                    "",
                    "",
                    "VP",
                    "",
                ]
            )
            ws_cod.append(_CODE_DOC_DF)
            ws_cod.append(_CODE_DOC_DP)
            changed = True
    if changed:
        wb.save(path)
    wb.close()


def seed_minimal_reference_workbook(path: Path) -> None:
    """Листы «Счета депо» и «Коды» с заголовками и минимальными строками для теста (ClientID 248 → 33563)."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_dep = wb.create_sheet("Счета депо")
    for col, title in enumerate(SCHEETA_DEPO_HEADERS, start=1):
        ws_dep.cell(row=1, column=col, value=title)
    # строка как в memory/3_GBP_cods_1 (залоговый 33563, код ДОМ 248)
    ws_dep.append(
        [
            "ДОМ",
            "Банк ГПБ АО",
            33563,
            "—",
            "БДРФ",
            "—",
            "залоговый",
            248,
            "",
            "ПАО ДОМ.РФ",
        ]
    )

    ws_cod = wb.create_sheet("Коды")
    code_headers = [
        "Код операции",
        "Код исполнения",
        "Тип операции",
        "Содержание ГПБ",
        "Содержание РСД",
        "Содержание РЕГИОН",
        "Содержание ВТБСД",
        "Тип поручения",
        "Комментарий",
    ]
    for col, title in enumerate(code_headers, start=1):
        ws_cod.cell(row=1, column=col, value=title)
    ws_cod.append(
        [
            12005,
            "",
            "Прекращение залога - зачисление",
            "Зачисление ЦБ - прекращение залога",
            "",
            "",
            "",
            "VF",
            "",
        ]
    )
    ws_cod.append(
        [
            12005,
            "",
            "Прекращение залога - списание",
            "Списание ЦБ - прекращение залога",
            "",
            "",
            "",
            "VP",
            "",
        ]
    )
    ws_cod.append(_CODE_DOC_DF)
    ws_cod.append(_CODE_DOC_DP)

    ws_svod = wb.create_sheet("СВОД_операций")
    for col, title in enumerate(SVOD_HEADERS, start=1):
        ws_svod.cell(row=1, column=col, value=title)

    ws_err = wb.create_sheet("Ошибки")
    for col, title in enumerate(ERRORS_SHEET_HEADERS, start=1):
        ws_err.cell(row=1, column=col, value=title)

    wb.save(path)


def ensure_workbook_structure(path: Path, create: bool, seed: bool) -> None:
    if not path.exists():
        if not create:
            print(
                f"Файл не найден: {path}. Укажите --create-workbook.",
                file=sys.stderr,
            )
            sys.exit(1)
        if seed:
            print(f"Создаю книгу с тестовыми справочниками: {path}")
            seed_minimal_reference_workbook(path)
        else:
            create_workbook_with_svod_only(path)
        return

    wb = load_workbook(path)
    changed = False
    if "СВОД_операций" not in wb.sheetnames:
        ensure_svod_sheet(wb)
        changed = True
    if "Ошибки" not in wb.sheetnames:
        ensure_errors_sheet(wb)
        changed = True
    if changed:
        wb.save(path)
    wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Парсинг I*.MSG, R-01 xlsx (РСД), REGION IS*.xls, ВТБ СД F752 XML+ZIP, D*.MSG; запись листа «СВОД_операций».",
    )
    add_period_dir_argument(parser)
    parser.add_argument(
        "--input-dir-rsd",
        type=Path,
        default=_ROOT / "2026_01" / "RSD" / "I",
        help="Каталог с I*.MSG депозитария РСД",
    )
    parser.add_argument(
        "--no-rsd-msg",
        action="store_true",
        help="Не обрабатывать I*.MSG из --input-dir-rsd",
    )
    parser.add_argument(
        "--input-dir-rsd-r01",
        type=Path,
        default=_ROOT / "2026_01" / "RSD" / "REP",
        help="Каталог с отчётами РСД R-01 (*.xlsx)",
    )
    parser.add_argument(
        "--no-rsd-r01",
        action="store_true",
        help="Не обрабатывать R-01*.xlsx из --input-dir-rsd-r01",
    )
    parser.add_argument(
        "--input-dir-region",
        type=Path,
        default=_ROOT / "2026_01" / "REGION" / "REP",
        help="Каталог с отчётами REGION IS*.xls / IS*.xlsx",
    )
    parser.add_argument(
        "--no-region-xls",
        action="store_true",
        help="Не обрабатывать IS*.xls из --input-dir-region",
    )
    parser.add_argument(
        "--input-dir-vtb-xml",
        type=Path,
        default=_ROOT / "2026_01" / "VTBSD" / "REP" / "XML",
        help="Каталог с отчётами ВТБ СД F752 (*.xml)",
    )
    parser.add_argument(
        "--input-dir-vtb-zip",
        type=Path,
        default=_ROOT / "2026_01" / "VTBSD" / "REP" / "XML_приложения",
        help="Каталог приложений: одноимённый *.zip (НГРИ в .xls внутри) или одноимённый *.xls, если архива нет",
    )
    parser.add_argument(
        "--no-vtb-xml",
        action="store_true",
        help="Не обрабатывать ВТБ СД *.xml",
    )
    parser.add_argument(
        "--no-vtb-zip",
        action="store_true",
        help="Не читать приложения для НГРИ (ни ZIP, ни одноимённый .xls из --input-dir-vtb-zip) — одна строка СВОД на XML с пустым НГРИ",
    )
    parser.add_argument(
        "--debug-vtb-ngri",
        action="store_true",
        help="Подробный вывод в stderr при извлечении НГРИ из ZIP ВТБ СД (архив, .xls/.xlsx, ошибки)",
    )
    parser.add_argument(
        "--input-dir-d",
        type=Path,
        default=_ROOT / "2026_01" / "GPB" / "D",
        help="Каталог с D*.MSG для строк СВОД (DOC_INSTRUCT)",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=_ROOT / "2026_01" / "2026_01_Ord_Quantity.xlsx",
        help="Путь к xlsx",
    )
    parser.add_argument(
        "--create-workbook",
        action="store_true",
        help="Создать файл, если отсутствует",
    )
    parser.add_argument(
        "--seed-demo-data",
        action="store_true",
        help="При создании книги добавить тестовые листы «Счета депо» и «Коды»",
    )
    parser.add_argument(
        "--ensure-reference-sheets",
        action="store_true",
        help="Если нет листов «Счета депо» или «Коды», создать их с демо-данными",
    )
    parser.add_argument(
        "--no-d-msg",
        action="store_true",
        help="Не добавлять строки из D*.MSG (только отчёт I)",
    )
    parser.add_argument(
        "--svod-d-only",
        action="store_true",
        help="Только строки из D*.MSG из --input-dir-d (I*.MSG из --input-dir для сопоставления DF/DP рекомендуется)",
    )
    parser.add_argument(
        "--include-empty-doc-name",
        action="store_true",
        help="Для D*.MSG не отбрасывать строки с пустым наименованием документа",
    )
    args = parser.parse_args()
    pp = apply_period_workbook(args, project_root=_ROOT)
    if pp is not None:
        args.input_dir = pp.gpb_i
        args.input_dir_rsd = pp.rsd_i
        args.input_dir_rsd_r01 = pp.rsd_rep
        args.input_dir_region = pp.region_rep
        args.input_dir_vtb_xml = pp.vtb_xml
        args.input_dir_vtb_zip = pp.vtb_zip
        args.input_dir_d = pp.gpb_d
    elif args.period_dir:
        args.input_dir = Path(args.period_dir).resolve()
    else:
        args.input_dir = (_ROOT / "2026_01" / "GPB" / "I").resolve()

    wb_path = args.workbook.resolve()
    ensure_workbook_structure(wb_path, args.create_workbook, args.seed_demo_data)
    if args.ensure_reference_sheets:
        append_reference_sheets_if_missing(wb_path)
    append_document_codifier_rows_if_missing(wb_path)

    depo_rows = read_depo_directory_rows(str(wb_path))
    cod_rows = read_codifier_rows(str(wb_path))
    vtb_sd_cod_rows = read_vtb_sd_codifier_rows(str(wb_path))
    if not depo_rows:
        print(
            "Предупреждение: лист «Счета депо» пуст или отсутствует — колонки B, E могут быть пустыми.",
            file=sys.stderr,
        )
    if not cod_rows:
        print(
            "Предупреждение: лист «Коды» пуст или отсутствует — колонки D, N могут быть пустыми.",
            file=sys.stderr,
        )
    if not vtb_sd_cod_rows:
        print(
            'Предупреждение: лист «Коды ВТБ СД» (или имя с «Коды» и «ВТБ») '
            "отсутствует или без строк данных — для ВТБ СД колонки C, D, N берутся из «Коды».",
            file=sys.stderr,
        )

    rows_out: list[list] = []

    error_rows: list[list] = []
    gpb_d_index: dict[tuple[str, str], str] = {}
    gpb_d_stmt_keys: set[str] = set()
    if not args.no_d_msg:
        gpb_i_for_d = args.input_dir.resolve()
        if gpb_i_for_d.is_dir():
            gpb_d_index, gpb_d_stmt_keys = build_gpb_i_d_match_index(
                gpb_i_for_d,
                depo_rows,
                cod_rows,
            )

    if not args.svod_d_only:
        i_dirs: list[Path] = [args.input_dir.resolve()]
        if not args.no_rsd_msg:
            i_dirs.append(args.input_dir_rsd.resolve())

        gpb_i_root = args.input_dir.resolve()

        seen_i: set[Path] = set()
        files_i: list[Path] = []
        i_msg_from_gpb_dir: dict[Path, bool] = {}
        for input_dir in i_dirs:
            if not input_dir.is_dir():
                print(
                    f"Каталог I не найден (пропуск): {input_dir}",
                    file=sys.stderr,
                )
                continue
            input_res = input_dir.resolve()
            from_gpb_catalog = input_res == gpb_i_root
            for pat in ("I*.MSG", "i*.MSG"):
                for f in sorted(input_dir.glob(pat)):
                    rp = f.resolve()
                    if rp not in seen_i:
                        seen_i.add(rp)
                        files_i.append(f)
                        i_msg_from_gpb_dir[rp] = from_gpb_catalog

        if not files_i:
            print("Нет файлов I*.MSG в указанных каталогах ГПБ/РСД")

        for fp in files_i:
            parsed = parse_i_msg_file(fp)
            client_id = (parsed.header.get("ClientID") or "").strip()
            from_gpb_catalog = i_msg_from_gpb_dir[fp.resolve()]

            for data in parsed.rows:
                parts = data.parts
                if from_gpb_catalog:
                    if not is_gpb_i_row(parts):
                        continue
                    if not passes_execution_filter(data):
                        continue
                    rows_out.append(
                        build_svod_row(
                            parsed,
                            data,
                            depo_rows,
                            cod_rows,
                            client_id,
                        ),
                    )
                elif is_rsd_i_row(parts):
                    if is_rsd_technical_row(parts):
                        continue
                    if not rsd_execution_code(parts):
                        continue
                    rows_out.append(
                        build_svod_row_rsd(
                            parsed,
                            data,
                            depo_rows,
                            cod_rows,
                            client_id,
                        ),
                    )

        if not args.no_rsd_r01:
            r01_dir = args.input_dir_rsd_r01.resolve()
            if not r01_dir.is_dir():
                print(
                    f"Каталог R-01 не найден (пропуск): {r01_dir}",
                    file=sys.stderr,
                )
            else:
                seen_r01: set[Path] = set()
                files_r01: list[Path] = []
                for pat in ("R-01*.xlsx", "R-01*.XLSX", "r-01*.xlsx"):
                    for f in sorted(r01_dir.glob(pat)):
                        rp = f.resolve()
                        if rp not in seen_r01:
                            seen_r01.add(rp)
                            files_r01.append(f)
                if not files_r01:
                    print(f"Нет файлов R-01*.xlsx в {r01_dir}")
                for fp in files_r01:
                    pr = parse_r01_xlsx(fp)
                    if pr is None:
                        print(
                            f"Не удалось разобрать R-01 (пропуск): {fp.name}",
                            file=sys.stderr,
                        )
                        continue
                    for dr in pr.rows:
                        row_svod = build_svod_row_r01(
                            pr,
                            dr,
                            depo_rows,
                            cod_rows,
                        )
                        if row_svod is not None:
                            rows_out.append(row_svod)

        if not args.no_region_xls:
            region_dir = args.input_dir_region.resolve()
            if not region_dir.is_dir():
                print(
                    f"Каталог REGION не найден (пропуск): {region_dir}",
                    file=sys.stderr,
                )
            else:
                seen_reg: set[Path] = set()
                files_reg: list[Path] = []
                for pat in (
                    "IS*.xls",
                    "IS*.XLS",
                    "IS*.xlsx",
                    "IS*.XLSX",
                ):
                    for f in sorted(region_dir.glob(pat)):
                        rp = f.resolve()
                        if rp not in seen_reg:
                            seen_reg.add(rp)
                            files_reg.append(f)
                if not files_reg:
                    print(f"Нет файлов IS*.xls в {region_dir}")
                for fp in files_reg:
                    pr, region_skip = parse_region_is_xls(fp)
                    if pr is None:
                        if region_skip == REGION_PARSE_SKIP_BAD_NAME:
                            msg = "Пропуск REGION (имя не IS*.xls):"
                        else:
                            msg = (
                                "Пропуск REGION (файл не прочитан, пустые листы "
                                "или неподдерживаемый формат):"
                            )
                        print(f"{msg} {fp.name}", file=sys.stderr)
                        continue
                    tbody = pr.table_rows
                    if not tbody:
                        tbody = [RegionTableRow(ngri="")]
                    for tr in tbody:
                        svod_r, region_err = build_svod_row_region(
                            pr,
                            tr,
                            depo_rows,
                            cod_rows,
                        )
                        if region_err is not None:
                            error_rows.append(region_err)
                        if svod_r is not None:
                            rows_out.append(svod_r)

        if not args.no_vtb_xml:
            vtb_xml_dir = args.input_dir_vtb_xml.resolve()
            vtb_zip_dir = args.input_dir_vtb_zip.resolve()
            if not vtb_xml_dir.is_dir():
                print(
                    f"Каталог ВТБ XML не найден (пропуск): {vtb_xml_dir}",
                    file=sys.stderr,
                )
            else:
                vtb_files: list[Path] = []
                for p in sorted(vtb_xml_dir.iterdir()):
                    if not p.is_file():
                        continue
                    if p.suffix.lower() != ".xml":
                        continue
                    vtb_files.append(p)
                if not vtb_files:
                    print(f"Нет файлов *.xml в {vtb_xml_dir}")
                dbg_z = args.debug_vtb_ngri
                if dbg_z and not args.no_vtb_zip and not vtb_zip_dir.is_dir():
                    print(
                        f"[VTB НГРИ] каталог приложений не найден (--input-dir-vtb-zip): {vtb_zip_dir}",
                        file=sys.stderr,
                    )
                for xp in vtb_files:
                    pr_vtb = parse_vtb_f752_xml(xp)
                    if pr_vtb is None:
                        print(
                            f"Не удалось разобрать ВТБ XML (пропуск): {xp.name}",
                            file=sys.stderr,
                        )
                        continue
                    ngri_list: list[str] = []
                    if not args.no_vtb_zip and vtb_zip_dir.is_dir():
                        zp = find_vtb_zip_for_xml(
                            xp,
                            vtb_zip_dir,
                            debug=dbg_z,
                            context=xp.name,
                        )
                        if zp is not None:
                            ngri_list = list_ngri_from_vtb_zip(
                                zp,
                                debug=dbg_z,
                                context=xp.name,
                            )
                        else:
                            xls_p = find_vtb_xls_for_xml(
                                xp,
                                vtb_zip_dir,
                                debug=dbg_z,
                                context=xp.name,
                            )
                            if xls_p is not None:
                                ngri_list = list_ngri_from_vtb_xls_path(
                                    xls_p,
                                    debug=dbg_z,
                                    context=xp.name,
                                )
                        if dbg_z and ngri_list:
                            print(
                                f"[VTB НГРИ] {xp.name}: итого НГРИ (ZIP или .xls): {len(ngri_list)}",
                                file=sys.stderr,
                            )
                    if not ngri_list:
                        ngri_list = [""]
                    vtb_file_err: list | None = None
                    vtb_svod_block: list[list] = []
                    for ng in ngri_list:
                        sv, err = build_svod_row_vtb(
                            pr_vtb,
                            ng,
                            depo_rows,
                            cod_rows,
                            vtb_sd_cod_rows,
                        )
                        if err is not None and sv is None:
                            error_rows.append(err)
                            vtb_file_err = err
                            break
                        if err is not None:
                            error_rows.append(err)
                        if sv is not None:
                            vtb_svod_block.append(sv)
                    if vtb_file_err is None:
                        rows_out.extend(vtb_svod_block)

    if not args.no_d_msg:
        input_d = args.input_dir_d.resolve()
        if not input_d.is_dir():
            print(f"Каталог D не найден: {input_d}", file=sys.stderr)
            if args.svod_d_only:
                return 1
        else:
            seen_d = set()
            files_d: list[Path] = []
            for pat in ("D*.MSG", "d*.MSG"):
                for f in sorted(input_d.glob(pat)):
                    rp = f.resolve()
                    if rp not in seen_d:
                        seen_d.add(rp)
                        files_d.append(f)
            if not files_d:
                print(f"Нет файлов D*.MSG в {input_d}")
            for fp in files_d:
                try:
                    parsed = parse_d_msg_file(fp)
                except OSError as e:
                    print(f"Ошибка чтения {fp}: {e}", file=sys.stderr)
                    continue
                rt = (parsed.header.get("ReportType") or "").strip()
                if rt and rt != "DOC_INSTRUCT":
                    print(f"Пропуск D (не DOC_INSTRUCT): {fp.name}", file=sys.stderr)
                    continue
                client_id = (parsed.header.get("ClientID") or "").strip()
                picked_rows = select_d_msg_rows_for_svod(
                    parsed.rows,
                    include_empty_doc_name=args.include_empty_doc_name,
                )
                for data in picked_rows:
                    stmt_key = gpb_instruction_pair_key(fp)
                    ng_key = normalize_ngri_for_match(data.ngri)
                    i_flow = None
                    if stmt_key is not None:
                        hit = gpb_d_index.get((stmt_key, ng_key))
                        if hit is not None:
                            i_flow = hit  # "debit" | "credit"
                    if i_flow is None:
                        if stmt_key is None:
                            reason = "key_unparsed"
                        elif stmt_key not in gpb_d_stmt_keys:
                            reason = "no_i_file"
                        else:
                            reason = "no_ngri_in_i"
                        error_rows.append(
                            [
                                fp.name,
                                stmt_key or "",
                                data.ngri.strip(),
                                reason,
                                data.doc_type.strip(),
                                (parsed.header.get("ReportDate") or "").strip(),
                            ],
                        )
                    rows_out.append(
                        build_svod_row_from_d(
                            parsed,
                            data,
                            depo_rows,
                            cod_rows,
                            client_id,
                            i_match_flow=i_flow,
                        ),
                    )

    write_svod_rows(
        wb_path,
        rows_out,
        error_rows,
    )
    print(f"Записано строк: {len(rows_out)} в {wb_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
